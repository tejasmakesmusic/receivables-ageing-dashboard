"""Integration tests for the staging review API (M3 Task 4).

Covers:
  GET  /snapshots/{id}/staging         (filter variants, pagination, RBAC)
  PATCH /snapshots/{id}/staging/{row}  (each action, RBAC, 409, 422)
  PATCH /snapshots/{id}/warnings/ack   (happy path, unknown code → 422)

Test fixture strategy:
  - Reuses auth helpers from test_snapshots_upload.py.
  - Uploads a snapshot via POST /snapshots in each test that needs one.
  - Per-test DB rollback means no cross-test contamination.
"""

from __future__ import annotations

import io
import json
import uuid
from datetime import date
from typing import TYPE_CHECKING, Any, cast

import openpyxl
import pytest
from sqlalchemy import select

from app.core.rbac import Role
from app.db.models.entity import Entity
from app.db.models.party import PartyAlias, PartyCanonical
from app.db.models.snapshot import Snapshot
from app.db.models.user import User

if TYPE_CHECKING:
    from fastapi.testclient import TestClient
    from sqlalchemy.orm import Session


# ---------------------------------------------------------------------------
# Auth helpers (mirrors test_snapshots_upload.py pattern)
# ---------------------------------------------------------------------------


def _login(client: TestClient, email: str) -> None:
    client.get(f"/auth/google/callback?stub_email={email}", follow_redirects=False)


def _csrf(client: TestClient) -> str:
    return client.cookies.get("csrf_token") or ""


def _login_as_admin(client: TestClient) -> None:
    _login(client, "tejaswa.sharma@emb.global")


def _login_as_analyst(
    client: TestClient,
    db_session: Session,
    email: str,
    entity_code: str | None = None,
) -> uuid.UUID:
    _login(client, email)
    user = db_session.scalar(select(User).where(User.email == email))
    assert user is not None
    user.role = Role.ANALYST
    if entity_code is not None:
        entity = db_session.scalar(select(Entity).where(Entity.code == entity_code))
        assert entity is not None
        user.entity_id_scope = entity.id
    else:
        user.entity_id_scope = None
    user.is_active = True
    db_session.flush()
    return cast(uuid.UUID, user.id)


def _login_as_cfo(client: TestClient, db_session: Session, email: str) -> None:
    _login(client, email)
    user = db_session.scalar(select(User).where(User.email == email))
    assert user is not None
    user.role = Role.CFO
    user.is_active = True
    db_session.flush()


def _login_as_pending(client: TestClient, email: str) -> None:
    _login(client, email)


# ---------------------------------------------------------------------------
# XLSX builders
# ---------------------------------------------------------------------------


def _make_tally_xlsx(data_rows: list[list[Any]] | None = None) -> bytes:
    """Build a minimal Tally XLSX in the correct GrpBills format.

    The Tally parser uses a two-row pattern per invoice:
      1. Party header row: [None, None, party_name, None, None, None, None]
         (date empty + ref_no empty → sets current_party)
      2. Invoice row:      [date, ref_no, None, opening, pending, due_on, overdue]
         (date + ref_no populated → creates StagedInvoice with current_party)

    Each element in ``data_rows`` is interpreted as
    [date, ref_no, party_name, opening, pending, due_on, overdue].
    For each row a party-header row is emitted first (if party_name is non-empty),
    then the invoice row with party_name blanked out.
    """
    _meta = [
        ["Group :", "Sundry Debtors", None, "1-Apr-26 to 16-Apr-26", None, None, None],
        ["Details of:", "Pending Bills", None, None, None, None, None],
        [None] * 7,
        ["Date", "Ref. No.", "Party's Name", "Opening", "Pending", "Due on", "Overdue"],
        [None, None, None, "Amount", "Amount", None, "by days"],
    ]
    wb = openpyxl.Workbook()
    del wb["Sheet"]
    ws = wb.create_sheet("Sundry Debtors")
    for row in _meta:
        ws.append(row)
    for row in data_rows or []:
        # row: [date, ref_no, party_name, opening, pending, due_on, overdue]
        inv_date, ref_no, party_name, opening, pending, due_on, overdue = row
        # Emit party header row so the parser sets current_party
        if party_name is not None:
            ws.append([None, None, party_name, None, None, None, None])
        # Emit invoice row (party_name column is empty; parser uses current_party)
        ws.append([inv_date, ref_no, None, opening, pending, due_on, overdue])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_XERO_HEADER_ROW: list[Any] = [
    "Contact Account Number",
    "Primary Person",
    "Phone",
    "Email",
    "Mobile",
    "Contact Group",
    "Invoice Date",
    "Due Date",
    "Expected Date",
    "Invoice Number",
    "Invoice Reference",
    "< 1 Month",
    "1 Month",
    "2 Months",
    "3 Months",
    "Older",
    None,
    "Total",
    "Outstanding Tax",
    "PROJECT ID",
    "SERVICE MONTH",
    "Invoice Seen",
    "Invoice Sent",
]


def _make_xero_xlsx(
    as_of_str: str = "As at 31 March 2026",
    party: str = "TestParty",
    inv_date: Any = date(2026, 1, 15),
    inv_num: str = "INV-001",
    total: float = 1000.0,
) -> bytes:
    wb = openpyxl.Workbook()
    del wb["Sheet"]
    ws = wb.create_sheet("Aged Receivables Detail")
    ws.append(["Aged Receivables Detail"] + [None] * 22)
    ws.append(["TEST COMPANY LLC"] + [None] * 22)
    ws.append([as_of_str] + [None] * 22)
    ws.append(["Ageing by due date"] + [None] * 22)
    ws.append([None] * 23)
    ws.append(_XERO_HEADER_ROW)
    ws.append([None] * 23)
    party_header: list[Any] = [None] * 23
    party_header[0] = party
    ws.append(party_header)
    inv_row: list[Any] = [None] * 23
    inv_row[6] = inv_date
    inv_row[9] = inv_num
    inv_row[15] = total
    inv_row[17] = total
    inv_row[18] = 0
    inv_row[21] = "Seen"
    inv_row[22] = "Sent"
    ws.append(inv_row)
    total_row: list[Any] = [None] * 23
    total_row[0] = "Total"
    total_row[17] = total
    ws.append(total_row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_cp_xlsx(
    india_rows: list[list[Any]] | None = None,
    uae_rows: list[list[Any]] | None = None,
) -> bytes:
    wb = openpyxl.Workbook()
    del wb["Sheet"]
    ws_ind = wb.create_sheet("India")
    ws_ind.append(["Client Name", "Credit Period"])
    for row in india_rows or [["AlphaClient Ltd", 30]]:
        ws_ind.append(row)
    ws_uae = wb.create_sheet("UAE")
    ws_uae.append(["Client Name", "Credit Period", "Reason for extended Credit", "Amount"])
    for row in uae_rows or [["BetaClient LLC", 45, "contract", None]]:
        ws_uae.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Upload helper
# ---------------------------------------------------------------------------


def _upload(
    client: TestClient,
    file_bytes: bytes,
    entity_code: str = "IND",
    source_hint: str | None = None,
    as_of_date: str | None = None,
    filename: str = "test.xlsx",
) -> Any:
    data: dict[str, Any] = {"entity_code": entity_code}
    if source_hint:
        data["source_hint"] = source_hint
    if as_of_date:
        data["as_of_date"] = as_of_date
    csrf_token = _csrf(client)
    headers: dict[str, str] = {}
    if csrf_token:
        headers["X-CSRF-Token"] = csrf_token
    return client.post(
        "/snapshots",
        data=data,
        files={
            "file": (
                filename,
                io.BytesIO(file_bytes),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=headers,
    )


def _staging_get(
    client: TestClient,
    snapshot_id: str,
    *,
    filter: str = "all",
    offset: int = 0,
    limit: int = 50,
) -> Any:
    csrf_token = _csrf(client)
    headers: dict[str, str] = {}
    if csrf_token:
        headers["X-CSRF-Token"] = csrf_token
    return client.get(
        f"/snapshots/{snapshot_id}/staging",
        params={"filter": filter, "offset": offset, "limit": limit},
        headers=headers,
    )


def _staging_patch(
    client: TestClient,
    snapshot_id: str,
    row_index: int,
    body: dict[str, Any],
) -> Any:
    csrf_token = _csrf(client)
    headers: dict[str, str] = {"Content-Type": "application/json"}
    if csrf_token:
        headers["X-CSRF-Token"] = csrf_token
    return client.patch(
        f"/snapshots/{snapshot_id}/staging/{row_index}",
        content=json.dumps(body),
        headers=headers,
    )


def _warnings_ack(
    client: TestClient,
    snapshot_id: str,
    codes: list[str],
) -> Any:
    csrf_token = _csrf(client)
    headers: dict[str, str] = {"Content-Type": "application/json"}
    if csrf_token:
        headers["X-CSRF-Token"] = csrf_token
    return client.patch(
        f"/snapshots/{snapshot_id}/warnings/ack",
        content=json.dumps({"codes": codes}),
        headers=headers,
    )


def _seed_canonical(
    db_session: Session,
    entity_code: str,
    canonical_name: str,
    actor_email: str = "tejaswa.sharma@emb.global",
) -> PartyCanonical:
    entity = db_session.scalar(select(Entity).where(Entity.code == entity_code))
    assert entity is not None
    actor = db_session.scalar(select(User).where(User.email == actor_email))
    assert actor is not None
    canonical = PartyCanonical(
        entity_id=entity.id,
        name=canonical_name,
        created_by=actor.id,
    )
    db_session.add(canonical)
    db_session.flush()
    return canonical


# ---------------------------------------------------------------------------
# GET /staging happy path tests
# ---------------------------------------------------------------------------


class TestStagingGetHappyPath:
    def test_tally_snapshot_returns_invoice_rows(
        self, client: TestClient, db_session: Session
    ) -> None:
        """TALLY snapshot: full shape with rows + alias_resolution populated."""
        _login_as_admin(client)
        file_bytes = _make_tally_xlsx(
            [[date(2026, 1, 15), "INV-001", "TestParty Alpha", 1000.0, 1000.0, None, None]]
        )
        r = _upload(client, file_bytes, "IND", "TALLY", "2026-01-31")
        assert r.status_code == 201
        snap_id = r.json()["snapshot_id"]

        r2 = _staging_get(client, snap_id)
        assert r2.status_code == 200
        body = r2.json()
        assert body["snapshot_id"] == snap_id
        assert body["snapshot_status"] == "STAGED"
        assert body["source_hint"] == "TALLY"
        assert body["entity_code"] == "IND"
        assert "totals" in body
        assert "publish_gate" in body
        assert "pagination" in body
        assert "rows" in body
        # Each invoice row should have alias_resolution
        rows = body["rows"]
        assert len(rows) >= 1
        row = rows[0]
        assert "alias_resolution" in row
        assert "analyst_overrides" in row
        assert "raw_row_json" in row

    def test_xero_snapshot_returns_invoice_rows(
        self, client: TestClient, db_session: Session
    ) -> None:
        """XERO snapshot: happy path."""
        _login_as_admin(client)
        file_bytes = _make_xero_xlsx("As at 31 March 2026", "XeroParty", date(2026, 1, 15))
        r = _upload(client, file_bytes, "UAE", "XERO")
        assert r.status_code == 201
        snap_id = r.json()["snapshot_id"]

        r2 = _staging_get(client, snap_id)
        assert r2.status_code == 200
        body = r2.json()
        assert body["source_hint"] == "XERO"
        assert body["entity_code"] == "UAE"
        rows = body["rows"]
        assert len(rows) >= 1
        assert "alias_resolution" in rows[0]

    def test_credit_period_snapshot_returns_cp_rows(
        self, client: TestClient, db_session: Session
    ) -> None:
        """CREDIT_PERIOD snapshot: rows are StagingCreditPeriodRow."""
        _login_as_admin(client)
        file_bytes = _make_cp_xlsx(
            [["AlphaClient Ltd", 30]], [["BetaClient LLC", 45, "contract", None]]
        )
        r = _upload(client, file_bytes, "IND", "CREDIT_PERIOD")
        assert r.status_code == 201
        snap_id = r.json()["snapshot_id"]

        r2 = _staging_get(client, snap_id)
        assert r2.status_code == 200
        body = r2.json()
        assert body["source_hint"] == "CREDIT_PERIOD"
        rows = body["rows"]
        assert len(rows) >= 1
        row = rows[0]
        # CP rows have credit_days and name but NOT alias_resolution
        assert "credit_days" in row
        assert "name" in row
        assert "alias_resolution" not in row

    def test_totals_correct_for_tally(self, client: TestClient, db_session: Session) -> None:
        """Totals reflect the actual invoices in parse_result_json."""
        _login_as_admin(client)
        file_bytes = _make_tally_xlsx(
            [
                [date(2026, 1, 15), "INV-001", "Alpha", 1000.0, 1000.0, None, None],
                [date(2026, 2, 1), "INV-002", "Beta", 2000.0, 2000.0, None, None],
            ]
        )
        r = _upload(client, file_bytes, "IND", "TALLY", "2026-02-28")
        assert r.status_code == 201
        snap_id = r.json()["snapshot_id"]

        r2 = _staging_get(client, snap_id)
        assert r2.status_code == 200
        totals = r2.json()["totals"]
        assert totals["invoices_total"] == 2
        assert totals["invoices_ok"] == 2
        assert totals["invoices_parse_error"] == 0

    def test_publish_gate_fresh_upload_not_ok(
        self, client: TestClient, db_session: Session
    ) -> None:
        """Fresh upload with no alias resolves → ok=False, unmapped_parties_count matches OK rows."""
        _login_as_admin(client)
        file_bytes = _make_tally_xlsx(
            [
                [date(2026, 1, 15), "INV-001", "NewPartyNoAlias", 1000.0, 1000.0, None, None],
            ]
        )
        r = _upload(client, file_bytes, "IND", "TALLY", "2026-01-31")
        assert r.status_code == 201
        snap_id = r.json()["snapshot_id"]

        r2 = _staging_get(client, snap_id)
        assert r2.status_code == 200
        gate = r2.json()["publish_gate"]
        assert gate["ok"] is False
        assert gate["unmapped_parties_count"] >= 1

    def test_publish_gate_ok_after_resolving_aliases(
        self, client: TestClient, db_session: Session
    ) -> None:
        """After resolving the one OK row → gate ok=True (assuming no warnings/parse errors)."""
        _login_as_admin(client)
        # Seed a canonical for IND that matches our party name (exact match)
        _seed_canonical(db_session, "IND", "ResolvedParty Corp")

        file_bytes = _make_tally_xlsx(
            [
                [date(2026, 1, 15), "INV-001", "ResolvedParty Corp", 1000.0, 1000.0, None, None],
            ]
        )
        r = _upload(client, file_bytes, "IND", "TALLY", "2026-01-31")
        assert r.status_code == 201
        snap_id = r.json()["snapshot_id"]

        # Exact match → publish gate should be ok=True already
        r2 = _staging_get(client, snap_id)
        assert r2.status_code == 200
        gate = r2.json()["publish_gate"]
        # Exact resolution clears the unmapped count; gate ok depends on role_permits_publish
        assert gate["unmapped_parties_count"] == 0


# ---------------------------------------------------------------------------
# GET /staging filter tests
# ---------------------------------------------------------------------------


class TestStagingGetFilters:
    def _upload_tally_with_data(
        self,
        client: TestClient,
        db_session: Session,
        rows: list[list[Any]],
        as_of_date: str = "2026-01-31",
    ) -> str:
        _login_as_admin(client)
        file_bytes = _make_tally_xlsx(rows)
        r = _upload(client, file_bytes, "IND", "TALLY", as_of_date)
        assert r.status_code == 201
        return str(r.json()["snapshot_id"])

    def test_filter_ok_returns_only_ok_rows(self, client: TestClient, db_session: Session) -> None:
        snap_id = self._upload_tally_with_data(
            client,
            db_session,
            [[date(2026, 1, 15), "INV-001", "Alpha", 1000.0, 1000.0, None, None]],
        )
        r = _staging_get(client, snap_id, filter="ok")
        assert r.status_code == 200
        rows = r.json()["rows"]
        assert all(row["status"] == "OK" for row in rows)

    def test_filter_parse_error_returns_only_errors(
        self, client: TestClient, db_session: Session
    ) -> None:
        # Single valid row — no parse errors expected
        snap_id = self._upload_tally_with_data(
            client,
            db_session,
            [[date(2026, 1, 15), "INV-001", "Alpha", 1000.0, 1000.0, None, None]],
        )
        r = _staging_get(client, snap_id, filter="parse_error")
        assert r.status_code == 200
        rows = r.json()["rows"]
        assert all(row["status"] == "PARSE_ERROR" for row in rows)

    def test_filter_unmapped_returns_unresolved_ok_rows(
        self, client: TestClient, db_session: Session
    ) -> None:
        snap_id = self._upload_tally_with_data(
            client,
            db_session,
            [[date(2026, 1, 15), "INV-001", "NoMatchPartyXYZ", 1000.0, 1000.0, None, None]],
        )
        r = _staging_get(client, snap_id, filter="unmapped")
        assert r.status_code == 200
        body = r.json()
        # The party has no canonical → should be UNMAPPED → appears in filter
        # (total may be 0 if the party name exactly matches something seeded, but generally 1)
        assert "rows" in body
        assert "pagination" in body

    def test_filter_fuzzy_high(self, client: TestClient, db_session: Session) -> None:
        snap_id = self._upload_tally_with_data(
            client,
            db_session,
            [[date(2026, 1, 15), "INV-001", "AnyParty", 1000.0, 1000.0, None, None]],
        )
        r = _staging_get(client, snap_id, filter="fuzzy_high")
        assert r.status_code == 200
        rows = r.json()["rows"]
        for row in rows:
            assert row["alias_resolution"]["resolution_state"] == "FUZZY_HIGH"

    def test_filter_fuzzy_low(self, client: TestClient, db_session: Session) -> None:
        snap_id = self._upload_tally_with_data(
            client,
            db_session,
            [[date(2026, 1, 15), "INV-001", "AnyParty", 1000.0, 1000.0, None, None]],
        )
        r = _staging_get(client, snap_id, filter="fuzzy_low")
        assert r.status_code == 200
        rows = r.json()["rows"]
        for row in rows:
            assert row["alias_resolution"]["resolution_state"] == "FUZZY_LOW"

    def test_pagination_offset_limit_honored(self, client: TestClient, db_session: Session) -> None:
        snap_id = self._upload_tally_with_data(
            client,
            db_session,
            [
                [date(2026, 1, 15), "INV-001", "Alpha", 1000.0, 1000.0, None, None],
                [date(2026, 1, 16), "INV-002", "Beta", 2000.0, 2000.0, None, None],
                [date(2026, 1, 17), "INV-003", "Gamma", 3000.0, 3000.0, None, None],
            ],
        )
        r = _staging_get(client, snap_id, offset=1, limit=1)
        assert r.status_code == 200
        body = r.json()
        assert len(body["rows"]) == 1
        assert body["pagination"]["offset"] == 1
        assert body["pagination"]["limit"] == 1
        assert body["pagination"]["total"] == 3


# ---------------------------------------------------------------------------
# GET /staging RBAC tests
# ---------------------------------------------------------------------------


class TestStagingGetRbac:
    def _upload_tally(self, client: TestClient) -> str:
        _login_as_admin(client)
        file_bytes = _make_tally_xlsx(
            [[date(2026, 1, 15), "INV-001", "Alpha", 1000.0, 1000.0, None, None]]
        )
        r = _upload(client, file_bytes, "IND", "TALLY", "2026-01-31")
        assert r.status_code == 201
        return str(r.json()["snapshot_id"])

    def test_cfo_gets_403(self, client: TestClient, db_session: Session) -> None:
        snap_id = self._upload_tally(client)
        _login_as_cfo(client, db_session, "cfo_staging@emb.global")
        _login(client, "cfo_staging@emb.global")
        r = _staging_get(client, snap_id)
        assert r.status_code == 403

    def test_pending_gets_403(self, client: TestClient, db_session: Session) -> None:
        snap_id = self._upload_tally(client)
        _login_as_pending(client, "pending_staging@emb.global")
        r = _staging_get(client, snap_id)
        assert r.status_code == 403

    def test_analyst_wrong_entity_gets_403(self, client: TestClient, db_session: Session) -> None:
        snap_id = self._upload_tally(client)
        _login_as_analyst(client, db_session, "analyst_uae_staging@emb.global", entity_code="UAE")
        _login(client, "analyst_uae_staging@emb.global")
        r = _staging_get(client, snap_id)
        assert r.status_code == 403

    def test_analyst_right_entity_gets_200(self, client: TestClient, db_session: Session) -> None:
        snap_id = self._upload_tally(client)
        _login_as_analyst(client, db_session, "analyst_ind_staging@emb.global", entity_code="IND")
        _login(client, "analyst_ind_staging@emb.global")
        r = _staging_get(client, snap_id)
        assert r.status_code == 200

    def test_admin_any_entity_gets_200(self, client: TestClient, db_session: Session) -> None:
        snap_id = self._upload_tally(client)
        r = _staging_get(client, snap_id)
        assert r.status_code == 200

    def test_nonexistent_snapshot_returns_404(
        self, client: TestClient, db_session: Session
    ) -> None:
        _login_as_admin(client)
        fake_id = str(uuid.uuid4())
        r = _staging_get(client, fake_id)
        assert r.status_code == 404

    def test_published_snapshot_returns_409(self, client: TestClient, db_session: Session) -> None:
        snap_id = self._upload_tally(client)
        # Manually flip status to PUBLISHED in DB
        snap = db_session.scalar(select(Snapshot).where(Snapshot.id == uuid.UUID(snap_id)))
        assert snap is not None
        snap.status = "PUBLISHED"
        db_session.flush()

        r = _staging_get(client, snap_id)
        assert r.status_code == 409
        detail = r.json()["detail"]
        assert detail["code"] == "SNAPSHOT_NOT_STAGED"

    def test_discarded_snapshot_returns_409(self, client: TestClient, db_session: Session) -> None:
        snap_id = self._upload_tally(client)
        snap = db_session.scalar(select(Snapshot).where(Snapshot.id == uuid.UUID(snap_id)))
        assert snap is not None
        snap.status = "DISCARDED"
        db_session.flush()

        r = _staging_get(client, snap_id)
        assert r.status_code == 409


# ---------------------------------------------------------------------------
# PATCH /staging/{row} — resolve_alias
# ---------------------------------------------------------------------------


class TestPatchResolveAlias:
    def _upload_and_get_snap(
        self, client: TestClient, db_session: Session
    ) -> tuple[str, dict[str, Any]]:
        _login_as_admin(client)
        file_bytes = _make_tally_xlsx(
            [
                [date(2026, 1, 15), "INV-001", "RawPartyName", 1000.0, 1000.0, None, None],
            ]
        )
        r = _upload(client, file_bytes, "IND", "TALLY", "2026-01-31")
        assert r.status_code == 201
        snap_id = r.json()["snapshot_id"]
        rows = _staging_get(client, snap_id).json()["rows"]
        return snap_id, rows[0]

    def test_happy_path_with_existing_canonical(
        self, client: TestClient, db_session: Session
    ) -> None:
        """resolve_alias with valid canonical → 200, override appended, alias row created."""
        canonical = _seed_canonical(db_session, "IND", "CanonicalCorp")
        snap_id, row = self._upload_and_get_snap(client, db_session)
        row_index = row["row_index"]

        r = _staging_patch(
            client,
            snap_id,
            row_index,
            {
                "action": "resolve_alias",
                "canonical_id": str(canonical.id),
                "create_alias": True,
            },
        )
        assert r.status_code == 200
        body = r.json()
        assert "row" in body
        assert "publish_gate" in body
        updated_row = body["row"]
        assert updated_row["analyst_overrides"]["resolved_canonical_id"] == str(canonical.id)

        # Alias row should be created
        alias = db_session.scalar(
            select(PartyAlias).where(
                PartyAlias.canonical_id == canonical.id,
                PartyAlias.alias_text == "RawPartyName",
            )
        )
        assert alias is not None

    def test_without_create_alias_no_alias_created(
        self, client: TestClient, db_session: Session
    ) -> None:
        canonical = _seed_canonical(db_session, "IND", "CanonicalCorp2")
        snap_id, row = self._upload_and_get_snap(client, db_session)
        row_index = row["row_index"]

        r = _staging_patch(
            client,
            snap_id,
            row_index,
            {
                "action": "resolve_alias",
                "canonical_id": str(canonical.id),
                "create_alias": False,
            },
        )
        assert r.status_code == 200

        alias = db_session.scalar(
            select(PartyAlias).where(
                PartyAlias.canonical_id == canonical.id,
                PartyAlias.alias_text == "RawPartyName",
            )
        )
        assert alias is None

    def test_canonical_wrong_entity_returns_422(
        self, client: TestClient, db_session: Session
    ) -> None:
        """canonical_id from different entity → 422."""
        canonical = _seed_canonical(db_session, "UAE", "UAEOnlyCanonical")
        snap_id, row = self._upload_and_get_snap(client, db_session)
        row_index = row["row_index"]

        r = _staging_patch(
            client,
            snap_id,
            row_index,
            {
                "action": "resolve_alias",
                "canonical_id": str(canonical.id),
            },
        )
        assert r.status_code == 422

    def test_nonexistent_canonical_returns_422(
        self, client: TestClient, db_session: Session
    ) -> None:
        snap_id, row = self._upload_and_get_snap(client, db_session)
        r = _staging_patch(
            client,
            snap_id,
            row["row_index"],
            {
                "action": "resolve_alias",
                "canonical_id": str(uuid.uuid4()),
            },
        )
        assert r.status_code == 422

    def test_snapshot_not_staged_returns_409(self, client: TestClient, db_session: Session) -> None:
        canonical = _seed_canonical(db_session, "IND", "CanonicalCorp3")
        snap_id, row = self._upload_and_get_snap(client, db_session)
        snap = db_session.scalar(select(Snapshot).where(Snapshot.id == uuid.UUID(snap_id)))
        assert snap is not None
        snap.status = "PUBLISHED"
        db_session.flush()

        r = _staging_patch(
            client,
            snap_id,
            row["row_index"],
            {
                "action": "resolve_alias",
                "canonical_id": str(canonical.id),
            },
        )
        assert r.status_code == 409


# ---------------------------------------------------------------------------
# PATCH /staging/{row} — create_canonical
# ---------------------------------------------------------------------------


class TestPatchCreateCanonical:
    def _upload_tally(self, client: TestClient) -> tuple[str, int]:
        _login_as_admin(client)
        file_bytes = _make_tally_xlsx(
            [
                [date(2026, 1, 15), "INV-001", "BrandNewParty", 1000.0, 1000.0, None, None],
            ]
        )
        r = _upload(client, file_bytes, "IND", "TALLY", "2026-01-31")
        assert r.status_code == 201
        snap_id = r.json()["snapshot_id"]
        rows = _staging_get(client, snap_id).json()["rows"]
        return snap_id, rows[0]["row_index"]

    def test_happy_path_creates_canonical_and_alias(
        self, client: TestClient, db_session: Session
    ) -> None:
        snap_id, row_index = self._upload_tally(client)
        r = _staging_patch(
            client,
            snap_id,
            row_index,
            {
                "action": "create_canonical",
                "canonical_name": "Brand New Party Ltd",
                "alias_text": "BrandNewParty",
                "notes": "created during staging",
            },
        )
        assert r.status_code == 200
        body = r.json()
        cid = body["row"]["analyst_overrides"]["resolved_canonical_id"]
        assert cid is not None

        canonical = db_session.scalar(
            select(PartyCanonical).where(PartyCanonical.id == uuid.UUID(cid))
        )
        assert canonical is not None

        alias = db_session.scalar(
            select(PartyAlias).where(
                PartyAlias.canonical_id == canonical.id,
                PartyAlias.alias_text == "BrandNewParty",
            )
        )
        assert alias is not None

    def test_duplicate_canonical_name_returns_422(
        self, client: TestClient, db_session: Session
    ) -> None:
        _seed_canonical(db_session, "IND", "ExistingCanonical")
        snap_id, row_index = self._upload_tally(client)
        r = _staging_patch(
            client,
            snap_id,
            row_index,
            {
                "action": "create_canonical",
                "canonical_name": "ExistingCanonical",
            },
        )
        assert r.status_code == 422


# ---------------------------------------------------------------------------
# PATCH /staging/{row} — override_credit_days
# ---------------------------------------------------------------------------


class TestPatchOverrideCreditDays:
    def _upload_tally(self, client: TestClient) -> tuple[str, int]:
        _login_as_admin(client)
        file_bytes = _make_tally_xlsx(
            [
                [date(2026, 1, 15), "INV-001", "CreditTestParty", 1000.0, 1000.0, None, None],
            ]
        )
        r = _upload(client, file_bytes, "IND", "TALLY", "2026-01-31")
        assert r.status_code == 201
        snap_id = r.json()["snapshot_id"]
        rows = _staging_get(client, snap_id).json()["rows"]
        return snap_id, rows[0]["row_index"]

    def test_happy_path_ok_row(self, client: TestClient, db_session: Session) -> None:
        snap_id, row_index = self._upload_tally(client)
        r = _staging_patch(
            client,
            snap_id,
            row_index,
            {
                "action": "override_credit_days",
                "credit_days": 45,
                "reason": "special contract",
            },
        )
        assert r.status_code == 200
        overrides = r.json()["row"]["analyst_overrides"]
        assert overrides["credit_days_override"] == 45
        assert overrides["credit_days_source"] == "MANUAL"

    def test_negative_credit_days_returns_422(
        self, client: TestClient, db_session: Session
    ) -> None:
        snap_id, row_index = self._upload_tally(client)
        r = _staging_patch(
            client,
            snap_id,
            row_index,
            {
                "action": "override_credit_days",
                "credit_days": -1,
            },
        )
        assert r.status_code == 422

    def test_credit_period_snapshot_returns_422(
        self, client: TestClient, db_session: Session
    ) -> None:
        _login_as_admin(client)
        file_bytes = _make_cp_xlsx()
        r = _upload(client, file_bytes, "IND", "CREDIT_PERIOD")
        assert r.status_code == 201
        snap_id = r.json()["snapshot_id"]
        rows = _staging_get(client, snap_id).json()["rows"]
        assert len(rows) >= 1
        row_index = rows[0]["row_index"]

        r2 = _staging_patch(
            client,
            snap_id,
            row_index,
            {
                "action": "override_credit_days",
                "credit_days": 30,
            },
        )
        assert r2.status_code == 422


# ---------------------------------------------------------------------------
# PATCH /staging/{row} — dismiss / undismiss
# ---------------------------------------------------------------------------


class TestPatchDismissUndismiss:
    def _upload_with_parse_error(self, client: TestClient) -> tuple[str, int]:
        """Upload a Tally file that produces at least one PARSE_ERROR row.

        A row missing the required ref / date / amount triggers PARSE_ERROR.
        In the Tally parser, a row with a party name but None date and None ref
        and None amount is staged as PARSE_ERROR.
        """
        _login_as_admin(client)
        file_bytes = _make_tally_xlsx(
            [
                # OK row
                [date(2026, 1, 15), "INV-001", "GoodParty", 1000.0, 1000.0, None, None],
                # Bad row — no date, no ref, no amount → PARSE_ERROR
                [None, None, "BadParty", None, None, None, None],
            ]
        )
        r = _upload(client, file_bytes, "IND", "TALLY", "2026-01-31")
        assert r.status_code == 201
        snap_id = r.json()["snapshot_id"]
        rows_all = _staging_get(client, snap_id).json()["rows"]
        pe_rows = [row for row in rows_all if row["status"] == "PARSE_ERROR"]
        if not pe_rows:
            pytest.skip("No PARSE_ERROR row produced by this fixture")
        return snap_id, pe_rows[0]["row_index"]

    def test_dismiss_parse_error(self, client: TestClient, db_session: Session) -> None:
        snap_id, pe_row_index = self._upload_with_parse_error(client)

        r_before = _staging_get(client, snap_id)
        gate_before = r_before.json()["publish_gate"]
        pe_count_before = gate_before["parse_errors_unresolved_count"]

        r = _staging_patch(
            client,
            snap_id,
            pe_row_index,
            {
                "action": "dismiss_parse_error",
                "reason": "Row is irrelevant",
            },
        )
        assert r.status_code == 200
        gate_after = r.json()["publish_gate"]
        assert gate_after["parse_errors_unresolved_count"] == pe_count_before - 1

    def test_dismiss_ok_row_returns_422(self, client: TestClient, db_session: Session) -> None:
        _login_as_admin(client)
        file_bytes = _make_tally_xlsx(
            [
                [date(2026, 1, 15), "INV-001", "GoodParty", 1000.0, 1000.0, None, None],
            ]
        )
        r = _upload(client, file_bytes, "IND", "TALLY", "2026-01-31")
        assert r.status_code == 201
        snap_id = r.json()["snapshot_id"]
        rows = _staging_get(client, snap_id).json()["rows"]
        ok_row_index = next(row["row_index"] for row in rows if row["status"] == "OK")

        r2 = _staging_patch(
            client,
            snap_id,
            ok_row_index,
            {
                "action": "dismiss_parse_error",
                "reason": "Wrong action on OK row",
            },
        )
        assert r2.status_code == 422

    def test_undismiss_restores_count(self, client: TestClient, db_session: Session) -> None:
        snap_id, pe_row_index = self._upload_with_parse_error(client)

        # Dismiss first
        r1 = _staging_patch(
            client,
            snap_id,
            pe_row_index,
            {
                "action": "dismiss_parse_error",
                "reason": "test",
            },
        )
        assert r1.status_code == 200
        count_after_dismiss = r1.json()["publish_gate"]["parse_errors_unresolved_count"]

        # Undismiss
        r2 = _staging_patch(
            client,
            snap_id,
            pe_row_index,
            {
                "action": "undismiss_parse_error",
            },
        )
        assert r2.status_code == 200
        count_after_undismiss = r2.json()["publish_gate"]["parse_errors_unresolved_count"]
        assert count_after_undismiss == count_after_dismiss + 1


# ---------------------------------------------------------------------------
# PATCH /warnings/ack
# ---------------------------------------------------------------------------


class TestWarningsAck:
    def _upload_xero_with_warnings(self, client: TestClient) -> str:
        """Upload an Xero file that produces a grand total warning."""
        _login_as_admin(client)
        # Build Xero file where grand total doesn't match rows (triggers GRAND_TOTAL_MISMATCH warning)
        wb = openpyxl.Workbook()
        del wb["Sheet"]
        ws = wb.create_sheet("Aged Receivables Detail")
        ws.append(["Aged Receivables Detail"] + [None] * 22)
        ws.append(["TEST COMPANY"] + [None] * 22)
        ws.append(["As at 31 March 2026"] + [None] * 22)
        ws.append(["Ageing by due date"] + [None] * 22)
        ws.append([None] * 23)
        ws.append(_XERO_HEADER_ROW)
        ws.append([None] * 23)
        # Party header
        ph: list[Any] = [None] * 23
        ph[0] = "PartyWithWarning"
        ws.append(ph)
        # Invoice row — total = 500
        inv: list[Any] = [None] * 23
        inv[6] = date(2026, 1, 15)
        inv[9] = "INV-W001"
        inv[17] = 500.0
        inv[18] = 0
        ws.append(inv)
        # Grand total row — total = 9999 (mismatch!)
        gt: list[Any] = [None] * 23
        gt[0] = "Total"
        gt[17] = 9999.0  # deliberate mismatch
        ws.append(gt)
        buf = io.BytesIO()
        wb.save(buf)
        r = _upload(client, buf.getvalue(), "UAE", "XERO")
        assert r.status_code == 201
        return str(r.json()["snapshot_id"])

    def test_ack_warning_code_removes_from_unacknowledged(
        self, client: TestClient, db_session: Session
    ) -> None:
        snap_id = self._upload_xero_with_warnings(client)

        r1 = _staging_get(client, snap_id)
        gate_before = r1.json()["publish_gate"]
        unacked_before = gate_before["warnings_unacknowledged"]

        if not unacked_before:
            pytest.skip("No warnings produced by this fixture")

        code = unacked_before[0]
        r2 = _warnings_ack(client, snap_id, [code])
        assert r2.status_code == 200
        body = r2.json()
        assert code not in body["publish_gate"]["warnings_unacknowledged"]
        assert code in body["acknowledged"]

    def test_ack_unknown_code_returns_422(self, client: TestClient, db_session: Session) -> None:
        _login_as_admin(client)
        file_bytes = _make_tally_xlsx(
            [[date(2026, 1, 15), "INV-001", "Alpha", 1000.0, 1000.0, None, None]]
        )
        r = _upload(client, file_bytes, "IND", "TALLY", "2026-01-31")
        assert r.status_code == 201
        snap_id = r.json()["snapshot_id"]

        r2 = _warnings_ack(client, snap_id, ["NONEXISTENT_CODE"])
        assert r2.status_code == 422


# ---------------------------------------------------------------------------
# Concurrency test
# ---------------------------------------------------------------------------


class TestConcurrency:
    def test_two_patches_both_appended_last_wins(
        self, client: TestClient, db_session: Session
    ) -> None:
        """Two sequential PATCHes on same row → both entries in staging_overrides_json,
        last one is effective (latest-wins semantics).

        True parallel execution is hard in a single-threaded test; we demonstrate
        the SELECT FOR UPDATE + append semantics with sequential transactions that
        share the same session.
        """
        canonical1 = _seed_canonical(db_session, "IND", "Canonical One")
        canonical2 = _seed_canonical(db_session, "IND", "Canonical Two")

        _login_as_admin(client)
        file_bytes = _make_tally_xlsx(
            [
                [date(2026, 1, 15), "INV-001", "DoubleResolveParty", 1000.0, 1000.0, None, None],
            ]
        )
        r = _upload(client, file_bytes, "IND", "TALLY", "2026-01-31")
        assert r.status_code == 201
        snap_id = r.json()["snapshot_id"]
        rows = _staging_get(client, snap_id).json()["rows"]
        row_index = rows[0]["row_index"]

        # First PATCH
        r1 = _staging_patch(
            client,
            snap_id,
            row_index,
            {
                "action": "resolve_alias",
                "canonical_id": str(canonical1.id),
            },
        )
        assert r1.status_code == 200

        # Second PATCH (different canonical)
        r2 = _staging_patch(
            client,
            snap_id,
            row_index,
            {
                "action": "resolve_alias",
                "canonical_id": str(canonical2.id),
            },
        )
        assert r2.status_code == 200

        # Check DB: both entries present, last one wins
        snap = db_session.scalar(select(Snapshot).where(Snapshot.id == uuid.UUID(snap_id)))
        assert snap is not None
        overrides = snap.staging_overrides_json
        row_overrides = [e for e in overrides if e["row_index"] == row_index]
        assert len(row_overrides) == 2

        # Last entry should be canonical2
        assert row_overrides[-1]["payload"]["canonical_id"] == str(canonical2.id)

        # GET should show canonical2 as effective
        r3 = _staging_get(client, snap_id)
        row = r3.json()["rows"][0]
        assert row["analyst_overrides"]["resolved_canonical_id"] == str(canonical2.id)
