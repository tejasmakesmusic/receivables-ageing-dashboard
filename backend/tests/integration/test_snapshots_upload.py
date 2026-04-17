"""Integration tests for POST /snapshots upload endpoint (M3 Task 2).

Uses the function-scoped ``client`` fixture (DB-backed, per-test rollback).
Auth via stub Google OAuth callback (same pattern as test_admin_users.py).

Seeded data (migration 0002):
    - tejaswa.sharma@emb.global → ADMIN
    - D9 entities seeded in 0003_m3_ingestion

Entity seed (0003):
    - EMB_IN → IND
    - MANTARAV_UAE → UAE
    (confirmed in test_m3_migration.py)

Partition coverage (0003_m3_ingestion):
    - invoice_snapshots_2026_q1: 2026-01-01 → 2026-04-01
    - invoice_snapshots_2026_q2: 2026-04-01 → 2026-07-01
"""

from __future__ import annotations

import io
import uuid
from datetime import date
from typing import TYPE_CHECKING, Any

import openpyxl
from sqlalchemy import select, text

from app.core.rbac import Role
from app.db.models.audit_log import AuditLog
from app.db.models.entity import Entity
from app.db.models.snapshot import Snapshot
from app.db.models.user import User

if TYPE_CHECKING:
    from fastapi.testclient import TestClient
    from sqlalchemy.orm import Session


# ---------------------------------------------------------------------------
# Auth / session helpers
# ---------------------------------------------------------------------------


def _login(client: TestClient, email: str) -> None:
    """Log in via stub Google OAuth callback."""
    client.get(
        f"/auth/google/callback?stub_email={email}",
        follow_redirects=False,
    )


def _csrf(client: TestClient) -> str:
    return client.cookies.get("csrf_token", "")


def _login_as_admin(client: TestClient) -> None:
    _login(client, "tejaswa.sharma@emb.global")


def _login_as_pending(client: TestClient, email: str) -> None:
    _login(client, email)


def _login_as_analyst(
    client: TestClient,
    db_session: Session,
    email: str,
    entity_code: str | None = None,
) -> uuid.UUID:
    """Create + approve a user as ANALYST, optionally binding them to an entity."""
    _login(client, email)
    user = db_session.scalar(select(User).where(User.email == email))
    assert user is not None, f"Expected user {email!r} to exist after OAuth callback"

    user.role = Role.ANALYST
    if entity_code is not None:
        entity = db_session.scalar(select(Entity).where(Entity.code == entity_code))
        assert entity is not None, f"Entity code {entity_code!r} not found"
        user.entity_id_scope = entity.id
    else:
        user.entity_id_scope = None

    user.is_active = True
    db_session.flush()
    return user.id


def _login_as_cfo(client: TestClient, db_session: Session, email: str) -> None:
    """Create + set CFO role."""
    _login(client, email)
    user = db_session.scalar(select(User).where(User.email == email))
    assert user is not None
    user.role = Role.CFO
    user.is_active = True
    db_session.flush()


# ---------------------------------------------------------------------------
# XLSX builders (re-use builder patterns from parser tests)
# ---------------------------------------------------------------------------


def _make_tally_xlsx(
    data_rows: list[list[Any]] | None = None,
    sheet_name: str = "Sundry Debtors",
) -> bytes:
    """Minimal Tally XLSX with correct sheet name."""
    _meta = [
        ["Group :", "Sundry Debtors", None, "1-Apr-26 to 16-Apr-26", None, None, None],
        ["Details of:", "Pending Bills", None, None, None, None, None],
        [None] * 7,
        ["Date", "Ref. No.", "Party's Name", "Opening", "Pending", "Due on", "Overdue"],
        [None, None, None, "Amount", "Amount", None, "by days"],
    ]
    wb = openpyxl.Workbook()
    del wb["Sheet"]
    ws = wb.create_sheet(sheet_name)
    for row in _meta:
        ws.append(row)
    for row in data_rows or []:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _tally_data_row(
    inv_date: Any = date(2026, 1, 15),
    ref: str = "INV-001",
    party: str = "TestParty",
    opening: float = 1000.0,
    pending: float = 1000.0,
    due_on: Any = None,
    overdue: Any = None,
) -> list[Any]:
    return [inv_date, ref, party, opening, pending, due_on, overdue]


_XERO_HEADER_ROW: list[Any] = [
    "Contact Account Number",
    "Primary Person",
    "Phone",
    "Email",
    "Mobile",
    "Contact Group",
    "Invoice Date",
    "Due Date",
    "Expected Date",
    "Invoice Number",
    "Invoice Reference",
    "< 1 Month",
    "1 Month",
    "2 Months",
    "3 Months",
    "Older",
    None,  # col 16 unnamed
    "Total",
    "Outstanding Tax",
    "PROJECT ID",
    "SERVICE MONTH",
    "Invoice Seen",
    "Invoice Sent",
]


def _make_xero_xlsx(
    as_of_str: str = "As at 31 March 2026",
    party: str = "TestParty",
    inv_date: Any = date(2026, 1, 15),
    inv_num: str = "INV-001",
    total: float = 1000.0,
    sheet_name: str = "Aged Receivables Detail",
) -> bytes:
    wb = openpyxl.Workbook()
    del wb["Sheet"]
    ws = wb.create_sheet(sheet_name)

    # Metadata rows 0-4 (0-indexed)
    ws.append(["Aged Receivables Detail"] + [None] * 22)  # row 1
    ws.append(["TEST COMPANY LLC"] + [None] * 22)  # row 2
    ws.append([as_of_str] + [None] * 22)  # row 3
    ws.append(["Ageing by due date"] + [None] * 22)  # row 4
    ws.append([None] * 23)  # row 5 blank
    ws.append(_XERO_HEADER_ROW)  # row 6 — headers
    ws.append([None] * 23)  # row 7 blank gap

    # Party header
    party_header = [None] * 23
    party_header[0] = party
    ws.append(party_header)

    # Invoice row
    inv_row: list[Any] = [None] * 23
    inv_row[6] = inv_date
    inv_row[9] = inv_num
    inv_row[15] = total  # Older bucket
    inv_row[17] = total  # Total
    inv_row[18] = 0
    inv_row[21] = "Seen"
    inv_row[22] = "Sent"
    ws.append(inv_row)

    # Grand total row
    total_row: list[Any] = [None] * 23
    total_row[0] = "Total"
    total_row[17] = total
    ws.append(total_row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_cp_xlsx(
    india_rows: list[list[Any]] | None = None,
    uae_rows: list[list[Any]] | None = None,
) -> bytes:
    wb = openpyxl.Workbook()
    del wb["Sheet"]

    ws_ind = wb.create_sheet("India")
    ws_ind.append(["Client Name", "Credit Period"])
    for row in india_rows or [["AlphaClient Ltd", 30]]:
        ws_ind.append(row)

    ws_uae = wb.create_sheet("UAE")
    ws_uae.append(["Client Name", "Credit Period", "Reason for extended Credit", "Amount"])
    for row in uae_rows or [["BetaClient LLC", 45, "Long-term contract", None]]:
        ws_uae.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(
    client: TestClient,
    file_bytes: bytes,
    entity_code: str = "IND",
    source_hint: str | None = None,
    as_of_date: str | None = None,
    filename: str = "test_upload.xlsx",
) -> Any:
    """POST /snapshots with multipart form data.

    CSRF token is taken from the cookie jar (set by any prior GET response such
    as the OAuth callback login) and passed as the X-CSRF-Token header per the
    CSRFMiddleware multipart handling.
    """
    data: dict[str, Any] = {"entity_code": entity_code}
    if source_hint:
        data["source_hint"] = source_hint
    if as_of_date:
        data["as_of_date"] = as_of_date

    csrf_token = _csrf(client)
    headers: dict[str, str] = {}
    if csrf_token:
        headers["X-CSRF-Token"] = csrf_token

    return client.post(
        "/snapshots",
        data=data,
        files={
            "file": (
                filename,
                io.BytesIO(file_bytes),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=headers,
    )


# ---------------------------------------------------------------------------
# RBAC gate tests
# ---------------------------------------------------------------------------


class TestSnapshotUploadRbac:
    def test_pending_user_gets_403(self, client: TestClient, db_session: Session) -> None:
        _login_as_pending(client, "pending_upload@emb.global")
        r = _upload(client, _make_tally_xlsx(), "IND", "TALLY", "2026-01-31")
        assert r.status_code == 403

    def test_cfo_gets_403(self, client: TestClient, db_session: Session) -> None:
        _login_as_cfo(client, db_session, "cfo_upload@emb.global")
        _login(client, "cfo_upload@emb.global")  # refresh cookie with updated role
        r = _upload(client, _make_tally_xlsx(), "IND", "TALLY", "2026-01-31")
        assert r.status_code == 403

    def test_analyst_wrong_entity_gets_403(self, client: TestClient, db_session: Session) -> None:
        """ANALYST scoped to UAE cannot upload to IND."""
        _login_as_analyst(client, db_session, "analyst_uae@emb.global", entity_code="UAE")
        _login(client, "analyst_uae@emb.global")
        r = _upload(client, _make_tally_xlsx(), "IND", "TALLY", "2026-01-31")
        assert r.status_code == 403

    def test_unauthenticated_gets_401(self, client: TestClient) -> None:
        # Hit /health to get a CSRF cookie without logging in,
        # so the CSRF middleware passes and auth check fires.
        client.get("/health")
        r = _upload(client, _make_tally_xlsx(), "IND", "TALLY", "2026-01-31")
        assert r.status_code == 401

    def test_csrf_missing_returns_403(self, client: TestClient, db_session: Session) -> None:
        """Valid session but no X-CSRF-Token header → CSRF middleware rejects with 403."""
        _login_as_admin(client)

        # Record snapshot count before — no snapshot should be created.
        snapshot_count_before = db_session.scalar(text("SELECT COUNT(*) FROM snapshots"))

        file_bytes = _make_tally_xlsx([_tally_data_row()])
        # Post WITHOUT the CSRF header (bypass _upload helper which adds it).
        r = client.post(
            "/snapshots",
            data={"entity_code": "IND", "source_hint": "TALLY", "as_of_date": "2026-01-31"},
            files={
                "file": (
                    "test_csrf.xlsx",
                    io.BytesIO(file_bytes),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            # Deliberately omit headers= so X-CSRF-Token is absent.
        )

        assert r.status_code == 403, f"Expected 403 from CSRF guard, got {r.status_code}: {r.text}"

        snapshot_count_after = db_session.scalar(text("SELECT COUNT(*) FROM snapshots"))
        assert (
            snapshot_count_after == snapshot_count_before
        ), "No snapshot should be created when CSRF token is missing"

    def test_admin_can_upload_any_entity(self, client: TestClient, db_session: Session) -> None:
        _login_as_admin(client)
        r = _upload(
            client,
            _make_tally_xlsx([_tally_data_row()]),
            entity_code="IND",
            source_hint="TALLY",
            as_of_date="2026-01-31",
        )
        assert r.status_code == 201

    def test_analyst_own_entity_allowed(self, client: TestClient, db_session: Session) -> None:
        """ANALYST scoped to IND can upload to IND."""
        _login_as_analyst(client, db_session, "analyst_ind@emb.global", entity_code="IND")
        _login(client, "analyst_ind@emb.global")
        r = _upload(
            client,
            _make_tally_xlsx([_tally_data_row()]),
            entity_code="IND",
            source_hint="TALLY",
            as_of_date="2026-01-31",
        )
        assert r.status_code == 201


# ---------------------------------------------------------------------------
# Happy-path tests
# ---------------------------------------------------------------------------


class TestSnapshotUploadHappyPath:
    def test_tally_upload_returns_201(self, client: TestClient, db_session: Session) -> None:
        _login_as_admin(client)
        # Synthetic file has exactly 1 invoice row (_tally_data_row has date + ref → invoice).
        file_bytes = _make_tally_xlsx([_tally_data_row()])
        r = _upload(client, file_bytes, "IND", "TALLY", "2026-01-31")
        assert r.status_code == 201
        body = r.json()
        assert body["status"] == "STAGED"
        assert body["source_hint"] == "TALLY"
        assert body["as_of_date"] == "2026-01-31"
        assert "snapshot_id" in body
        assert "file_sha256" in body
        assert "parse_summary" in body
        summary = body["parse_summary"]
        assert summary["invoices_parsed"] == 1  # exactly 1 well-formed invoice row
        assert summary["parse_error_count"] == 0  # synthetic file is well-formed

    def test_xero_upload_returns_201(self, client: TestClient, db_session: Session) -> None:
        _login_as_admin(client)
        file_bytes = _make_xero_xlsx("As at 31 March 2026")
        r = _upload(client, file_bytes, "UAE", "XERO", "2026-03-31")
        assert r.status_code == 201
        body = r.json()
        assert body["status"] == "STAGED"
        assert body["source_hint"] == "XERO"
        # Sniffed date from file should win: 31 March 2026
        assert body["as_of_date"] == "2026-03-31"

    def test_xero_auto_detect_source(self, client: TestClient, db_session: Session) -> None:
        """source_hint not supplied → auto-detected from sheet names."""
        _login_as_admin(client)
        file_bytes = _make_xero_xlsx("As at 15 January 2026")
        r = _upload(client, file_bytes, "UAE", source_hint=None, as_of_date=None)
        assert r.status_code == 201
        body = r.json()
        assert body["source_hint"] == "XERO"
        assert body["as_of_date"] == "2026-01-15"

    def test_tally_auto_detect_source(self, client: TestClient, db_session: Session) -> None:
        _login_as_admin(client)
        file_bytes = _make_tally_xlsx([_tally_data_row()])
        r = _upload(client, file_bytes, "IND", source_hint=None, as_of_date="2026-02-28")
        assert r.status_code == 201
        body = r.json()
        assert body["source_hint"] == "TALLY"

    def test_credit_period_upload_returns_201(
        self, client: TestClient, db_session: Session
    ) -> None:
        _login_as_admin(client)
        file_bytes = _make_cp_xlsx()
        r = _upload(client, file_bytes, "IND", "CREDIT_PERIOD", as_of_date=None)
        assert r.status_code == 201
        body = r.json()
        assert body["status"] == "STAGED"
        assert body["source_hint"] == "CREDIT_PERIOD"
        assert body["as_of_date"] is None

    def test_parse_summary_structure(self, client: TestClient, db_session: Session) -> None:
        _login_as_admin(client)
        file_bytes = _make_tally_xlsx([_tally_data_row()])
        r = _upload(client, file_bytes, "IND", "TALLY", "2026-01-31")
        body = r.json()
        summary = body["parse_summary"]
        assert "invoices_parsed" in summary
        assert "credit_periods_parsed" in summary
        assert "parse_error_count" in summary
        assert "warnings" in summary


# ---------------------------------------------------------------------------
# Duplicate file (sha256)
# ---------------------------------------------------------------------------


class TestSnapshotDuplicateSha256:
    def test_same_file_second_upload_returns_409(
        self, client: TestClient, db_session: Session
    ) -> None:
        _login_as_admin(client)
        file_bytes = _make_tally_xlsx([_tally_data_row(ref="INV-DUP-001")])

        r1 = _upload(client, file_bytes, "IND", "TALLY", "2026-01-31")
        assert r1.status_code == 201
        first_id = r1.json()["snapshot_id"]

        r2 = _upload(client, file_bytes, "IND", "TALLY", "2026-01-31")
        assert r2.status_code == 409
        body = r2.json()
        detail = body["detail"]
        assert detail["code"] == "DUPLICATE_FILE"
        assert detail["existing_snapshot_id"] == first_id


# ---------------------------------------------------------------------------
# as_of_date validation
# ---------------------------------------------------------------------------


class TestAsOfDateValidation:
    def test_tally_missing_as_of_date_returns_422(
        self, client: TestClient, db_session: Session
    ) -> None:
        _login_as_admin(client)
        file_bytes = _make_tally_xlsx([_tally_data_row()])
        r = _upload(client, file_bytes, "IND", "TALLY", as_of_date=None)
        assert r.status_code == 422
        detail = r.json()["detail"]
        assert detail["code"] == "AS_OF_DATE_MISSING"

    def test_xero_sniff_used_when_form_absent(
        self, client: TestClient, db_session: Session
    ) -> None:
        """Xero without form as_of_date uses sniffed value."""
        _login_as_admin(client)
        file_bytes = _make_xero_xlsx("As at 15 April 2026")
        r = _upload(client, file_bytes, "UAE", "XERO", as_of_date=None)
        assert r.status_code == 201
        assert r.json()["as_of_date"] == "2026-04-15"

    def test_xero_both_absent_returns_422(self, client: TestClient, db_session: Session) -> None:
        """Xero with no sniffable date and no form value → 422."""
        _login_as_admin(client)
        # Build Xero file with no "As at" text.
        wb = openpyxl.Workbook()
        del wb["Sheet"]
        ws = wb.create_sheet("Aged Receivables Detail")
        ws.append(["Aged Receivables Detail"] + [None] * 22)
        ws.append(["Test Company"] + [None] * 22)
        ws.append(["Not an as-at date row"] + [None] * 22)
        ws.append([None] * 23)
        ws.append([None] * 23)
        ws.append(_XERO_HEADER_ROW)
        buf = io.BytesIO()
        wb.save(buf)
        r = _upload(client, buf.getvalue(), "UAE", "XERO", as_of_date=None)
        assert r.status_code == 422
        detail = r.json()["detail"]
        assert detail["code"] == "AS_OF_DATE_MISSING"


# ---------------------------------------------------------------------------
# Partition pre-flight
# ---------------------------------------------------------------------------


class TestPartitionPreflight:
    def test_missing_partition_returns_422(self, client: TestClient, db_session: Session) -> None:
        """Upload with as_of_date outside Q1/Q2 2026 → 422 MISSING_PARTITION."""
        _login_as_admin(client)
        file_bytes = _make_tally_xlsx([_tally_data_row()])
        # 2025-12-31 has no partition.
        r = _upload(client, file_bytes, "IND", "TALLY", "2025-12-31")
        assert r.status_code == 422
        detail = r.json()["detail"]
        assert detail["code"] == "MISSING_PARTITION"
        assert "as_of_date" in detail
        assert (
            "runbook" in detail["hint"].lower()
        ), "MISSING_PARTITION hint must reference the runbook so operators know where to look"

    def test_credit_period_skips_partition_check(
        self, client: TestClient, db_session: Session
    ) -> None:
        """CREDIT_PERIOD uploads skip partition pre-flight (no as_of_date)."""
        _login_as_admin(client)
        file_bytes = _make_cp_xlsx()
        # No as_of_date — CREDIT_PERIOD doesn't need one.
        r = _upload(client, file_bytes, "IND", "CREDIT_PERIOD", as_of_date=None)
        assert r.status_code == 201  # not 422


# ---------------------------------------------------------------------------
# Source auto-detect and mismatch
# ---------------------------------------------------------------------------


class TestSourceDetection:
    def test_ambiguous_source_returns_400(self, client: TestClient, db_session: Session) -> None:
        """File with both Tally + Xero sheets → 400 ambiguous."""
        _login_as_admin(client)
        wb = openpyxl.Workbook()
        del wb["Sheet"]
        wb.create_sheet("Sundry Debtors")
        wb.create_sheet("Aged Receivables Detail")
        buf = io.BytesIO()
        wb.save(buf)
        r = _upload(client, buf.getvalue(), "IND", source_hint=None, as_of_date="2026-01-31")
        assert r.status_code == 400

    def test_no_match_source_returns_400(self, client: TestClient, db_session: Session) -> None:
        """File with unrecognized sheet names → 400."""
        _login_as_admin(client)
        wb = openpyxl.Workbook()
        wb.active.title = "RandomSheet"  # type: ignore[union-attr]
        buf = io.BytesIO()
        wb.save(buf)
        r = _upload(client, buf.getvalue(), "IND", source_hint=None, as_of_date="2026-01-31")
        assert r.status_code == 400

    def test_source_hint_mismatch_returns_400(
        self, client: TestClient, db_session: Session
    ) -> None:
        """Caller supplies TALLY but file has Xero sheet → 400."""
        _login_as_admin(client)
        file_bytes = _make_xero_xlsx()
        r = _upload(client, file_bytes, "UAE", "TALLY", "2026-01-31")
        assert r.status_code == 400

    def test_unknown_source_hint_returns_400(self, client: TestClient, db_session: Session) -> None:
        _login_as_admin(client)
        r = _upload(client, _make_tally_xlsx(), "IND", "BOGUS", "2026-01-31")
        assert r.status_code == 400

    def test_empty_file_returns_400(self, client: TestClient, db_session: Session) -> None:
        _login_as_admin(client)
        r = _upload(client, b"", "IND", "TALLY", "2026-01-31")
        assert r.status_code == 400


# ---------------------------------------------------------------------------
# File-level parse errors block snapshot creation
# ---------------------------------------------------------------------------


class TestParseErrorsBlockSnapshot:
    def test_file_level_parse_error_returns_422_no_snapshot_created(
        self, client: TestClient, db_session: Session
    ) -> None:
        """TALLY upload with wrong sheet name → SHEET_NOT_FOUND → 422, no snapshot row.

        Using source_hint="TALLY" with a file whose only sheet is named
        "WrongSheet" (not "Sundry Debtors") deterministically triggers the
        parser's SHEET_NOT_FOUND file-level error.  The source-hint validation
        step passes because we also include "Sundry Debtors" in the workbook
        for detection, then replace it with the wrong name for the actual data.

        Simpler approach: build a file that has "Sundry Debtors" as required by
        source detection but a *second* sheet that is the only one with data —
        no, the parser opens _SHEET_NAME directly.  Instead we build a file
        whose only sheet name is "WrongSheet" and supply source_hint="TALLY" so
        auto-detection is skipped; the parser then fails to find "Sundry Debtors".
        """
        _login_as_admin(client)

        # Build a minimal XLSX with no "Sundry Debtors" sheet.
        wb = openpyxl.Workbook()
        wb.active.title = "WrongSheet"  # type: ignore[union-attr]
        wb.active.append(["some", "data", "here"])  # type: ignore[union-attr]
        buf = io.BytesIO()
        wb.save(buf)
        file_bytes = buf.getvalue()

        # Record snapshot count before upload attempt.
        sha256_hex = __import__("hashlib").sha256(file_bytes).hexdigest()
        snapshot_count_before = db_session.scalar(text("SELECT COUNT(*) FROM snapshots"))

        # source_hint="TALLY" bypasses auto-detect; parser opens "Sundry Debtors" → fails.
        r = _upload(client, file_bytes, "IND", "TALLY", "2026-01-31")

        assert r.status_code == 422, f"Expected 422, got {r.status_code}: {r.text}"
        body = r.json()
        detail = body["detail"]
        assert detail["code"] == "PARSE_ERROR"
        error_codes = [e["code"] for e in detail["errors"]]
        assert "SHEET_NOT_FOUND" in error_codes, f"Expected SHEET_NOT_FOUND in {error_codes}"

        # No snapshot row must have been persisted for this file.
        snapshot_count_after = db_session.scalar(text("SELECT COUNT(*) FROM snapshots"))
        assert (
            snapshot_count_after == snapshot_count_before
        ), "Snapshot must NOT be created when file-level parse errors are present"
        rows_for_sha = db_session.scalar(
            text("SELECT COUNT(*) FROM snapshots WHERE upload_file_sha256 = :sha"),
            {"sha": sha256_hex},
        )
        assert rows_for_sha == 0, "No snapshot row should exist for the rejected file's sha256"


# ---------------------------------------------------------------------------
# Audit log
# ---------------------------------------------------------------------------


class TestSnapshotAuditLog:
    def test_upload_creates_audit_log_row(self, client: TestClient, db_session: Session) -> None:
        _login_as_admin(client)
        file_bytes = _make_tally_xlsx([_tally_data_row(ref="INV-AUDIT-001")])
        r = _upload(client, file_bytes, "IND", "TALLY", "2026-02-28")
        assert r.status_code == 201

        snapshot_id_str = r.json()["snapshot_id"]
        snapshot_id = uuid.UUID(snapshot_id_str)

        audit_rows = list(
            db_session.scalars(
                select(AuditLog).where(
                    AuditLog.action == "snapshot.upload",
                    AuditLog.entity_id == snapshot_id,
                )
            )
        )
        assert len(audit_rows) == 1, "Exactly one snapshot.upload audit row expected"
        audit = audit_rows[0]
        assert audit.entity_type == "snapshots"
        assert audit.after is not None
        assert audit.after.get("source_hint") == "TALLY"
        assert audit.after.get("file_sha256") == r.json()["file_sha256"]

    def test_snapshot_row_created_in_db(self, client: TestClient, db_session: Session) -> None:
        _login_as_admin(client)
        file_bytes = _make_tally_xlsx([_tally_data_row(ref="INV-DB-001")])
        r = _upload(client, file_bytes, "IND", "TALLY", "2026-03-15")
        assert r.status_code == 201

        snapshot_id = uuid.UUID(r.json()["snapshot_id"])
        snapshot = db_session.get(Snapshot, snapshot_id)
        assert snapshot is not None
        assert snapshot.status == "STAGED"
        assert snapshot.source_hint == "TALLY"
        assert snapshot.as_of_date == date(2026, 3, 15)
        assert snapshot.parse_result_json is not None
