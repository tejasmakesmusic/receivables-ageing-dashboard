"""Integration tests for POST /snapshots/:id/publish (M3 Task 5).

Test strategy:
- Reuses auth helpers and XLSX builders from test_staging_api.py.
- Per-test DB rollback (function-scoped `client` + `db_session` fixtures).
- Each test that needs a fully-publishable snapshot:
    1. Uploads a TALLY or XERO xlsx via POST /snapshots.
    2. Resolves all party aliases via PATCH /snapshots/{id}/staging/{row}.
    3. Calls POST /snapshots/{id}/publish.

Seeded data (migration 0002 + 0003):
    - tejaswa.sharma@emb.global → ADMIN
    - Entity IND (EMB_IN), Entity UAE (MANTARAV_UAE)
    - Entity IND default_credit_days = NULL (set in tests as needed)
    - Partitions: invoice_snapshots_2026_q1, invoice_snapshots_2026_q2
"""

from __future__ import annotations

import io
import threading
import uuid
from datetime import date
from decimal import Decimal
from typing import TYPE_CHECKING, Any

import openpyxl
import pytest
from sqlalchemy import select

from app.core.rbac import Role
from app.db.models.audit_log import AuditLog
from app.db.models.email_outbox import EmailOutbox
from app.db.models.entity import Entity
from app.db.models.exception_bucket_type import ExceptionBucketType
from app.db.models.exception_tag import ExceptionTag
from app.db.models.invoice import Invoice
from app.db.models.invoice_snapshot import InvoiceSnapshot
from app.db.models.party import PartyAlias, PartyCanonical
from app.db.models.snapshot import Snapshot
from app.db.models.user import User

if TYPE_CHECKING:
    from fastapi.testclient import TestClient
    from sqlalchemy.orm import Session


# ---------------------------------------------------------------------------
# Reconciliation helper (satisfies §13 #6 gate between publishes)
# ---------------------------------------------------------------------------


def _reconcile_snapshot_directly(
    db_session: Session,
    snapshot_id: str,
) -> None:
    """Insert a MATCHED ReconciliationEntry row directly via db_session.

    Used in multi-snapshot tests to satisfy the §13 #6 gate that blocks the
    next publish until the prior published snapshot is MATCHED.  We bypass the
    HTTP API here because setting up an ADMIN session just for reconciliation in
    every test would be heavier than required and the gate only cares about the
    DB row status, not the route that created it.

    delta = dashboard_ar + exception_bucket_total - tally_xero_closing_ar (D19)
    Setting all three to 0.00 → delta = 0 → MATCHED (well within ₹100 tolerance).
    """
    from sqlalchemy import select

    from app.db.models.reconciliation_entry import ReconciliationEntry

    admin = db_session.scalar(select(User).where(User.email == "tejaswa.sharma@emb.global"))
    assert admin is not None

    # Upsert: if an entry exists already, update it; otherwise insert.
    existing = db_session.scalar(
        select(ReconciliationEntry).where(ReconciliationEntry.snapshot_id == uuid.UUID(snapshot_id))
    )
    if existing is not None:
        existing.tally_xero_closing_ar = Decimal("0.00")
        existing.delta = Decimal("0.00")
        existing.status = "MATCHED"
        existing.entered_by = admin.id
        existing.notes = "test-helper auto-reconcile"
    else:
        db_session.add(
            ReconciliationEntry(
                snapshot_id=uuid.UUID(snapshot_id),
                dashboard_ar=Decimal("0.00"),
                exception_bucket_total=Decimal("0.00"),
                exception_bucket_breakdown={},
                tally_xero_closing_ar=Decimal("0.00"),
                delta=Decimal("0.00"),
                status="MATCHED",
                entered_by=admin.id,
                notes="test-helper auto-reconcile",
            )
        )
    db_session.flush()


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------


def _login(client: TestClient, email: str) -> None:
    client.get(f"/auth/google/callback?stub_email={email}", follow_redirects=False)


def _csrf(client: TestClient) -> str:
    return client.cookies.get("csrf_token", "")


def _login_as_admin(client: TestClient) -> None:
    _login(client, "tejaswa.sharma@emb.global")


def _login_as_analyst(
    client: TestClient,
    db_session: Session,
    email: str,
    entity_code: str | None = None,
) -> uuid.UUID:
    _login(client, email)
    user = db_session.scalar(select(User).where(User.email == email))
    assert user is not None
    user.role = Role.ANALYST
    if entity_code is not None:
        entity = db_session.scalar(select(Entity).where(Entity.code == entity_code))
        assert entity is not None
        user.entity_id_scope = entity.id
    else:
        user.entity_id_scope = None
    user.is_active = True
    db_session.flush()
    return user.id


def _login_as_cfo(client: TestClient, db_session: Session, email: str) -> None:
    _login(client, email)
    user = db_session.scalar(select(User).where(User.email == email))
    assert user is not None
    user.role = Role.CFO
    user.is_active = True
    db_session.flush()


def _login_as_pending(client: TestClient, email: str) -> None:
    _login(client, email)


# ---------------------------------------------------------------------------
# XLSX builders
# ---------------------------------------------------------------------------


def _make_tally_xlsx(
    data_rows: list[list[Any]] | None = None,
    sheet_name: str = "Sundry Debtors",
) -> bytes:
    _meta = [
        ["Group :", "Sundry Debtors", None, "1-Apr-26 to 16-Apr-26", None, None, None],
        ["Details of:", "Pending Bills", None, None, None, None, None],
        [None] * 7,
        ["Date", "Ref. No.", "Party's Name", "Opening", "Pending", "Due on", "Overdue"],
        [None, None, None, "Amount", "Amount", None, "by days"],
    ]
    wb = openpyxl.Workbook()
    del wb["Sheet"]
    ws = wb.create_sheet(sheet_name)
    for row in _meta:
        ws.append(row)
    for row in data_rows or []:
        inv_date, ref_no, party_name, opening, pending, due_on, overdue = row
        if party_name is not None:
            ws.append([None, None, party_name, None, None, None, None])
        ws.append([inv_date, ref_no, None, opening, pending, due_on, overdue])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_XERO_HEADER_ROW: list[Any] = [
    "Contact Account Number",
    "Primary Person",
    "Phone",
    "Email",
    "Mobile",
    "Contact Group",
    "Invoice Date",
    "Due Date",
    "Expected Date",
    "Invoice Number",
    "Invoice Reference",
    "< 1 Month",
    "1 Month",
    "2 Months",
    "3 Months",
    "Older",
    None,
    "Total",
    "Outstanding Tax",
    "PROJECT ID",
    "SERVICE MONTH",
    "Invoice Seen",
    "Invoice Sent",
]


def _make_xero_xlsx(
    as_of_str: str = "As at 31 March 2026",
    party: str = "XeroParty Ltd",
    inv_date: Any = date(2026, 1, 15),
    inv_num: str = "XERO-001",
    total: float = 2000.0,
    sheet_name: str = "Aged Receivables Detail",
) -> bytes:
    wb = openpyxl.Workbook()
    del wb["Sheet"]
    ws = wb.create_sheet(sheet_name)
    ws.append(["Aged Receivables Detail"] + [None] * 22)
    ws.append(["TEST COMPANY LLC"] + [None] * 22)
    ws.append([as_of_str] + [None] * 22)
    ws.append(["Ageing by due date"] + [None] * 22)
    ws.append([None] * 23)
    ws.append(_XERO_HEADER_ROW)
    ws.append([None] * 23)
    party_header = [None] * 23
    party_header[0] = party
    ws.append(party_header)
    inv_row: list[Any] = [None] * 23
    inv_row[6] = inv_date
    inv_row[9] = inv_num
    inv_row[15] = total
    inv_row[17] = total
    inv_row[18] = 0
    inv_row[21] = "Seen"
    inv_row[22] = "Sent"
    ws.append(inv_row)
    total_row: list[Any] = [None] * 23
    total_row[0] = "Total"
    total_row[17] = total
    ws.append(total_row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Upload + staging helpers
# ---------------------------------------------------------------------------


def _upload(
    client: TestClient,
    file_bytes: bytes,
    entity_code: str = "IND",
    source_hint: str | None = None,
    as_of_date: str | None = None,
    filename: str = "test.xlsx",
) -> Any:
    data: dict[str, Any] = {"entity_code": entity_code}
    if source_hint:
        data["source_hint"] = source_hint
    if as_of_date:
        data["as_of_date"] = as_of_date
    csrf_token = _csrf(client)
    headers: dict[str, str] = {}
    if csrf_token:
        headers["X-CSRF-Token"] = csrf_token
    return client.post(
        "/snapshots",
        data=data,
        files={
            "file": (
                filename,
                io.BytesIO(file_bytes),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=headers,
    )


def _publish(
    client: TestClient,
    snapshot_id: str,
    override_reason: str | None = None,
) -> Any:
    csrf_token = _csrf(client)
    headers: dict[str, str] = {}
    if csrf_token:
        headers["X-CSRF-Token"] = csrf_token
    body: dict[str, Any] = {}
    if override_reason:
        body["override_reason"] = override_reason
    return client.post(
        f"/snapshots/{snapshot_id}/publish",
        json=body,
        headers=headers,
    )


def _set_entity_default_credit_days(
    db_session: Session, entity_code: str, days: int | None
) -> None:
    entity = db_session.scalar(select(Entity).where(Entity.code == entity_code))
    assert entity is not None
    entity.default_credit_days = days
    db_session.flush()


def _create_canonical_for_party(
    db_session: Session,
    entity_code: str,
    canonical_name: str,
    alias_text: str | None = None,
    admin_user_id: uuid.UUID | None = None,
) -> uuid.UUID:
    """Create a canonical party + optional exact alias for a given entity."""
    entity = db_session.scalar(select(Entity).where(Entity.code == entity_code))
    assert entity is not None
    admin = db_session.scalar(select(User).where(User.email == "tejaswa.sharma@emb.global"))
    assert admin is not None

    canonical = PartyCanonical(
        entity_id=entity.id,
        name=canonical_name,
        created_by=admin.id,
    )
    db_session.add(canonical)
    db_session.flush()

    if alias_text:
        alias = PartyAlias(
            canonical_id=canonical.id,
            alias_text=alias_text,
            source="MANUAL",
            confidence=None,
            created_by=admin.id,
        )
        db_session.add(alias)
        db_session.flush()

    return canonical.id


def _resolve_all_rows_via_canonical(
    client: TestClient,
    db_session: Session,
    snapshot_id: str,
    canonical_id: uuid.UUID,
) -> None:
    """Resolve all unmapped invoice rows in a snapshot to the given canonical."""
    csrf_token = _csrf(client)
    headers: dict[str, str] = {"X-CSRF-Token": csrf_token} if csrf_token else {}

    staging_resp = client.get(f"/snapshots/{snapshot_id}/staging")
    assert staging_resp.status_code == 200
    data = staging_resp.json()

    for row in data["rows"]:
        row_index = row["row_index"]
        if row.get("status") == "OK":
            resp = client.patch(
                f"/snapshots/{snapshot_id}/staging/{row_index}",
                json={"action": "resolve_alias", "canonical_id": str(canonical_id)},
                headers=headers,
            )
            assert resp.status_code == 200


def _ack_all_warnings(client: TestClient, db_session: Session, snapshot_id: str) -> list[str]:
    """Acknowledge every warning on the snapshot so the publish gate passes.

    Tally parser always emits UNALLOCATED_CREDITS_DELTA per ADR-0003 addendum;
    without an explicit ack, the publish gate blocks.  Tests that exercise the
    happy path must ack warnings first.
    """
    snap = db_session.scalar(select(Snapshot).where(Snapshot.id == uuid.UUID(snapshot_id)))
    assert snap is not None
    pr = snap.parse_result_json or {}
    codes = sorted({w.get("code") for w in pr.get("warnings", []) if w.get("code")})
    if not codes:
        return []
    csrf_token = _csrf(client)
    headers: dict[str, str] = {"X-CSRF-Token": csrf_token} if csrf_token else {}
    resp = client.patch(
        f"/snapshots/{snapshot_id}/warnings/ack",
        json={"codes": codes},
        headers=headers,
    )
    assert resp.status_code == 200, resp.json()
    return codes


def _setup_publishable_tally_snapshot(
    client: TestClient,
    db_session: Session,
    party_name: str = "AlphaClient Ltd",
    inv_date: date = date(2026, 2, 1),
    inv_ref: str = "INV-001",
    amount: float = 5000.0,
    as_of_date: str = "2026-03-31",
    entity_code: str = "IND",
    default_credit_days: int = 30,
) -> tuple[str, uuid.UUID]:
    """Upload + resolve a single-invoice TALLY snapshot. Returns (snapshot_id, canonical_id)."""
    _set_entity_default_credit_days(db_session, entity_code, default_credit_days)

    xlsx = _make_tally_xlsx(data_rows=[[inv_date, inv_ref, party_name, amount, amount, None, None]])
    upload_resp = _upload(
        client,
        xlsx,
        entity_code=entity_code,
        source_hint="TALLY",
        as_of_date=as_of_date,
    )
    assert upload_resp.status_code == 201, upload_resp.json()
    snapshot_id = upload_resp.json()["snapshot_id"]

    canonical_id = _create_canonical_for_party(
        db_session, entity_code, party_name, alias_text=party_name
    )
    # alias is EXACT, no need to patch staging — it auto-resolves.
    # But Tally always emits UNALLOCATED_CREDITS_DELTA warning — ack it so publish gate passes.
    _ack_all_warnings(client, db_session, snapshot_id)
    return snapshot_id, canonical_id


# ---------------------------------------------------------------------------
# Test 1: TALLY happy path
# ---------------------------------------------------------------------------


def test_publish_tally_happy_path_201_creates_invoices_and_snapshots(
    client: TestClient, db_session: Session
) -> None:
    _login_as_admin(client)
    snapshot_id, canonical_id = _setup_publishable_tally_snapshot(
        client,
        db_session,
        party_name="AlphaClient Ltd",
        inv_date=date(2026, 2, 1),
        inv_ref="INV-TALLY-001",
        amount=5000.0,
        as_of_date="2026-03-31",
    )

    resp = _publish(client, snapshot_id)
    assert resp.status_code == 200, resp.json()
    body = resp.json()

    assert body["status"] == "PUBLISHED"
    assert body["result"]["invoices_inserted"] == 1
    assert body["result"]["invoices_updated"] == 0
    assert body["result"]["invoices_settled"] == 0
    assert body["result"]["invoice_snapshots_written"] == 1
    assert body["result"]["publish_notif_enqueued"] is True

    # Verify DB state
    snap = db_session.scalar(select(Snapshot).where(Snapshot.id == uuid.UUID(snapshot_id)))
    assert snap is not None
    assert snap.status == "PUBLISHED"
    assert snap.published_at is not None

    invoice = db_session.scalar(select(Invoice).where(Invoice.invoice_ref == "INV-TALLY-001"))
    assert invoice is not None
    assert invoice.status == "OPEN"
    assert invoice.currency == "INR"
    assert invoice.amount == Decimal("5000.00")

    inv_snap = db_session.scalar(
        select(InvoiceSnapshot).where(InvoiceSnapshot.invoice_id == invoice.id)
    )
    assert inv_snap is not None
    assert inv_snap.as_of_date == date(2026, 3, 31)


# ---------------------------------------------------------------------------
# Test 2: XERO happy path — writes xero_metadata
# ---------------------------------------------------------------------------


def test_publish_xero_happy_path_writes_xero_metadata_to_invoices(
    client: TestClient, db_session: Session
) -> None:
    _login_as_admin(client)
    _set_entity_default_credit_days(db_session, "UAE", 30)

    xlsx = _make_xero_xlsx(
        as_of_str="As at 31 March 2026",
        party="XeroClientUAE",
        inv_date=date(2026, 1, 20),
        inv_num="XERO-UAE-001",
        total=8000.0,
    )
    upload_resp = _upload(
        client,
        xlsx,
        entity_code="UAE",
        source_hint="XERO",
        filename="xero_test.xlsx",
    )
    assert upload_resp.status_code == 201, upload_resp.json()
    snapshot_id = upload_resp.json()["snapshot_id"]

    # Create exact-match canonical
    _create_canonical_for_party(db_session, "UAE", "XeroClientUAE", alias_text="XeroClientUAE")

    resp = _publish(client, snapshot_id)
    assert resp.status_code == 200, resp.json()
    body = resp.json()
    assert body["result"]["invoices_inserted"] == 1

    invoice = db_session.scalar(select(Invoice).where(Invoice.invoice_ref == "XERO-UAE-001"))
    assert invoice is not None
    assert invoice.currency == "AED"
    # Xero parser sets xero_metadata — should be non-null
    assert invoice.xero_metadata is not None


# ---------------------------------------------------------------------------
# Test 3: Publish blocked — unmapped parties → 422 with gate detail
# ---------------------------------------------------------------------------


def test_publish_blocked_when_unmapped_parties_returns_422_with_gate_detail(
    client: TestClient, db_session: Session
) -> None:
    _login_as_admin(client)
    _set_entity_default_credit_days(db_session, "IND", 30)

    xlsx = _make_tally_xlsx(
        data_rows=[
            [date(2026, 2, 1), "INV-UNMAPPED", "UnknownParty XYZ", 1000.0, 1000.0, None, None]
        ]
    )
    upload_resp = _upload(
        client, xlsx, entity_code="IND", source_hint="TALLY", as_of_date="2026-03-31"
    )
    assert upload_resp.status_code == 201
    snapshot_id = upload_resp.json()["snapshot_id"]

    # NO canonical created → party is unmapped
    resp = _publish(client, snapshot_id)
    assert resp.status_code == 422
    detail = resp.json()["detail"]
    assert detail["code"] == "PUBLISH_GATE_BLOCKED"
    assert detail["publish_gate"]["unmapped_parties_count"] > 0


# ---------------------------------------------------------------------------
# Test 4: Publish blocked — warnings unacknowledged → 422
# ---------------------------------------------------------------------------


def test_publish_blocked_when_warnings_unacknowledged_returns_422(
    client: TestClient, db_session: Session
) -> None:
    """Snapshot with a warning that hasn't been acknowledged should block publish."""
    _login_as_admin(client)
    _set_entity_default_credit_days(db_session, "IND", 30)

    xlsx = _make_tally_xlsx(
        data_rows=[[date(2026, 2, 1), "INV-WARN", "WarnParty", 1000.0, 1000.0, None, None]]
    )
    upload_resp = _upload(
        client, xlsx, entity_code="IND", source_hint="TALLY", as_of_date="2026-03-31"
    )
    assert upload_resp.status_code == 201
    snapshot_id = upload_resp.json()["snapshot_id"]

    # Inject a warning into parse_result_json
    snap = db_session.scalar(select(Snapshot).where(Snapshot.id == uuid.UUID(snapshot_id)))
    assert snap is not None
    pr = dict(snap.parse_result_json or {})
    pr["warnings"] = [{"code": "GRAND_TOTAL_MISMATCH", "message": "test warning"}]
    snap.parse_result_json = pr
    db_session.flush()

    # Create exact-match canonical so party is not unmapped
    _create_canonical_for_party(db_session, "IND", "WarnParty", alias_text="WarnParty")

    resp = _publish(client, snapshot_id)
    assert resp.status_code == 422
    detail = resp.json()["detail"]
    assert detail["code"] == "PUBLISH_GATE_BLOCKED"
    assert "GRAND_TOTAL_MISMATCH" in detail["publish_gate"]["warnings_unacknowledged"]


# ---------------------------------------------------------------------------
# Test 5: Publish blocked — parse errors unresolved → 422
# ---------------------------------------------------------------------------


def test_publish_blocked_when_parse_errors_unresolved_returns_422(
    client: TestClient, db_session: Session
) -> None:
    _login_as_admin(client)
    _set_entity_default_credit_days(db_session, "IND", 30)

    xlsx = _make_tally_xlsx(
        data_rows=[[date(2026, 2, 1), "INV-PE", "ParseErrParty", 1000.0, 1000.0, None, None]]
    )
    upload_resp = _upload(
        client, xlsx, entity_code="IND", source_hint="TALLY", as_of_date="2026-03-31"
    )
    assert upload_resp.status_code == 201
    snapshot_id = upload_resp.json()["snapshot_id"]

    # Inject a PARSE_ERROR row
    snap = db_session.scalar(select(Snapshot).where(Snapshot.id == uuid.UUID(snapshot_id)))
    assert snap is not None
    pr = dict(snap.parse_result_json or {})
    existing_invoices = pr.get("invoices", [])
    existing_invoices.append(
        {
            "row_index": 99,
            "status": "PARSE_ERROR",
            "party_name_raw": "ParseErrParty",
            "invoice_ref": None,
            "invoice_date": None,
            "amount": None,
            "source_currency": "INR",
            "parse_error_reason": "Missing invoice date",
            "raw_row_json": {},
        }
    )
    pr["invoices"] = existing_invoices
    snap.parse_result_json = pr
    db_session.flush()

    _create_canonical_for_party(db_session, "IND", "ParseErrParty", alias_text="ParseErrParty")

    resp = _publish(client, snapshot_id)
    assert resp.status_code == 422
    detail = resp.json()["detail"]
    assert detail["code"] == "PUBLISH_GATE_BLOCKED"
    assert detail["publish_gate"]["parse_errors_unresolved_count"] > 0


# ---------------------------------------------------------------------------
# Tests 6-8: RBAC negatives
# ---------------------------------------------------------------------------


def test_publish_cfo_403(client: TestClient, db_session: Session) -> None:
    _login_as_admin(client)
    snapshot_id, _ = _setup_publishable_tally_snapshot(client, db_session)

    _login_as_cfo(client, db_session, "cfo@emb.global")
    resp = _publish(client, snapshot_id)
    assert resp.status_code == 403


def test_publish_pending_403(client: TestClient, db_session: Session) -> None:
    _login_as_admin(client)
    snapshot_id, _ = _setup_publishable_tally_snapshot(client, db_session)

    _login_as_pending(client, "pending@emb.global")
    resp = _publish(client, snapshot_id)
    assert resp.status_code == 403


def test_publish_analyst_wrong_entity_403(client: TestClient, db_session: Session) -> None:
    _login_as_admin(client)
    snapshot_id, _ = _setup_publishable_tally_snapshot(client, db_session, entity_code="IND")

    # Analyst scoped to UAE, trying to publish IND snapshot
    _login_as_analyst(client, db_session, "analyst.uae@emb.global", entity_code="UAE")
    resp = _publish(client, snapshot_id)
    assert resp.status_code == 403


# ---------------------------------------------------------------------------
# Test 9: Admin publishing another analyst's snapshot → published_as=OVERRIDE
# ---------------------------------------------------------------------------


def test_publish_admin_other_entity_sets_published_as_override(
    client: TestClient, db_session: Session
) -> None:
    # Upload as an analyst
    _login_as_analyst(client, db_session, "analyst.ind@emb.global", entity_code="IND")
    _set_entity_default_credit_days(db_session, "IND", 30)

    xlsx = _make_tally_xlsx(
        data_rows=[[date(2026, 2, 1), "INV-OVERRIDE", "OverrideParty", 1000.0, 1000.0, None, None]]
    )
    upload_resp = _upload(
        client, xlsx, entity_code="IND", source_hint="TALLY", as_of_date="2026-03-31"
    )
    assert upload_resp.status_code == 201
    snapshot_id = upload_resp.json()["snapshot_id"]

    _create_canonical_for_party(db_session, "IND", "OverrideParty", alias_text="OverrideParty")

    # Ack warnings while still logged in as analyst (before switching to admin)
    _ack_all_warnings(client, db_session, snapshot_id)

    # Now login as admin (different user) and publish
    _login_as_admin(client)
    resp = _publish(client, snapshot_id, override_reason="Admin override test")
    assert resp.status_code == 200, resp.json()
    body = resp.json()
    assert body["published_as"] == "OVERRIDE"

    snap = db_session.scalar(select(Snapshot).where(Snapshot.id == uuid.UUID(snapshot_id)))
    assert snap is not None
    assert snap.published_as == "OVERRIDE"


# ---------------------------------------------------------------------------
# Test 10: Publish twice → 409
# ---------------------------------------------------------------------------


def test_publish_twice_returns_409(client: TestClient, db_session: Session) -> None:
    _login_as_admin(client)
    snapshot_id, _ = _setup_publishable_tally_snapshot(client, db_session)

    resp1 = _publish(client, snapshot_id)
    assert resp1.status_code == 200

    resp2 = _publish(client, snapshot_id)
    assert resp2.status_code == 409
    assert resp2.json()["detail"]["code"] == "SNAPSHOT_NOT_STAGED"


# ---------------------------------------------------------------------------
# Test 11: Publish DISCARDED snapshot → 409
# ---------------------------------------------------------------------------


def test_publish_discarded_snapshot_returns_409(client: TestClient, db_session: Session) -> None:
    _login_as_admin(client)
    _set_entity_default_credit_days(db_session, "IND", 30)

    xlsx = _make_tally_xlsx(
        data_rows=[[date(2026, 2, 1), "INV-DISC", "DiscardParty", 1000.0, 1000.0, None, None]]
    )
    upload_resp = _upload(
        client, xlsx, entity_code="IND", source_hint="TALLY", as_of_date="2026-03-31"
    )
    assert upload_resp.status_code == 201
    snapshot_id = upload_resp.json()["snapshot_id"]

    # Forcibly set status to DISCARDED
    snap = db_session.scalar(select(Snapshot).where(Snapshot.id == uuid.UUID(snapshot_id)))
    assert snap is not None
    snap.status = "DISCARDED"
    db_session.flush()

    resp = _publish(client, snapshot_id)
    assert resp.status_code == 409
    assert resp.json()["detail"]["code"] == "SNAPSHOT_NOT_STAGED"


# ---------------------------------------------------------------------------
# Test 12: CREDIT_PERIOD → 422 not implemented
# ---------------------------------------------------------------------------


def test_publish_credit_period_returns_422_not_yet_implemented(
    client: TestClient, db_session: Session
) -> None:
    _login_as_admin(client)

    # Create a minimal CP xlsx
    wb = openpyxl.Workbook()
    del wb["Sheet"]
    ws_ind = wb.create_sheet("India")
    ws_ind.append(["Client Name", "Credit Period"])
    ws_ind.append(["SomeClient", 30])
    ws_uae = wb.create_sheet("UAE")
    ws_uae.append(["Client Name", "Credit Period", "Reason for extended Credit", "Amount"])
    ws_uae.append(["BetaClient", 45, "contract", None])
    buf = io.BytesIO()
    wb.save(buf)
    cp_bytes = buf.getvalue()

    upload_resp = _upload(
        client, cp_bytes, entity_code="IND", source_hint="CREDIT_PERIOD", filename="cp.xlsx"
    )
    assert upload_resp.status_code == 201, upload_resp.json()
    snapshot_id = upload_resp.json()["snapshot_id"]

    resp = _publish(client, snapshot_id)
    assert resp.status_code == 422
    detail = resp.json()["detail"]
    assert detail["code"] == "CREDIT_PERIOD_PUBLISH_NOT_IMPLEMENTED_YET_SEE_TASK_6"


# ---------------------------------------------------------------------------
# Test 13: Three-snapshot upsert test (spec §12)
# ---------------------------------------------------------------------------


def test_three_snapshot_upsert_insert_update_settle(
    client: TestClient, db_session: Session
) -> None:
    """Spec §12: insert → update → settle.

    Snapshot 1: 3 invoices (INV-A, INV-B, INV-C) → all inserted.
    Snapshot 2: 2 invoices (INV-A updated amount, INV-B same) → C not present (SETTLED).
    Snapshot 3: 1 invoice (INV-A only) → B also SETTLED.
    """
    _login_as_admin(client)
    _set_entity_default_credit_days(db_session, "IND", 30)

    party = "ThreeSnapshotParty"
    _create_canonical_for_party(db_session, "IND", party, alias_text=party)

    # --- Snapshot 1: insert 3 invoices ---
    xlsx1 = _make_tally_xlsx(
        data_rows=[
            [date(2026, 1, 1), "INV-A", party, 1000.0, 1000.0, None, None],
            [date(2026, 1, 2), "INV-B", party, 2000.0, 2000.0, None, None],
            [date(2026, 1, 3), "INV-C", party, 3000.0, 3000.0, None, None],
        ]
    )
    r1 = _upload(client, xlsx1, entity_code="IND", source_hint="TALLY", as_of_date="2026-02-28")
    assert r1.status_code == 201
    s1_id = r1.json()["snapshot_id"]

    _ack_all_warnings(client, db_session, s1_id)
    pub1 = _publish(client, s1_id)
    assert pub1.status_code == 200, pub1.json()
    r1_body = pub1.json()["result"]
    assert r1_body["invoices_inserted"] == 3
    assert r1_body["invoices_updated"] == 0
    assert r1_body["invoices_settled"] == 0
    assert r1_body["invoice_snapshots_written"] == 3

    # Satisfy §13 #6 gate: prior snapshot must be MATCHED before next publish
    _reconcile_snapshot_directly(db_session, s1_id)

    # --- Snapshot 2: update INV-A, keep INV-B; omit INV-C → SETTLED ---
    xlsx2 = _make_tally_xlsx(
        data_rows=[
            [date(2026, 1, 1), "INV-A", party, 1500.0, 1500.0, None, None],  # updated amount
            [date(2026, 1, 2), "INV-B", party, 2000.0, 2000.0, None, None],  # same
        ]
    )
    r2 = _upload(
        client,
        xlsx2,
        entity_code="IND",
        source_hint="TALLY",
        as_of_date="2026-03-31",
        filename="snap2.xlsx",
    )
    assert r2.status_code == 201
    s2_id = r2.json()["snapshot_id"]

    _ack_all_warnings(client, db_session, s2_id)
    pub2 = _publish(client, s2_id)
    assert pub2.status_code == 200, pub2.json()
    r2_body = pub2.json()["result"]
    assert r2_body["invoices_inserted"] == 0
    assert r2_body["invoices_updated"] == 2
    assert r2_body["invoices_settled"] == 1

    # Satisfy §13 #6 gate before snapshot 3
    _reconcile_snapshot_directly(db_session, s2_id)

    # Verify INV-C is SETTLED with settled_snapshot_id = s2
    inv_c = db_session.scalar(select(Invoice).where(Invoice.invoice_ref == "INV-C"))
    assert inv_c is not None
    assert inv_c.status == "SETTLED"
    assert inv_c.settled_snapshot_id == uuid.UUID(s2_id)

    # Verify INV-A amount updated
    inv_a = db_session.scalar(select(Invoice).where(Invoice.invoice_ref == "INV-A"))
    assert inv_a is not None
    assert inv_a.amount == Decimal("1500.00")

    # --- Snapshot 3: only INV-A → INV-B also SETTLED ---
    xlsx3 = _make_tally_xlsx(
        data_rows=[
            [date(2026, 1, 1), "INV-A", party, 1500.0, 1500.0, None, None],
        ]
    )
    r3 = _upload(
        client,
        xlsx3,
        entity_code="IND",
        source_hint="TALLY",
        as_of_date="2026-04-30",
        filename="snap3.xlsx",
    )
    assert r3.status_code == 201
    s3_id = r3.json()["snapshot_id"]

    _ack_all_warnings(client, db_session, s3_id)
    pub3 = _publish(client, s3_id)
    assert pub3.status_code == 200, pub3.json()
    r3_body = pub3.json()["result"]
    assert r3_body["invoices_settled"] == 1

    inv_b = db_session.scalar(select(Invoice).where(Invoice.invoice_ref == "INV-B"))
    assert inv_b is not None
    assert inv_b.status == "SETTLED"
    assert inv_b.settled_snapshot_id == uuid.UUID(s3_id)


# ---------------------------------------------------------------------------
# Test 14: SETTLED invoice reopens if reappears in new snapshot
# ---------------------------------------------------------------------------


def test_settled_invoice_reopens_if_reappears_in_new_snapshot(
    client: TestClient, db_session: Session
) -> None:
    _login_as_admin(client)
    _set_entity_default_credit_days(db_session, "IND", 30)

    party = "ReopenParty"
    _create_canonical_for_party(db_session, "IND", party, alias_text=party)

    # Snapshot 1: INV-REOPEN present
    xlsx1 = _make_tally_xlsx(
        data_rows=[
            [date(2026, 1, 1), "INV-REOPEN", party, 1000.0, 1000.0, None, None],
        ]
    )
    r1 = _upload(client, xlsx1, entity_code="IND", source_hint="TALLY", as_of_date="2026-02-28")
    s1_id = r1.json()["snapshot_id"]
    _ack_all_warnings(client, db_session, s1_id)
    p1 = _publish(client, s1_id)
    assert p1.status_code == 200
    _reconcile_snapshot_directly(db_session, s1_id)

    # Snapshot 2: empty — settles INV-REOPEN
    xlsx2 = _make_tally_xlsx(
        data_rows=[
            [date(2026, 1, 5), "INV-OTHER", party, 500.0, 500.0, None, None],
        ]
    )
    r2 = _upload(
        client,
        xlsx2,
        entity_code="IND",
        source_hint="TALLY",
        as_of_date="2026-03-31",
        filename="s2.xlsx",
    )
    s2_id = r2.json()["snapshot_id"]
    _ack_all_warnings(client, db_session, s2_id)
    p2 = _publish(client, s2_id)
    assert p2.status_code == 200
    _reconcile_snapshot_directly(db_session, s2_id)

    inv = db_session.scalar(select(Invoice).where(Invoice.invoice_ref == "INV-REOPEN"))
    assert inv is not None
    assert inv.status == "SETTLED"

    # Snapshot 3: INV-REOPEN reappears — should be OPEN again
    xlsx3 = _make_tally_xlsx(
        data_rows=[
            [date(2026, 1, 1), "INV-REOPEN", party, 1000.0, 1000.0, None, None],
            [date(2026, 1, 5), "INV-OTHER", party, 500.0, 500.0, None, None],
        ]
    )
    r3 = _upload(
        client,
        xlsx3,
        entity_code="IND",
        source_hint="TALLY",
        as_of_date="2026-04-30",
        filename="s3.xlsx",
    )
    s3_id = r3.json()["snapshot_id"]
    _ack_all_warnings(client, db_session, s3_id)
    p3 = _publish(client, s3_id)
    assert p3.status_code == 200, p3.json()
    r3_body = p3.json()["result"]
    assert r3_body["invoices_updated"] >= 1  # reopened invoice counted as updated

    db_session.expire(inv)
    db_session.refresh(inv)
    assert inv.status == "OPEN"
    assert inv.settled_snapshot_id is None


# ---------------------------------------------------------------------------
# Test 15: Exception auto-resolved on settled cascade
# ---------------------------------------------------------------------------


def test_exception_auto_resolved_on_settled_cascade(
    client: TestClient, db_session: Session
) -> None:
    _login_as_admin(client)
    _set_entity_default_credit_days(db_session, "IND", 30)

    party = "ExceptionParty"
    _create_canonical_for_party(db_session, "IND", party, alias_text=party)

    # Snapshot 1: INV-EXC present
    xlsx1 = _make_tally_xlsx(
        data_rows=[
            [date(2026, 1, 1), "INV-EXC", party, 1000.0, 1000.0, None, None],
        ]
    )
    r1 = _upload(client, xlsx1, entity_code="IND", source_hint="TALLY", as_of_date="2026-02-28")
    s1_id = r1.json()["snapshot_id"]
    _ack_all_warnings(client, db_session, s1_id)
    p1 = _publish(client, s1_id)
    assert p1.status_code == 200
    _reconcile_snapshot_directly(db_session, s1_id)

    # Fetch the invoice and add an ACTIVE exception tag
    inv = db_session.scalar(select(Invoice).where(Invoice.invoice_ref == "INV-EXC"))
    assert inv is not None
    admin = db_session.scalar(select(User).where(User.email == "tejaswa.sharma@emb.global"))
    bucket_type = db_session.scalar(select(ExceptionBucketType).limit(1))
    assert bucket_type is not None

    tag = ExceptionTag(
        invoice_id=inv.id,
        bucket_type_id=bucket_type.id,
        reason="Test exception for auto-resolve",
        tagged_by=admin.id,
        status="ACTIVE",
    )
    db_session.add(tag)
    db_session.flush()

    # Snapshot 2: INV-EXC absent → SETTLED → exception AUTO_RESOLVED
    xlsx2 = _make_tally_xlsx(
        data_rows=[
            [date(2026, 2, 1), "INV-KEEPALIVE", party, 500.0, 500.0, None, None],
        ]
    )
    r2 = _upload(
        client,
        xlsx2,
        entity_code="IND",
        source_hint="TALLY",
        as_of_date="2026-03-31",
        filename="s2.xlsx",
    )
    s2_id = r2.json()["snapshot_id"]
    _ack_all_warnings(client, db_session, s2_id)
    p2 = _publish(client, s2_id)
    assert p2.status_code == 200, p2.json()
    assert p2.json()["result"]["exceptions_auto_resolved"] == 1
    assert p2.json()["result"]["invoices_settled"] == 1

    db_session.expire(tag)
    db_session.refresh(tag)
    assert tag.status == "AUTO_RESOLVED"
    assert tag.resolved_at is not None
    assert "settled in snapshot" in (tag.resolution_note or "")


# ---------------------------------------------------------------------------
# Test 16: Material-change flagged when amount delta >5% on active exception
# ---------------------------------------------------------------------------


def test_material_change_flagged_when_amount_delta_gt_5_percent_on_active_exception(
    client: TestClient, db_session: Session
) -> None:
    _login_as_admin(client)
    _set_entity_default_credit_days(db_session, "IND", 30)

    party = "MaterialParty"
    _create_canonical_for_party(db_session, "IND", party, alias_text=party)

    # Snapshot 1: INV-MAT at 1000
    xlsx1 = _make_tally_xlsx(
        data_rows=[
            [date(2026, 1, 1), "INV-MAT", party, 1000.0, 1000.0, None, None],
        ]
    )
    r1 = _upload(client, xlsx1, entity_code="IND", source_hint="TALLY", as_of_date="2026-02-28")
    s1_id = r1.json()["snapshot_id"]
    _ack_all_warnings(client, db_session, s1_id)
    p1 = _publish(client, s1_id)
    assert p1.status_code == 200
    _reconcile_snapshot_directly(db_session, s1_id)

    # Add ACTIVE exception tag on INV-MAT
    inv = db_session.scalar(select(Invoice).where(Invoice.invoice_ref == "INV-MAT"))
    assert inv is not None
    admin = db_session.scalar(select(User).where(User.email == "tejaswa.sharma@emb.global"))
    bucket_type = db_session.scalar(select(ExceptionBucketType).limit(1))
    assert bucket_type is not None

    tag = ExceptionTag(
        invoice_id=inv.id,
        bucket_type_id=bucket_type.id,
        reason="Material change test",
        tagged_by=admin.id,
        status="ACTIVE",
    )
    db_session.add(tag)
    db_session.flush()

    # Snapshot 2: INV-MAT at 1100 (10% increase — above 5% threshold)
    xlsx2 = _make_tally_xlsx(
        data_rows=[
            [date(2026, 1, 1), "INV-MAT", party, 1100.0, 1100.0, None, None],
        ]
    )
    r2 = _upload(
        client,
        xlsx2,
        entity_code="IND",
        source_hint="TALLY",
        as_of_date="2026-03-31",
        filename="s2.xlsx",
    )
    s2_id = r2.json()["snapshot_id"]
    _ack_all_warnings(client, db_session, s2_id)
    p2 = _publish(client, s2_id)
    assert p2.status_code == 200, p2.json()
    assert p2.json()["result"]["exceptions_material_change_flagged"] == 1

    # Verify flag stored on snapshot
    snap = db_session.scalar(select(Snapshot).where(Snapshot.id == uuid.UUID(s2_id)))
    assert snap is not None
    flags = snap.material_change_flags_json
    assert len(flags) == 1
    assert flags[0]["invoice_id"] == str(inv.id)
    assert Decimal(flags[0]["delta_pct"]) > Decimal("5")


# ---------------------------------------------------------------------------
# Test 17: invoice_snapshots row written with correct bucket and overdue_days
# ---------------------------------------------------------------------------


def test_invoice_snapshots_row_written_with_correct_bucket_and_overdue_days(
    client: TestClient, db_session: Session
) -> None:
    """Use compute_ageing ground truth to verify bucket assignment."""
    from app.services.ageing import AgeingBucket, compute_ageing

    _login_as_admin(client)
    _set_entity_default_credit_days(db_session, "IND", 30)

    party = "BucketParty"
    _create_canonical_for_party(db_session, "IND", party, alias_text=party)

    inv_date = date(2026, 1, 1)
    as_of = date(2026, 3, 31)
    # With 30 credit days, due = 2026-01-31, overdue_days = (2026-03-31 - 2026-01-31).days = 59
    expected = compute_ageing(inv_date, 30, as_of)
    assert expected.bucket == AgeingBucket.BUCKET_31_60

    xlsx = _make_tally_xlsx(
        data_rows=[
            [inv_date, "INV-BUCKET", party, 1000.0, 1000.0, None, None],
        ]
    )
    r = _upload(client, xlsx, entity_code="IND", source_hint="TALLY", as_of_date="2026-03-31")
    snap_id = r.json()["snapshot_id"]
    _ack_all_warnings(client, db_session, snap_id)
    pub = _publish(client, snap_id)
    assert pub.status_code == 200, pub.json()

    inv = db_session.scalar(select(Invoice).where(Invoice.invoice_ref == "INV-BUCKET"))
    assert inv is not None
    inv_snap = db_session.scalar(
        select(InvoiceSnapshot).where(
            InvoiceSnapshot.invoice_id == inv.id,
            InvoiceSnapshot.snapshot_id == uuid.UUID(snap_id),
        )
    )
    assert inv_snap is not None
    assert inv_snap.overdue_days == expected.overdue_days
    assert inv_snap.bucket == expected.bucket.value


# ---------------------------------------------------------------------------
# Test 18: Settled invoices do NOT get invoice_snapshot row this publish
# ---------------------------------------------------------------------------


def test_settled_invoices_do_not_get_invoice_snapshot_row_this_publish(
    client: TestClient, db_session: Session
) -> None:
    _login_as_admin(client)
    _set_entity_default_credit_days(db_session, "IND", 30)

    party = "SettleSnapParty"
    _create_canonical_for_party(db_session, "IND", party, alias_text=party)

    # Snap 1: 2 invoices
    xlsx1 = _make_tally_xlsx(
        data_rows=[
            [date(2026, 1, 1), "INV-SNAP1", party, 1000.0, 1000.0, None, None],
            [date(2026, 1, 2), "INV-SNAP2", party, 2000.0, 2000.0, None, None],
        ]
    )
    r1 = _upload(client, xlsx1, entity_code="IND", source_hint="TALLY", as_of_date="2026-02-28")
    s1_id = r1.json()["snapshot_id"]
    _ack_all_warnings(client, db_session, s1_id)
    p1 = _publish(client, s1_id)
    assert p1.status_code == 200
    assert p1.json()["result"]["invoice_snapshots_written"] == 2
    _reconcile_snapshot_directly(db_session, s1_id)

    # Snap 2: only INV-SNAP1 → INV-SNAP2 settled
    xlsx2 = _make_tally_xlsx(
        data_rows=[
            [date(2026, 1, 1), "INV-SNAP1", party, 1000.0, 1000.0, None, None],
        ]
    )
    r2 = _upload(
        client,
        xlsx2,
        entity_code="IND",
        source_hint="TALLY",
        as_of_date="2026-03-31",
        filename="s2.xlsx",
    )
    s2_id = r2.json()["snapshot_id"]
    _ack_all_warnings(client, db_session, s2_id)
    p2 = _publish(client, s2_id)
    assert p2.status_code == 200, p2.json()
    # Only 1 invoice snapshot written (not 2)
    assert p2.json()["result"]["invoice_snapshots_written"] == 1
    assert p2.json()["result"]["invoices_settled"] == 1

    inv2 = db_session.scalar(select(Invoice).where(Invoice.invoice_ref == "INV-SNAP2"))
    assert inv2 is not None
    # No invoice_snapshot row for INV-SNAP2 in snapshot s2
    snap_row_count = (
        db_session.query(InvoiceSnapshot)
        .filter(
            InvoiceSnapshot.invoice_id == inv2.id,
            InvoiceSnapshot.snapshot_id == uuid.UUID(s2_id),
        )
        .count()
    )
    assert snap_row_count == 0


# ---------------------------------------------------------------------------
# Test 19: audit_log written on publish
# ---------------------------------------------------------------------------


def test_audit_log_written_on_publish(client: TestClient, db_session: Session) -> None:
    _login_as_admin(client)
    snapshot_id, _ = _setup_publishable_tally_snapshot(client, db_session, inv_ref="INV-AUDIT")

    resp = _publish(client, snapshot_id)
    assert resp.status_code == 200

    audit = db_session.scalar(
        select(AuditLog).where(
            AuditLog.action == "snapshot.publish",
            AuditLog.entity_id == uuid.UUID(snapshot_id),
        )
    )
    assert audit is not None
    assert audit.before == {"status": "STAGED"}
    assert audit.after["status"] == "PUBLISHED"
    assert "result" in audit.after
    assert audit.after["result"]["invoices_inserted"] == 1


# ---------------------------------------------------------------------------
# Test 20: publish_notif row in email_outbox after publish
# ---------------------------------------------------------------------------


def test_publish_notif_row_in_email_outbox_after_publish(
    client: TestClient, db_session: Session
) -> None:
    _login_as_admin(client)
    snapshot_id, _ = _setup_publishable_tally_snapshot(client, db_session, inv_ref="INV-EMAIL")

    resp = _publish(client, snapshot_id)
    assert resp.status_code == 200

    outbox = db_session.scalar(
        select(EmailOutbox).where(
            EmailOutbox.snapshot_id == uuid.UUID(snapshot_id),
            EmailOutbox.rule_type == "PUBLISH_NOTIF",
        )
    )
    assert outbox is not None
    assert outbox.status == "QUEUED"
    assert "PUBLISHED" in outbox.subject.upper() or "published" in outbox.subject.lower()
    assert outbox.recipients_json == []  # M6 populates these


# ---------------------------------------------------------------------------
# Test 21: publish_gate recomputed fresh (not trusting stale state)
# ---------------------------------------------------------------------------


def test_publish_gate_recomputed_fresh(client: TestClient, db_session: Session) -> None:
    """Mutate staging between GET and publish; gate must use fresh state."""
    _login_as_admin(client)
    _set_entity_default_credit_days(db_session, "IND", 30)

    xlsx = _make_tally_xlsx(
        data_rows=[[date(2026, 2, 1), "INV-FRESH", "FreshParty", 1000.0, 1000.0, None, None]]
    )
    upload_resp = _upload(
        client, xlsx, entity_code="IND", source_hint="TALLY", as_of_date="2026-03-31"
    )
    assert upload_resp.status_code == 201
    snapshot_id = upload_resp.json()["snapshot_id"]

    # First: publish WITHOUT canonical — should be blocked
    resp = _publish(client, snapshot_id)
    assert resp.status_code == 422
    assert resp.json()["detail"]["code"] == "PUBLISH_GATE_BLOCKED"

    # Now create exact-match canonical → gate should pass
    _create_canonical_for_party(db_session, "IND", "FreshParty", alias_text="FreshParty")

    # Ack warnings now that canonical exists (gate was blocked on unmapped parties, not warnings)
    _ack_all_warnings(client, db_session, snapshot_id)

    resp2 = _publish(client, snapshot_id)
    assert resp2.status_code == 200, resp2.json()


# ---------------------------------------------------------------------------
# Test 22: credit_days_source MANUAL override wins over CONFIG
# ---------------------------------------------------------------------------


def test_credit_days_source_manual_override_wins_over_config(
    client: TestClient, db_session: Session
) -> None:
    """D8: MANUAL wins even when CONFIG row exists."""
    from app.db.models.credit_period_config import CreditPeriodConfig
    from app.db.models.user import User

    _login_as_admin(client)
    _set_entity_default_credit_days(db_session, "IND", 60)  # DEFAULT

    party = "ManualCreditParty"
    canonical_id = _create_canonical_for_party(db_session, "IND", party, alias_text=party)

    # Add CONFIG row (30 days)
    admin = db_session.scalar(select(User).where(User.email == "tejaswa.sharma@emb.global"))
    assert admin is not None
    config_row = CreditPeriodConfig(
        canonical_id=canonical_id,
        days=30,
        valid_from=date(2026, 1, 1),
        valid_to=None,
        updated_by=admin.id,
    )
    db_session.add(config_row)
    db_session.flush()

    xlsx = _make_tally_xlsx(
        data_rows=[
            [date(2026, 1, 1), "INV-MANUAL-CD", party, 1000.0, 1000.0, None, None],
        ]
    )
    r = _upload(client, xlsx, entity_code="IND", source_hint="TALLY", as_of_date="2026-03-31")
    snapshot_id = r.json()["snapshot_id"]

    # Apply MANUAL override_credit_days = 7 via staging
    csrf_token = _csrf(client)
    headers: dict[str, str] = {"X-CSRF-Token": csrf_token} if csrf_token else {}
    staging = client.get(f"/snapshots/{snapshot_id}/staging").json()
    row_index = staging["rows"][0]["row_index"]
    patch_resp = client.patch(
        f"/snapshots/{snapshot_id}/staging/{row_index}",
        json={"action": "override_credit_days", "credit_days": 7, "reason": "test"},
        headers=headers,
    )
    assert patch_resp.status_code == 200

    _ack_all_warnings(client, db_session, snapshot_id)
    pub = _publish(client, snapshot_id)
    assert pub.status_code == 200, pub.json()

    inv = db_session.scalar(select(Invoice).where(Invoice.invoice_ref == "INV-MANUAL-CD"))
    assert inv is not None
    assert inv.credit_days_applied == 7
    assert inv.credit_days_source == "MANUAL"


# ---------------------------------------------------------------------------
# Test 23: credit_days_source entity DEFAULT used when no config row
# ---------------------------------------------------------------------------


def test_credit_days_source_entity_default_used_when_no_config_row(
    client: TestClient, db_session: Session
) -> None:
    _login_as_admin(client)
    _set_entity_default_credit_days(db_session, "IND", 45)

    party = "DefaultCreditParty"
    _create_canonical_for_party(db_session, "IND", party, alias_text=party)

    xlsx = _make_tally_xlsx(
        data_rows=[
            [date(2026, 1, 1), "INV-DEFAULT-CD", party, 1000.0, 1000.0, None, None],
        ]
    )
    r = _upload(client, xlsx, entity_code="IND", source_hint="TALLY", as_of_date="2026-03-31")
    snapshot_id = r.json()["snapshot_id"]
    _ack_all_warnings(client, db_session, snapshot_id)
    pub = _publish(client, snapshot_id)
    assert pub.status_code == 200, pub.json()

    inv = db_session.scalar(select(Invoice).where(Invoice.invoice_ref == "INV-DEFAULT-CD"))
    assert inv is not None
    assert inv.credit_days_applied == 45
    assert inv.credit_days_source == "DEFAULT"


# ---------------------------------------------------------------------------
# Test 24: entity default_credit_days NULL → 422
# ---------------------------------------------------------------------------


def test_entity_default_credit_days_null_returns_422(
    client: TestClient, db_session: Session
) -> None:
    _login_as_admin(client)
    _set_entity_default_credit_days(db_session, "IND", None)  # NULL

    party = "NullCreditParty"
    _create_canonical_for_party(db_session, "IND", party, alias_text=party)

    xlsx = _make_tally_xlsx(
        data_rows=[
            [date(2026, 1, 1), "INV-NULL-CD", party, 1000.0, 1000.0, None, None],
        ]
    )
    r = _upload(client, xlsx, entity_code="IND", source_hint="TALLY", as_of_date="2026-03-31")
    snapshot_id = r.json()["snapshot_id"]
    _ack_all_warnings(client, db_session, snapshot_id)
    pub = _publish(client, snapshot_id)
    assert pub.status_code == 422
    detail = pub.json()["detail"]
    assert detail["code"] == "CREDIT_DAYS_UNRESOLVABLE"


# ---------------------------------------------------------------------------
# Test 25: Concurrent publish — row lock serialises; one wins, one 409
# ---------------------------------------------------------------------------


@pytest.mark.xfail(
    reason=(
        "FOLLOW-UP: Threading against the shared test_engine hits SQLAlchemy "
        "state-change errors under concurrent transactional DDL. The production "
        "SELECT FOR UPDATE lock is exercised in publish_service.publish_snapshot "
        "and sequentially covered by test_publish_twice_returns_409. A proper "
        "concurrency test needs per-thread engine instances against a non-"
        "branched Postgres, not the session-scoped Neon branch. "
        "Track in M7 hardening (RBAC suite + race tests)."
    ),
    strict=False,
)
def test_concurrent_publish_serialised_via_row_lock(
    client: TestClient, db_session: Session, test_engine: Any
) -> None:
    """Two concurrent publish calls on same snapshot; exactly one wins (200).

    Each thread gets a FRESH `Session` bound to the same `test_engine` (Neon
    test branch). Using the shared `db_session` across threads would hit
    "Session is already flushing" because SQLAlchemy Sessions are not
    thread-safe — the correctness signal we want (row-level lock serialisation
    at the DB) is independent of the session object and is surfaced per-thread
    as 200 vs 409.
    """
    from sqlalchemy.orm import sessionmaker

    _login_as_admin(client)
    snapshot_id, _ = _setup_publishable_tally_snapshot(client, db_session, inv_ref="INV-CONCURRENT")
    # Commit so the snapshot + canonical are visible to the per-thread sessions.
    db_session.commit()

    SessionLocal = sessionmaker(bind=test_engine, expire_on_commit=False)
    results: list[int] = []

    def do_publish() -> None:
        from fastapi.testclient import TestClient

        from app.api.deps import db_session as db_dep
        from app.main import app

        thread_session = SessionLocal()

        def override():
            try:
                yield thread_session
            finally:
                thread_session.close()

        app.dependency_overrides[db_dep] = override
        try:
            with TestClient(app, cookies=dict(client.cookies)) as tc:
                r = _publish(tc, snapshot_id)
                results.append(r.status_code)
        finally:
            app.dependency_overrides.pop(db_dep, None)

    t1 = threading.Thread(target=do_publish)
    t2 = threading.Thread(target=do_publish)
    t1.start()
    t2.start()
    t1.join(timeout=60)
    t2.join(timeout=60)

    # One should succeed (200) and one should fail (409 — already published)
    assert sorted(results) == [200, 409], f"Expected [200, 409], got {sorted(results)}"
