"""Integration tests for GET /snapshots/:id/cp-diff (Task 15).

Covers:
  - Happy path: all 3 diff categories (ADDED, SUPERSEDED, UNCHANGED)
  - 422 when snapshot.source_hint != CREDIT_PERIOD
  - 404 on unknown snapshot_id
  - RBAC matrix:
      ANALYST in scope (IND)  → 200
      ANALYST out of scope    → 403
      CFO                     → 200 (read-only allowed)
      PENDING                 → 403

Strategy:
  - Upload a CP snapshot via POST /snapshots to get a real parsed snapshot.
  - Seed CreditPeriodConfig rows directly into the DB to exercise the
    three-category classification without going through a full publish.
  - Per-test DB rollback via the function-scoped `client` + `db_session`
    fixtures from conftest.py.
"""

from __future__ import annotations

import io
import uuid
from datetime import date
from typing import TYPE_CHECKING, Any, cast

import openpyxl
from sqlalchemy import select

from app.core.rbac import Role
from app.db.models.credit_period_config import CreditPeriodConfig
from app.db.models.entity import Entity
from app.db.models.party import PartyAlias, PartyCanonical
from app.db.models.snapshot import Snapshot
from app.db.models.user import User

if TYPE_CHECKING:
    from fastapi.testclient import TestClient
    from sqlalchemy.orm import Session


# ---------------------------------------------------------------------------
# Auth helpers (mirrors test_staging_api.py pattern)
# ---------------------------------------------------------------------------


def _login(client: TestClient, email: str) -> None:
    client.get(f"/auth/google/callback?stub_email={email}", follow_redirects=False)


def _csrf(client: TestClient) -> str:
    return client.cookies.get("csrf_token") or ""


def _login_as_admin(client: TestClient) -> None:
    _login(client, "tejaswa.sharma@emb.global")


def _login_as_analyst(
    client: TestClient,
    db_session: Session,
    email: str,
    entity_code: str | None = None,
) -> uuid.UUID:
    _login(client, email)
    user = db_session.scalar(select(User).where(User.email == email))
    assert user is not None
    user.role = Role.ANALYST
    if entity_code is not None:
        entity = db_session.scalar(select(Entity).where(Entity.code == entity_code))
        assert entity is not None
        user.entity_id_scope = entity.id
    else:
        user.entity_id_scope = None
    user.is_active = True
    db_session.flush()
    return cast(uuid.UUID, user.id)


def _login_as_cfo(client: TestClient, db_session: Session, email: str) -> None:
    _login(client, email)
    user = db_session.scalar(select(User).where(User.email == email))
    assert user is not None
    user.role = Role.CFO
    user.is_active = True
    db_session.flush()


def _login_as_pending(client: TestClient, db_session: Session, email: str) -> None:
    """Create a PENDING user and ensure the session reflects that role.

    The stub auth auto-creates the user as PENDING.  We clear the session
    cookie first so the admin session from a prior upload step doesn't linger,
    then hit the callback to establish a new PENDING session.
    """
    client.cookies.clear()
    _login(client, email)
    # Confirm user was created and is PENDING
    user = db_session.scalar(select(User).where(User.email == email))
    if user is not None:
        # Ensure role is PENDING (auto-assigned on first login via stub)
        assert user.role == Role.PENDING


# ---------------------------------------------------------------------------
# XLSX builders
# ---------------------------------------------------------------------------


def _make_cp_xlsx(
    india_rows: list[list[Any]] | None = None,
    uae_rows: list[list[Any]] | None = None,
) -> bytes:
    """Build a minimal CP master XLSX with India + UAE sheets."""
    wb = openpyxl.Workbook()
    del wb["Sheet"]
    ws_ind = wb.create_sheet("India")
    ws_ind.append(["Client Name", "Credit Period"])
    for row in india_rows or []:
        ws_ind.append(row)
    ws_uae = wb.create_sheet("UAE")
    ws_uae.append(["Client Name", "Credit Period", "Reason for extended Credit", "Amount"])
    for row in uae_rows or []:
        ws_uae.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_tally_xlsx_minimal() -> bytes:
    """Minimal single-invoice Tally XLSX for non-CP snapshot test."""
    meta = [
        ["Group :", "Sundry Debtors", None, "1-Apr-26 to 16-Apr-26", None, None, None],
        ["Details of:", "Pending Bills", None, None, None, None, None],
        [None] * 7,
        ["Date", "Ref. No.", "Party's Name", "Opening", "Pending", "Due on", "Overdue"],
        [None, None, None, "Amount", "Amount", None, "by days"],
    ]
    wb = openpyxl.Workbook()
    del wb["Sheet"]
    ws = wb.create_sheet("Sundry Debtors")
    for row in meta:
        ws.append(row)
    # party header then invoice
    ws.append([None, None, "AcmeCorp Ltd", None, None, None, None])
    ws.append([date(2026, 2, 1), "INV-001", None, 50000, 50000, date(2026, 3, 1), 30])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Upload helpers
# ---------------------------------------------------------------------------


def _upload_cp(
    client: TestClient,
    file_bytes: bytes,
    entity_code: str = "IND",
    filename: str = "cp.xlsx",
) -> Any:
    csrf = _csrf(client)
    headers = {"X-CSRF-Token": csrf} if csrf else {}
    return client.post(
        "/snapshots",
        data={"entity_code": entity_code, "source_hint": "CREDIT_PERIOD"},
        files={
            "file": (
                filename,
                io.BytesIO(file_bytes),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=headers,
    )


def _upload_tally(
    client: TestClient,
    file_bytes: bytes,
    entity_code: str = "IND",
) -> Any:
    csrf = _csrf(client)
    headers = {"X-CSRF-Token": csrf} if csrf else {}
    return client.post(
        "/snapshots",
        data={
            "entity_code": entity_code,
            "source_hint": "TALLY",
            "as_of_date": "2026-03-31",
        },
        files={
            "file": (
                "tally.xlsx",
                io.BytesIO(file_bytes),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=headers,
    )


def _get_cp_diff(client: TestClient, snapshot_id: str) -> Any:
    csrf = _csrf(client)
    headers = {"X-CSRF-Token": csrf} if csrf else {}
    return client.get(f"/snapshots/{snapshot_id}/cp-diff", headers=headers)


# ---------------------------------------------------------------------------
# DB seeding helpers
# ---------------------------------------------------------------------------


def _seed_canonical_and_config(
    db_session: Session,
    name: str,
    entity_code: str,
    days: int,
    reason_note: str | None,
    user_id: uuid.UUID,
) -> tuple[PartyCanonical, CreditPeriodConfig]:
    """Create a canonical + open CreditPeriodConfig row for diff testing."""
    entity = db_session.scalar(select(Entity).where(Entity.code == entity_code))
    assert entity is not None

    canonical = PartyCanonical(
        entity_id=entity.id,
        name=name,
        notes="seeded for test",
        created_by=user_id,
    )
    db_session.add(canonical)
    db_session.flush()

    alias = PartyAlias(
        canonical_id=canonical.id,
        alias_text=name,
        source="MANUAL",
        confidence=None,
        created_by=user_id,
    )
    db_session.add(alias)

    config = CreditPeriodConfig(
        canonical_id=canonical.id,
        days=days,
        reason_note=reason_note,
        valid_from=date(2026, 1, 1),
        valid_to=None,
        updated_by=user_id,
    )
    db_session.add(config)
    db_session.flush()
    return canonical, config


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


class TestCpDiffHappyPath:
    """Happy path: snapshot with ADDED + SUPERSEDED + UNCHANGED rows."""

    def test_all_three_categories(self, client: TestClient, db_session: Session) -> None:
        _login_as_admin(client)
        admin = db_session.scalar(
            select(User).where(User.email == "tejaswa.sharma@emb.global")
        )
        assert admin is not None

        # Seed existing canonicals + configs
        # "ExistingMatch" in India — days=30, will be UNCHANGED (CP has 30 too)
        _seed_canonical_and_config(
            db_session, "ExistingMatch Ltd", "IND", 30, None, admin.id
        )
        # "ExistingDiffer" in India — days=45, will be SUPERSEDED (CP has 60)
        _seed_canonical_and_config(
            db_session, "ExistingDiffer Ltd", "IND", 45, None, admin.id
        )
        # "NewClient" in India — no existing canonical → ADDED

        # Build CP xlsx:
        #   India: ExistingMatch Ltd → 30 (unchanged), ExistingDiffer Ltd → 60 (superseded), NewClient → 30 (added)
        india_rows = [
            ["ExistingMatch Ltd", 30],
            ["ExistingDiffer Ltd", 60],
            ["NewClient Ltd", 30],
        ]
        cp_bytes = _make_cp_xlsx(india_rows=india_rows)
        upload_resp = _upload_cp(client, cp_bytes)
        assert upload_resp.status_code == 201, upload_resp.text
        snapshot_id = upload_resp.json()["snapshot_id"]

        resp = _get_cp_diff(client, snapshot_id)
        assert resp.status_code == 200, resp.text
        body = resp.json()

        assert body["snapshot_id"] == snapshot_id
        assert isinstance(body["added"], list)
        assert isinstance(body["superseded"], list)
        assert isinstance(body["unchanged"], list)

        # Verify names appear in correct categories
        added_names = {r["canonical_name"] for r in body["added"]}
        superseded_names = {r["canonical_name"] for r in body["superseded"]}
        unchanged_names = {r["canonical_name"] for r in body["unchanged"]}

        assert "NewClient Ltd" in added_names
        assert "ExistingDiffer Ltd" in superseded_names
        assert "ExistingMatch Ltd" in unchanged_names

    def test_superseded_entry_has_prior_days(
        self, client: TestClient, db_session: Session
    ) -> None:
        _login_as_admin(client)
        admin = db_session.scalar(
            select(User).where(User.email == "tejaswa.sharma@emb.global")
        )
        assert admin is not None

        _seed_canonical_and_config(
            db_session, "OldTerms Ltd", "IND", 45, "legacy", admin.id
        )
        india_rows = [["OldTerms Ltd", 60]]
        cp_bytes = _make_cp_xlsx(india_rows=india_rows)
        upload_resp = _upload_cp(client, cp_bytes)
        assert upload_resp.status_code == 201
        snapshot_id = upload_resp.json()["snapshot_id"]

        resp = _get_cp_diff(client, snapshot_id)
        assert resp.status_code == 200
        body = resp.json()

        assert len(body["superseded"]) == 1
        entry = body["superseded"][0]
        assert entry["canonical_name"] == "OldTerms Ltd"
        assert entry["days"] == 60          # new
        assert entry["prior_days"] == 45    # was
        assert entry["prior_reason_note"] == "legacy"

    def test_added_entry_has_null_prior(
        self, client: TestClient, db_session: Session
    ) -> None:
        _login_as_admin(client)
        india_rows = [["BrandNewClient Ltd", 30]]
        cp_bytes = _make_cp_xlsx(india_rows=india_rows)
        upload_resp = _upload_cp(client, cp_bytes)
        assert upload_resp.status_code == 201
        snapshot_id = upload_resp.json()["snapshot_id"]

        resp = _get_cp_diff(client, snapshot_id)
        assert resp.status_code == 200
        body = resp.json()

        assert len(body["added"]) >= 1
        entry = next(r for r in body["added"] if r["canonical_name"] == "BrandNewClient Ltd")
        assert entry["prior_days"] is None
        assert entry["prior_reason_note"] is None


class TestCpDiffErrors:
    def test_422_for_non_cp_snapshot(
        self, client: TestClient, db_session: Session
    ) -> None:
        """Returns 422 when snapshot.source_hint != CREDIT_PERIOD."""
        _login_as_admin(client)
        tally_bytes = _make_tally_xlsx_minimal()
        upload_resp = _upload_tally(client, tally_bytes)
        assert upload_resp.status_code == 201, upload_resp.text
        snapshot_id = upload_resp.json()["snapshot_id"]

        resp = _get_cp_diff(client, snapshot_id)
        assert resp.status_code == 422
        body = resp.json()
        assert body["detail"]["code"] == "NOT_CREDIT_PERIOD_SNAPSHOT"

    def test_404_for_unknown_snapshot(
        self, client: TestClient, db_session: Session
    ) -> None:
        """Returns 404 for a snapshot UUID that does not exist."""
        _login_as_admin(client)
        fake_id = str(uuid.uuid4())
        resp = _get_cp_diff(client, fake_id)
        assert resp.status_code == 404


class TestCpDiffRbac:
    """RBAC matrix tests for cp-diff endpoint."""

    def _upload_cp_snapshot(
        self, client: TestClient, db_session: Session
    ) -> str:
        """Upload a CP snapshot as admin, return snapshot_id."""
        _login_as_admin(client)
        india_rows = [["SomeClient Ltd", 30]]
        cp_bytes = _make_cp_xlsx(india_rows=india_rows)
        upload_resp = _upload_cp(client, cp_bytes)
        assert upload_resp.status_code == 201, upload_resp.text
        return cast(str, upload_resp.json()["snapshot_id"])

    def test_analyst_in_scope_allowed(
        self, client: TestClient, db_session: Session
    ) -> None:
        """ANALYST with entity_id_scope == snapshot.entity_id gets 200."""
        snapshot_id = self._upload_cp_snapshot(client, db_session)

        # Fetch snapshot entity to set scope correctly
        snapshot = db_session.scalar(
            select(Snapshot).where(Snapshot.id == uuid.UUID(snapshot_id))
        )
        assert snapshot is not None
        entity = db_session.scalar(
            select(Entity).where(Entity.id == snapshot.entity_id)
        )
        assert entity is not None

        _login_as_analyst(client, db_session, "analyst_inscope@emb.global", entity.code)
        resp = _get_cp_diff(client, snapshot_id)
        assert resp.status_code == 200

    def test_analyst_out_of_scope_forbidden(
        self, client: TestClient, db_session: Session
    ) -> None:
        """ANALYST scoped to the other entity gets 403."""
        snapshot_id = self._upload_cp_snapshot(client, db_session)

        snapshot = db_session.scalar(
            select(Snapshot).where(Snapshot.id == uuid.UUID(snapshot_id))
        )
        assert snapshot is not None
        entity = db_session.scalar(
            select(Entity).where(Entity.id == snapshot.entity_id)
        )
        assert entity is not None
        # Pick the opposite entity
        other_code = "UAE" if entity.code == "IND" else "IND"

        _login_as_analyst(client, db_session, "analyst_outscope@emb.global", other_code)
        resp = _get_cp_diff(client, snapshot_id)
        assert resp.status_code == 403

    def test_cfo_read_allowed(
        self, client: TestClient, db_session: Session
    ) -> None:
        """CFO role gets 200 on cp-diff (read-only route)."""
        snapshot_id = self._upload_cp_snapshot(client, db_session)
        _login_as_cfo(client, db_session, "cfo_cpdiff@emb.global")
        resp = _get_cp_diff(client, snapshot_id)
        assert resp.status_code == 200

    def test_pending_role_forbidden(
        self, client: TestClient, db_session: Session
    ) -> None:
        """PENDING role gets 403."""
        snapshot_id = self._upload_cp_snapshot(client, db_session)
        _login_as_pending(client, db_session, "pending_cpdiff@emb.global")
        resp = _get_cp_diff(client, snapshot_id)
        assert resp.status_code == 403
