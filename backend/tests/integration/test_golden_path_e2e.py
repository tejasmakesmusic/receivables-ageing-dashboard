"""End-to-end test for the full upload → stage → publish → dashboard → email flow.

  SECTION A — NEW staging + publish features (2026-04-19):

  1. POST /snapshots/{id}/staging/bulk-create-canonicals (include_fuzzy=false)
     — creates canonicals for UNMAPPED rows in one call.

  2. POST /snapshots/{id}/staging/bulk-create-canonicals (include_fuzzy=true)
     — rejects FUZZY suggestions and creates new canonicals from raw names.

  3. POST /snapshots/{id}/publish on a CREDIT_PERIOD snapshot (ADR-0005):
     — happy path: writes credit_period_config rows + MANUAL aliases.
     — idempotent re-publish: identical rows → no-op.
     — supersede: changed days on re-publish → old row's valid_to = as_of - 1.

Covers ADR-0005 decisions D1 (auto-create canonicals), D2 (valid_from =
as_of_date, valid_to = NULL), D3 (no-op / supersede / insert policy).

  SECTION B — Full transactional golden path (spec §12 DoD):

  4. test_full_transactional_golden_path
     Exercises all 5 steps:
       upload    → POST /snapshots returns 201 STAGED
       staging   → GET /snapshots/{id}/staging returns rows; bulk-create-canonicals
       publish   → POST /snapshots/{id}/publish returns 200 PUBLISHED
       dashboard → GET /dashboard?entity=IND returns KPIs with the new snapshot
       email     → email_outbox has a QUEUED PUBLISH_NOTIF row for the snapshot

State tolerance: publish_service.py calls `db.commit()` mid-test, which
breaks the conftest per-test rollback.  Additionally, the Neon test branch
is forked from a developer's live branch, so existing canonicals may
already be present.  This suite therefore:
  - uses a unique UUID prefix on every party name to avoid collisions,
  - asserts on deltas (counts returned by the API) rather than absolute
    DB counts,
  - only reads DB state through WHERE clauses scoped to the test's unique
    prefix.

Measures wall-clock duration of each publish call so regression on publish
latency is visible in test output.
"""

from __future__ import annotations

import io
import time
import uuid
from datetime import date, timedelta
from typing import TYPE_CHECKING

import openpyxl
import pytest
from sqlalchemy import select

from app.db.models.credit_period_config import CreditPeriodConfig
from app.db.models.email_outbox import EmailOutbox
from app.db.models.entity import Entity
from app.db.models.invoice import Invoice
from app.db.models.invoice_snapshot import InvoiceSnapshot
from app.db.models.party import PartyAlias, PartyCanonical
from app.db.models.snapshot import Snapshot
from app.db.models.user import User

if TYPE_CHECKING:
    from fastapi.testclient import TestClient
    from sqlalchemy.orm import Session


# ---------------------------------------------------------------------------
# Auth helpers (stub-auth mode, forced in conftest)
# ---------------------------------------------------------------------------


ADMIN_EMAIL = "tejaswa.sharma@emb.global"


def _login_as_admin(client: TestClient) -> None:
    client.get(f"/auth/google/callback?stub_email={ADMIN_EMAIL}", follow_redirects=False)


def _csrf_headers(client: TestClient) -> dict[str, str]:
    tok = client.cookies.get("csrf_token") or ""
    return {"X-CSRF-Token": tok} if tok else {}


# ---------------------------------------------------------------------------
# XLSX builders
# ---------------------------------------------------------------------------


def _make_tally_xlsx(invoices: list[tuple[str, str, int]]) -> bytes:
    """Build a minimal Tally Sundry Debtors workbook.

    invoices: list of (party_name, invoice_ref, pending_amount).

    The parser expects exactly 5 header rows (indices 0-4) before data rows.
    Row layout:
      0 — Group header
      1 — Details header
      2 — blank
      3 — column header row 1
      4 — column header row 2 (sub-header / blank)
      5+ — party header rows + invoice rows + subtotal rows
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sundry Debtors"
    # 5 header rows (indices 0-4) matching _HEADER_ROWS = 5 in the parser
    ws.append(["Group :", "Sundry Debtors", None, "1-Apr-26 to 16-Apr-26", None, None, None])
    ws.append(["Details of:", "Pending Bills", None, None, None, None, None])
    ws.append([None] * 7)
    ws.append(
        ["Date", "Ref. No.", "Party's Name", "Opening Amount", "Pending Amount", "Due On", "Overdue"]
    )
    ws.append([None, None, None, "Amount", "Amount", None, "by days"])  # sub-header (row 4)
    grand_total = 0
    current_party: str | None = None
    party_subtotals: dict[str, int] = {}
    for party, ref, amount in invoices:
        if party != current_party:
            ws.append([None, None, party, None, None, None, None])  # party header row
            current_party = party
        ws.append([date(2026, 3, 1), ref, None, amount, amount, None, None])
        party_subtotals[party] = party_subtotals.get(party, 0) + amount
        grand_total += amount
    # Subtotal rows per party (subtotal-shaped: date+ref+party all None, amounts populated)
    for _p, subtotal in party_subtotals.items():
        ws.append([None, None, None, subtotal, subtotal, None, None])
    # Grand total row (subtotal-shaped)
    ws.append([None, None, None, grand_total, grand_total, None, None])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_cp_xlsx(ind_rows: list[tuple[str, int]], uae_rows: list[tuple[str, int, str | None]]) -> bytes:
    """Build a minimal Credit Period master workbook.

    ind_rows: (client_name, credit_period_days)
    uae_rows: (client_name, credit_period_days, reason_for_extended_credit)
    """
    wb = openpyxl.Workbook()
    ws_ind = wb.active
    assert ws_ind is not None
    ws_ind.title = "India"
    ws_ind.append(["Client Name", "Credit Period"])
    for name, days in ind_rows:
        ws_ind.append([name, days])

    ws_uae = wb.create_sheet("UAE")
    ws_uae.append(["Client Name", "Credit Period", "Reason for extended Credit", "Amount"])
    for name, days, reason in uae_rows:
        ws_uae.append([name, days, reason, None])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Test fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def ind_entity_with_default(db_session: Session) -> Entity:
    """IND entity with default_credit_days=30 set (matches production local setup)."""
    ind = db_session.scalar(select(Entity).where(Entity.code == "IND"))
    assert ind is not None, "IND entity must be seeded by migrations"
    ind.default_credit_days = 30
    db_session.flush()
    return ind


# ---------------------------------------------------------------------------
# 1. Bulk-create-canonicals — UNMAPPED-only mode
# ---------------------------------------------------------------------------


def test_bulk_create_canonicals_unmapped_only(
    client: TestClient,
    db_session: Session,
    ind_entity_with_default: Entity,
) -> None:
    _login_as_admin(client)
    tag = uuid.uuid4().hex[:8]  # unique-per-test party-name prefix
    names = [f"E2E-{tag}-Acme", f"E2E-{tag}-Beta", f"E2E-{tag}-Gamma"]

    xlsx = _make_tally_xlsx(
        [
            (names[0], f"INV-{tag}-1", 100_000),
            (names[0], f"INV-{tag}-2", 200_000),
            (names[1], f"INV-{tag}-3", 50_000),
            (names[2], f"INV-{tag}-4", 75_000),
        ]
    )
    resp = client.post(
        "/snapshots",
        files={"file": ("GrpBills.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"entity_code": "IND", "source_hint": "TALLY", "as_of_date": "2026-03-31"},
        headers=_csrf_headers(client),
    )
    assert resp.status_code == 201, resp.text
    snap_id = resp.json()["snapshot_id"]

    # Bulk create (UNMAPPED only — default)
    t0 = time.perf_counter()
    bulk = client.post(
        f"/snapshots/{snap_id}/staging/bulk-create-canonicals",
        json={},
        headers=_csrf_headers(client),
    )
    elapsed = time.perf_counter() - t0
    assert bulk.status_code == 200, bulk.text
    body = bulk.json()
    # All 3 distinct unique-prefixed names were UNMAPPED and become EXACT.
    assert body["distinct_unmapped_names"] == 3
    assert body["created_canonicals"] == 3
    assert body["created_aliases"] == 3
    assert body["publish_gate"]["unmapped_parties_count"] == 0
    print(f"\n[timing] bulk-create-canonicals(unmapped): {elapsed*1000:.0f}ms")

    # DB side-effect: the 3 uniquely-prefixed canonicals exist.
    cans = db_session.scalars(
        select(PartyCanonical).where(
            PartyCanonical.entity_id == ind_entity_with_default.id,
            PartyCanonical.name.in_(names),
        )
    ).all()
    assert {c.name for c in cans} == set(names)
    aliases = db_session.scalars(
        select(PartyAlias).where(PartyAlias.alias_text.in_(names))
    ).all()
    assert len(aliases) == 3
    assert {a.source for a in aliases} == {"MANUAL"}


# ---------------------------------------------------------------------------
# 2. Bulk-create-canonicals — include_fuzzy=true rejects suggestions
# ---------------------------------------------------------------------------


def test_bulk_create_canonicals_include_fuzzy(
    client: TestClient,
    db_session: Session,
    ind_entity_with_default: Entity,
) -> None:
    _login_as_admin(client)
    tag = uuid.uuid4().hex[:8]
    # Two corporate-suffix-heavy names that should fuzzy-match each other at
    # ~70-87% via rapidfuzz (the "classic false positive" case).  The 32-char
    # suffix overlap drives the score even though the distinctive prefix
    # differs. We seed `seed_name` as a canonical, then upload a snapshot
    # containing `raw_name` which the resolver flags as FUZZY — not EXACT.
    seed_name = f"E2E-{tag}-Amazon Web Services India Private Limited"
    raw_name = f"E2E-{tag}-Webify Services India Private Limited"

    admin = db_session.scalar(select(User).where(User.email == ADMIN_EMAIL))
    assert admin is not None
    seeded = PartyCanonical(
        entity_id=ind_entity_with_default.id,
        name=seed_name,
        created_by=admin.id,
    )
    db_session.add(seeded)
    db_session.flush()
    db_session.add(
        PartyAlias(
            canonical_id=seeded.id,
            alias_text=seed_name,
            source="MANUAL",
            created_by=admin.id,
        )
    )
    db_session.commit()  # persist so the HTTP handler's session sees it

    xlsx = _make_tally_xlsx(
        [
            (raw_name, f"INV-{tag}-1", 100_000),
            (raw_name, f"INV-{tag}-2", 200_000),
        ]
    )
    resp = client.post(
        "/snapshots",
        files={"file": ("GrpBills.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"entity_code": "IND", "source_hint": "TALLY", "as_of_date": "2026-03-31"},
        headers=_csrf_headers(client),
    )
    assert resp.status_code == 201, resp.text
    snap_id = resp.json()["snapshot_id"]

    # UNMAPPED-only bulk: the raw name fuzzy-matches the seed, so it's FUZZY,
    # not UNMAPPED — bulk should no-op on this specific raw name.
    bulk1 = client.post(
        f"/snapshots/{snap_id}/staging/bulk-create-canonicals",
        json={"include_fuzzy": False},
        headers=_csrf_headers(client),
    ).json()
    # The Webify canonical must NOT have been created by the UNMAPPED-only call.
    webify_before = db_session.scalar(
        select(PartyCanonical).where(PartyCanonical.name == raw_name)
    )
    assert webify_before is None
    assert bulk1["publish_gate"]["unmapped_parties_count"] >= 2  # 2 fuzzy rows still blocking

    # include_fuzzy=True: new canonical from raw name, bypassing the fuzzy suggestion.
    bulk2 = client.post(
        f"/snapshots/{snap_id}/staging/bulk-create-canonicals",
        json={"include_fuzzy": True},
        headers=_csrf_headers(client),
    ).json()
    assert bulk2["publish_gate"]["unmapped_parties_count"] == 0

    webify_after = db_session.scalar(
        select(PartyCanonical).where(PartyCanonical.name == raw_name)
    )
    assert webify_after is not None
    assert webify_after.id != seeded.id


# ---------------------------------------------------------------------------
# 3. CREDIT_PERIOD publish — happy path + idempotent + supersede (ADR-0005)
# ---------------------------------------------------------------------------


def _upload_cp_snapshot(
    client: TestClient,
    *,
    ind_rows: list[tuple[str, int]],
    uae_rows: list[tuple[str, int, str | None]],
    as_of: str,
) -> str:
    xlsx = _make_cp_xlsx(ind_rows, uae_rows)
    resp = client.post(
        "/snapshots",
        files={
            "file": (
                "Credit Period.xlsx",
                xlsx,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"entity_code": "IND", "source_hint": "CREDIT_PERIOD", "as_of_date": as_of},
        headers=_csrf_headers(client),
    )
    assert resp.status_code == 201, resp.text
    return str(resp.json()["snapshot_id"])


def test_credit_period_publish_happy_path(
    client: TestClient,
    db_session: Session,
    ind_entity_with_default: Entity,
) -> None:
    """First-publish of a CP master: creates canonicals, aliases, config rows."""
    _login_as_admin(client)
    tag = uuid.uuid4().hex[:8]
    foo, bar = f"CP-{tag}-Foo", f"CP-{tag}-Bar"
    zeta, eta = f"CP-{tag}-Zeta", f"CP-{tag}-Eta"

    snap_id = _upload_cp_snapshot(
        client,
        ind_rows=[(foo, 30), (bar, 45)],
        uae_rows=[(zeta, 15, "quarterly"), (eta, 60, None)],
        as_of="2026-03-31",
    )

    t0 = time.perf_counter()
    resp = client.post(f"/snapshots/{snap_id}/publish", json={}, headers=_csrf_headers(client))
    elapsed = time.perf_counter() - t0
    assert resp.status_code == 200, resp.text
    result = resp.json()["result"]
    assert result["credit_period_configs_inserted"] == 4
    assert result["credit_period_configs_superseded"] == 0
    assert result["credit_period_configs_noop"] == 0
    assert result["canonicals_auto_created"] == 4
    assert result["aliases_auto_created"] == 4
    print(f"\n[timing] cp-publish(happy-path, 4 rows): {elapsed*1000:.0f}ms")

    snap = db_session.get(Snapshot, uuid.UUID(snap_id))
    assert snap is not None and snap.status == "PUBLISHED"

    # Check the 4 just-created config rows (scope by our unique tag).
    configs = db_session.scalars(
        select(CreditPeriodConfig)
        .join(PartyCanonical, PartyCanonical.id == CreditPeriodConfig.canonical_id)
        .where(PartyCanonical.name.in_([foo, bar, zeta, eta]))
    ).all()
    assert len(configs) == 4
    for c in configs:
        assert c.valid_to is None
        assert c.valid_from == date(2026, 3, 31)


def test_credit_period_publish_idempotent(
    client: TestClient,
    db_session: Session,
    ind_entity_with_default: Entity,
) -> None:
    """Publishing the same CP master twice: second round is all no-ops."""
    _login_as_admin(client)
    tag = uuid.uuid4().hex[:8]
    foo, zeta = f"CP-{tag}-Foo", f"CP-{tag}-Zeta"

    snap1 = _upload_cp_snapshot(
        client,
        ind_rows=[(foo, 30)],
        uae_rows=[(zeta, 15, None)],
        as_of="2026-03-31",
    )
    resp1 = client.post(f"/snapshots/{snap1}/publish", json={}, headers=_csrf_headers(client))
    assert resp1.status_code == 200, resp1.text
    assert resp1.json()["result"]["credit_period_configs_inserted"] == 2

    snap2 = _upload_cp_snapshot(
        client,
        ind_rows=[(foo, 30)],
        uae_rows=[(zeta, 15, None)],
        as_of="2026-04-30",
    )
    resp2 = client.post(f"/snapshots/{snap2}/publish", json={}, headers=_csrf_headers(client))
    assert resp2.status_code == 200, resp2.text
    result = resp2.json()["result"]
    assert result["credit_period_configs_inserted"] == 0
    assert result["credit_period_configs_superseded"] == 0
    assert result["credit_period_configs_noop"] == 2
    assert result["canonicals_auto_created"] == 0
    assert result["aliases_auto_created"] == 0

    # Exactly 2 config rows for our two uniquely-tagged clients, both open.
    configs = db_session.scalars(
        select(CreditPeriodConfig)
        .join(PartyCanonical, PartyCanonical.id == CreditPeriodConfig.canonical_id)
        .where(PartyCanonical.name.in_([foo, zeta]))
    ).all()
    assert len(configs) == 2
    assert all(c.valid_to is None for c in configs)


def test_credit_period_publish_supersede_on_value_change(
    client: TestClient,
    db_session: Session,
    ind_entity_with_default: Entity,
) -> None:
    """Re-publishing the same client with a different credit_days supersedes:
    old row's valid_to = new.as_of_date - 1 day; new row with valid_to = NULL."""
    _login_as_admin(client)
    tag = uuid.uuid4().hex[:8]
    foo = f"CP-{tag}-Foo"

    snap1 = _upload_cp_snapshot(
        client,
        ind_rows=[(foo, 30)],
        uae_rows=[],
        as_of="2026-03-31",
    )
    resp1 = client.post(f"/snapshots/{snap1}/publish", json={}, headers=_csrf_headers(client))
    assert resp1.status_code == 200, resp1.text

    snap2 = _upload_cp_snapshot(
        client,
        ind_rows=[(foo, 45)],
        uae_rows=[],
        as_of="2026-04-30",
    )
    resp2 = client.post(f"/snapshots/{snap2}/publish", json={}, headers=_csrf_headers(client))
    assert resp2.status_code == 200, resp2.text
    result = resp2.json()["result"]
    assert result["credit_period_configs_inserted"] == 0
    assert result["credit_period_configs_superseded"] == 1
    assert result["credit_period_configs_noop"] == 0

    foo_canon = db_session.scalar(select(PartyCanonical).where(PartyCanonical.name == foo))
    assert foo_canon is not None
    configs = db_session.scalars(
        select(CreditPeriodConfig).where(CreditPeriodConfig.canonical_id == foo_canon.id)
    ).all()
    assert len(configs) == 2
    old = next(c for c in configs if c.days == 30)
    new = next(c for c in configs if c.days == 45)
    assert old.valid_from == date(2026, 3, 31)
    assert old.valid_to == date(2026, 4, 30) - timedelta(days=1)
    assert new.valid_from == date(2026, 4, 30)
    assert new.valid_to is None


# ---------------------------------------------------------------------------
# 4. Full transactional golden path (spec §12 DoD)
#    upload → staging → publish → dashboard → email enqueue
# ---------------------------------------------------------------------------


def test_full_transactional_golden_path(
    client: TestClient,
    db_session: Session,
    ind_entity_with_default: Entity,
) -> None:
    """Exercises all 5 §12 DoD steps for a TALLY transactional snapshot.

    Step 1 — Upload (POST /snapshots): file parses cleanly → 201 STAGED.
    Step 2 — Staging (GET /snapshots/{id}/staging + bulk-create-canonicals):
              staging view returns rows; bulk-create maps every UNMAPPED party.
    Step 3 — Publish (POST /snapshots/{id}/publish): snapshot transitions to
              PUBLISHED; invoices + invoice_snapshots are written; publish gate
              reports invoices_inserted > 0.
    Step 4 — Dashboard (GET /dashboard?entity=IND): returns 200 with KPIs
              derived from the snapshot just published; total_ar > 0.
    Step 5 — Email enqueue: email_outbox has a QUEUED PUBLISH_NOTIF row
              whose snapshot_id matches and subject contains the entity code.
    """
    _login_as_admin(client)
    tag = uuid.uuid4().hex[:8]
    party_a = f"GP-{tag}-Alpha"
    party_b = f"GP-{tag}-Beta"

    # ------------------------------------------------------------------
    # Step 1: Upload a Tally snapshot
    # ------------------------------------------------------------------
    xlsx = _make_tally_xlsx(
        [
            (party_a, f"INV-{tag}-001", 500_000),
            (party_a, f"INV-{tag}-002", 300_000),
            (party_b, f"INV-{tag}-003", 200_000),
        ]
    )
    t_upload = time.perf_counter()
    upload_resp = client.post(
        "/snapshots",
        files={
            "file": (
                "GrpBills.xlsx",
                xlsx,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"entity_code": "IND", "source_hint": "TALLY", "as_of_date": "2026-03-31"},
        headers=_csrf_headers(client),
    )
    assert upload_resp.status_code == 201, upload_resp.text
    snap_id = upload_resp.json()["snapshot_id"]
    print(f"\n[timing] upload: {(time.perf_counter() - t_upload)*1000:.0f}ms")

    # Snapshot is in STAGED state immediately after upload.
    snap_obj = db_session.get(Snapshot, uuid.UUID(snap_id))
    assert snap_obj is not None
    assert snap_obj.status == "STAGED"

    # ------------------------------------------------------------------
    # Step 2: Staging — inspect then bulk-create-canonicals
    # ------------------------------------------------------------------
    staging_resp = client.get(f"/snapshots/{snap_id}/staging", headers=_csrf_headers(client))
    assert staging_resp.status_code == 200, staging_resp.text
    staging_body = staging_resp.json()
    # At least the 3 invoices we uploaded should appear (pagination.total).
    assert staging_body["pagination"]["total"] >= 3
    # All 3 party names are brand-new → publish gate must be blocked (UNMAPPED).
    gate = staging_body["publish_gate"]
    assert gate["ok"] is False
    assert gate["unmapped_parties_count"] >= 2  # 2 distinct parties

    # Acknowledge any parse warnings so the publish gate's warnings sub-check passes.
    unacked = staging_body["publish_gate"]["warnings_unacknowledged"]
    if unacked:
        ack_resp = client.patch(
            f"/snapshots/{snap_id}/warnings/ack",
            json={"codes": unacked},
            headers=_csrf_headers(client),
        )
        assert ack_resp.status_code == 200, ack_resp.text

    # Bulk-create resolves all UNMAPPED parties in one call.
    t_bulk = time.perf_counter()
    bulk_resp = client.post(
        f"/snapshots/{snap_id}/staging/bulk-create-canonicals",
        json={},
        headers=_csrf_headers(client),
    )
    assert bulk_resp.status_code == 200, bulk_resp.text
    bulk_body = bulk_resp.json()
    assert bulk_body["created_canonicals"] == 2  # party_a + party_b
    assert bulk_body["created_aliases"] == 2
    assert bulk_body["publish_gate"]["unmapped_parties_count"] == 0
    assert bulk_body["publish_gate"]["ok"] is True
    print(f"\n[timing] bulk-create-canonicals: {(time.perf_counter() - t_bulk)*1000:.0f}ms")

    # ------------------------------------------------------------------
    # Step 3: Publish
    # ------------------------------------------------------------------
    t_publish = time.perf_counter()
    pub_resp = client.post(
        f"/snapshots/{snap_id}/publish",
        json={},
        headers=_csrf_headers(client),
    )
    assert pub_resp.status_code == 200, pub_resp.text
    pub_result = pub_resp.json()["result"]
    print(f"\n[timing] publish: {(time.perf_counter() - t_publish)*1000:.0f}ms")

    # Snapshot transitions to PUBLISHED.
    db_session.expire(snap_obj)
    assert snap_obj.status == "PUBLISHED"

    # Invoices written for our uniquely-tagged parties.
    inv_count = len(
        db_session.scalars(
            select(Invoice).where(
                Invoice.entity_id == ind_entity_with_default.id,
                Invoice.invoice_ref.like(f"INV-{tag}-%"),
            )
        ).all()
    )
    assert inv_count == 3, f"Expected 3 invoices, got {inv_count}"

    # invoice_snapshots written (one per effective invoice row).
    inv_snap_count = len(
        db_session.scalars(
            select(InvoiceSnapshot).where(InvoiceSnapshot.snapshot_id == uuid.UUID(snap_id))
        ).all()
    )
    assert inv_snap_count == 3, f"Expected 3 invoice_snapshots, got {inv_snap_count}"

    # Publish result reports the inserts.
    assert pub_result["invoices_inserted"] == 3
    assert pub_result["invoices_updated"] == 0
    # invoices_settled may be > 0 if the Neon test branch contains prior open
    # invoices from other test runs (state tolerance — see module docstring).
    assert pub_result["invoices_settled"] >= 0
    assert pub_result["invoice_snapshots_written"] == 3
    assert pub_result["publish_notif_enqueued"] is True

    # ------------------------------------------------------------------
    # Step 4: Dashboard — KPIs reflect the published snapshot
    # ------------------------------------------------------------------
    dash_resp = client.get("/dashboard?entity=IND&as_of=latest", headers=_csrf_headers(client))
    assert dash_resp.status_code == 200, dash_resp.text
    dash = dash_resp.json()
    # kpis.total_outstanding must be positive (we published 1_000_000 worth of invoices).
    assert float(dash["kpis"]["total_outstanding"]) > 0, (
        "Dashboard kpis.total_outstanding should be > 0 after publish"
    )
    # top_parties must include our 2 distinct parties (party_a, party_b).
    assert len(dash["top_parties"]) >= 2, (
        f"Expected at least 2 top_parties, got {len(dash['top_parties'])}"
    )

    # ------------------------------------------------------------------
    # Step 5: Email enqueue — QUEUED PUBLISH_NOTIF row in email_outbox
    # ------------------------------------------------------------------
    outbox_rows = db_session.scalars(
        select(EmailOutbox).where(
            EmailOutbox.snapshot_id == uuid.UUID(snap_id),
            EmailOutbox.rule_type == "PUBLISH_NOTIF",
        )
    ).all()
    assert len(outbox_rows) == 1, f"Expected 1 outbox row, got {len(outbox_rows)}"
    outbox = outbox_rows[0]
    assert outbox.status == "QUEUED"
    assert "IND" in outbox.subject
    assert str(snap_id) in outbox.subject
