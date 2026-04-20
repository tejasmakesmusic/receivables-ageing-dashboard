"""Unit tests for app.parsers.credit_period (M2 Task 4).

Real-fixture tests skip gracefully when the fixture is absent from the
fixture directory (CI-safe: spec §12, plan open-decision #2).

Data-handling rule (CLAUDE.md §"Data handling"): no raw client names in
assertion messages.  All assertions are structural (counts, entity codes,
field presence).  Synthetic client names like "AlphaClient Ltd" are fine.
"""

from __future__ import annotations

import io
import json
from pathlib import Path
from typing import Any

import openpyxl
import pytest

from app.parsers.common import StagedCreditPeriod
from app.parsers.credit_period import parse_credit_period_master

# ---------------------------------------------------------------------------
# Fixture path
# ---------------------------------------------------------------------------

FIXTURE_PATH = (
    Path(__file__).resolve().parents[2]
    / "fixtures"
    / "sample_files"
    / "Credit Period for Accounts - India & UAE.xlsx"
)


@pytest.fixture
def credit_period_file_bytes() -> bytes:
    # Per M2 plan open decision #2 — real fixtures are .gitignored; skip in
    # envs (CI) that don't have them locally. Revisit with sanitized fixtures
    # in M3 if CI coverage gap bites.
    if not FIXTURE_PATH.exists():
        pytest.skip(f"fixture not present: {FIXTURE_PATH}")
    return FIXTURE_PATH.read_bytes()


# ---------------------------------------------------------------------------
# Synthetic XLSX builder
# ---------------------------------------------------------------------------
#
# Builds a minimal Credit Period master workbook in memory.
# India sheet: 2 cols (Client Name | Credit Period)
# UAE sheet:   4 cols (Client Name | Credit Period | Reason for extended Credit | Amount)
#
# D20: Amount column is present in the UAE sheet so the parser must DROP it.
# The builder always includes Amount to make the D20 test load-bearing.

_INDIA_HEADER = ["Client Name", "Credit Period"]
_UAE_HEADER = ["Client Name", "Credit Period", "Reason for extended Credit", "Amount"]


def _make_ind_row(name: str | None, credit_days: Any) -> list[Any]:
    return [name, credit_days]


def _make_uae_row(
    name: str,
    credit_days: Any,
    reason: str | None = None,
    amount: Any = None,
) -> list[Any]:
    return [name, credit_days, reason, amount]


def _build_wb_bytes(
    india_rows: list[list[Any]],
    uae_rows: list[list[Any]],
    include_india: bool = True,
    include_uae: bool = True,
) -> bytes:
    """Build a minimal Credit Period master XLSX in memory.

    Args:
        india_rows: Data rows for the India sheet (excluding header).
        uae_rows: Data rows for the UAE sheet (excluding header).
        include_india: If False, the India sheet is omitted entirely.
        include_uae: If False, the UAE sheet is omitted entirely.

    Returns:
        Raw bytes of the ``.xlsx`` file.
    """
    wb = openpyxl.Workbook()
    del wb["Sheet"]  # remove default sheet

    if include_india:
        ws_india = wb.create_sheet("India")
        ws_india.append(_INDIA_HEADER)
        for row in india_rows:
            ws_india.append(row)

    if include_uae:
        ws_uae = wb.create_sheet("UAE")
        ws_uae.append(_UAE_HEADER)
        for row in uae_rows:
            ws_uae.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Minimal valid synthetic fixture
# ---------------------------------------------------------------------------

_SIMPLE_IND_ROWS: list[list[Any]] = [
    _make_ind_row("AlphaClient Ltd", 30),
    _make_ind_row("BetaClient Inc", 0),
    _make_ind_row("GammaClient Co", 15),
]
_SIMPLE_UAE_ROWS: list[list[Any]] = [
    _make_uae_row("DeltaClient LLC", 30, "Long-term contract", 99999),
    _make_uae_row("EpsilonClient FZ", 0, None, None),
]


@pytest.fixture
def simple_cp_bytes() -> bytes:
    return _build_wb_bytes(_SIMPLE_IND_ROWS, _SIMPLE_UAE_ROWS)


# ---------------------------------------------------------------------------
# Test 1: source_hint and sha256 shape
# ---------------------------------------------------------------------------


def test_source_hint_credit_period_and_sha256_64_char_hex(simple_cp_bytes: bytes) -> None:
    result = parse_credit_period_master(simple_cp_bytes)
    assert result.source_hint == "CREDIT_PERIOD"
    assert len(result.file_sha256) == 64
    assert all(c in "0123456789abcdef" for c in result.file_sha256)


# ---------------------------------------------------------------------------
# Test 2: invoices == [] and as_of_date is None on all results
# ---------------------------------------------------------------------------


def test_invoices_empty_and_as_of_date_none(simple_cp_bytes: bytes) -> None:
    result = parse_credit_period_master(simple_cp_bytes)
    assert result.invoices == [], "Credit Period parser must never emit invoices"
    assert result.as_of_date is None, "as_of_date must always be None for Credit Period parser"


def test_invoices_empty_on_real_fixture(credit_period_file_bytes: bytes) -> None:
    result = parse_credit_period_master(credit_period_file_bytes)
    assert result.invoices == []
    assert result.as_of_date is None


# ---------------------------------------------------------------------------
# Test 3: real fixture — structural parse (fixture has data quality issues)
#
# Fixture inspection 2026-04-17 revealed real data quality issues:
#   - India sheet: 3 duplicate client names → DUPLICATE_CLIENT → is_valid False
#   - UAE sheet:   2 duplicate client names + 7 empty credit_days → errors
#
# The parser correctly flags these as errors per spec §4.3 rule 5.
# The fixture would need cleansing before re-upload; that is the analyst's job.
# We assert the structural parser behaviour (error codes present, correct counts)
# rather than asserting is_valid=True (which is incorrect for this fixture).
# ---------------------------------------------------------------------------


def test_real_fixture_parser_runs_and_emits_expected_error_codes(
    credit_period_file_bytes: bytes,
) -> None:
    """Real fixture parses cleanly after 2026-04-19 cleanse.

    Pre-cleanse the fixture held 2 DUPLICATE_CLIENT + 7 UNPARSEABLE_CREDIT_DAYS
    errors. Those were removed at the source per policy (identical dups
    collapsed; conflicting dups + blank-credit-days rows deleted so affected
    clients fall to entity default via D8). Parser now returns is_valid=True.
    """
    result = parse_credit_period_master(credit_period_file_bytes)
    # Parser must complete without exception (structural integrity).
    assert result.source_hint == "CREDIT_PERIOD"
    assert result.file_sha256, "file_sha256 must be set"

    # Post-cleanse: no errors expected.
    assert (
        result.errors == []
    ), f"Expected clean fixture; got errors={[e.code for e in result.errors]}"
    assert result.is_valid is True, "Cleansed fixture must have is_valid=True."


# ---------------------------------------------------------------------------
# Test 4: real fixture — zero errors post-cleanse (2026-04-19)
# ---------------------------------------------------------------------------


def test_real_fixture_exact_error_counts(credit_period_file_bytes: bytes) -> None:
    """Exact error counts post-cleanse (2026-04-19).

    Fixture was cleansed per policy — identical dups deduped, conflicting dups
    and blank-credit-days rows deleted. Both error codes that previously fired
    (DUPLICATE_CLIENT, UNPARSEABLE_CREDIT_DAYS) now have zero instances.

    If the fixture is replaced, re-measure and update.
    """
    result = parse_credit_period_master(credit_period_file_bytes)
    dup_errs = [e for e in result.errors if e.code == "DUPLICATE_CLIENT"]
    invalid_errs = [e for e in result.errors if e.code == "UNPARSEABLE_CREDIT_DAYS"]

    assert len(dup_errs) == 0, (
        f"Expected 0 DUPLICATE_CLIENT errors on cleansed fixture; got {len(dup_errs)}."
    )
    assert len(invalid_errs) == 0, (
        f"Expected 0 UNPARSEABLE_CREDIT_DAYS errors on cleansed fixture; "
        f"got {len(invalid_errs)}."
    )


# ---------------------------------------------------------------------------
# Test 5: entity_code invariants — IND has reason_note=None, UAE has entity_code="UAE"
# ---------------------------------------------------------------------------


def test_entity_code_invariants(simple_cp_bytes: bytes) -> None:
    result = parse_credit_period_master(simple_cp_bytes)
    for cp in result.credit_periods:
        if cp.entity_code == "IND":
            assert (
                cp.reason_note is None
            ), f"IND StagedCreditPeriod at row {cp.row_index} must have reason_note=None"
        else:
            assert (
                cp.entity_code == "UAE"
            ), f"entity_code must be 'IND' or 'UAE'; got {cp.entity_code!r}"


def test_entity_code_invariants_real_fixture(credit_period_file_bytes: bytes) -> None:
    result = parse_credit_period_master(credit_period_file_bytes)
    for cp in result.credit_periods:
        assert cp.entity_code in (
            "IND",
            "UAE",
        ), f"Unexpected entity_code {cp.entity_code!r} at row {cp.row_index}"
        if cp.entity_code == "IND":
            assert cp.reason_note is None, f"IND row {cp.row_index} must have reason_note=None"


# ---------------------------------------------------------------------------
# Test 6: D20 — Amount column NEVER reaches any StagedCreditPeriod
# ---------------------------------------------------------------------------


def test_d20_amount_not_in_model_fields() -> None:
    """StagedCreditPeriod.model_fields must not contain 'amount' or any Amount variant."""
    field_names = set(StagedCreditPeriod.model_fields.keys())
    for bad in ("amount", "Amount", "amount_col", "uae_amount"):
        assert (
            bad not in field_names
        ), f"D20 violation: StagedCreditPeriod.model_fields contains {bad!r}"


def test_d20_amount_value_not_in_parsed_output() -> None:
    """Synthetic UAE row with Amount=99999 — that value must NOT appear anywhere in output.

    This test is load-bearing for D20 compliance: if the Amount column leaks
    into any field, model_dump(), or ParseError.detail, the sentinel value
    '99999' will appear in the JSON dump and this test will catch it.
    """
    sentinel = "99999"
    rows_with_sentinel: list[list[Any]] = [
        _make_uae_row("AlphaClient Ltd", 30, "Extended", 99999),
        _make_uae_row("BetaClient Ltd", 0, None, 99999),
    ]
    file_bytes = _build_wb_bytes(
        india_rows=[_make_ind_row("IndClient Ltd", 30)],
        uae_rows=rows_with_sentinel,
    )
    result = parse_credit_period_master(file_bytes)

    # Check every StagedCreditPeriod for the sentinel.
    for cp in result.credit_periods:
        dumped = json.dumps(cp.model_dump())
        assert sentinel not in dumped, (
            f"D20 violation: sentinel value {sentinel!r} found in "
            f"StagedCreditPeriod.model_dump() at row {cp.row_index}: {dumped}"
        )

    # Check every ParseError.detail as well.
    for err in result.errors + result.warnings:
        if err.detail is not None:
            detail_json = json.dumps(err.detail)
            assert sentinel not in detail_json, (
                f"D20 violation: sentinel value {sentinel!r} found in "
                f"ParseError.detail (code={err.code}): {detail_json}"
            )


# ---------------------------------------------------------------------------
# Test 7: credit_days = 0 is valid
# ---------------------------------------------------------------------------


def test_credit_days_zero_is_valid() -> None:
    """credit_days=0 (immediate payment) must be emitted as a valid StagedCreditPeriod."""
    file_bytes = _build_wb_bytes(
        india_rows=[_make_ind_row("ZeroDayClient Ltd", 0)],
        uae_rows=[_make_uae_row("ZeroUaeClient LLC", 0)],
    )
    result = parse_credit_period_master(file_bytes)
    assert (
        result.is_valid is True
    ), f"credit_days=0 must be valid. errors={[e.code for e in result.errors]}"
    zero_day_rows = [cp for cp in result.credit_periods if cp.credit_days == 0]
    assert len(zero_day_rows) == 2, f"Expected 2 rows with credit_days=0; got {len(zero_day_rows)}"
    for cp in zero_day_rows:
        assert cp.credit_days == 0


# ---------------------------------------------------------------------------
# Test 8: credit_days = -5 → UNPARSEABLE_CREDIT_DAYS in errors; is_valid False
# ---------------------------------------------------------------------------


def test_negative_credit_days_emits_invalid_error() -> None:
    """credit_days=-5 must emit UNPARSEABLE_CREDIT_DAYS in errors; is_valid False."""
    file_bytes = _build_wb_bytes(
        india_rows=[_make_ind_row("NegativeClient Ltd", -5)],
        uae_rows=[_make_uae_row("ValidUaeClient LLC", 30)],
    )
    result = parse_credit_period_master(file_bytes)
    assert result.is_valid is False, "Negative credit_days must make is_valid False"
    invalid_errs = [e for e in result.errors if e.code == "UNPARSEABLE_CREDIT_DAYS"]
    assert invalid_errs, (
        f"Expected UNPARSEABLE_CREDIT_DAYS in errors. " f"errors={[e.code for e in result.errors]}"
    )


# ---------------------------------------------------------------------------
# Test 9: credit_days = "not a number" → UNPARSEABLE_CREDIT_DAYS; is_valid False
# ---------------------------------------------------------------------------


def test_non_numeric_credit_days_emits_invalid_error() -> None:
    """credit_days='not a number' must emit UNPARSEABLE_CREDIT_DAYS in errors."""
    file_bytes = _build_wb_bytes(
        india_rows=[_make_ind_row("BadValueClient Ltd", "not a number")],
        uae_rows=[_make_uae_row("ValidUaeClient LLC", 30)],
    )
    result = parse_credit_period_master(file_bytes)
    assert result.is_valid is False, "Unparseable credit_days must make is_valid False"
    invalid_errs = [e for e in result.errors if e.code == "UNPARSEABLE_CREDIT_DAYS"]
    assert invalid_errs, (
        f"Expected UNPARSEABLE_CREDIT_DAYS in errors. " f"errors={[e.code for e in result.errors]}"
    )
    # detail must include the sheet and value.
    err = invalid_errs[0]
    assert err.detail is not None
    assert "sheet" in err.detail


# ---------------------------------------------------------------------------
# Test 10: empty Client Name rows mixed with valid rows → empty rows skipped
# ---------------------------------------------------------------------------


def test_empty_client_name_rows_skipped() -> None:
    """Empty Client Name rows (blank separators) must be silently skipped.

    Valid rows around the empty rows must still be emitted.
    is_valid must be True (empty-name rows are not errors).
    """
    india_rows: list[list[Any]] = [
        _make_ind_row("ValidClientA Ltd", 30),
        _make_ind_row("", 0),  # empty string → skip
        _make_ind_row(None, 0),  # None → skip
        _make_ind_row("ValidClientB Inc", 15),
        _make_ind_row("   ", 0),  # whitespace-only → skip
        _make_ind_row("ValidClientC Co", 45),
    ]
    file_bytes = _build_wb_bytes(
        india_rows=india_rows,
        uae_rows=[_make_uae_row("ValidUaeClient LLC", 30)],
    )
    result = parse_credit_period_master(file_bytes)
    assert result.is_valid is True, (
        f"Empty client name rows must not produce errors. "
        f"errors={[e.code for e in result.errors]}"
    )
    ind_rows = [cp for cp in result.credit_periods if cp.entity_code == "IND"]
    assert len(ind_rows) == 3, f"Expected 3 IND rows (empty rows skipped); got {len(ind_rows)}"


# ---------------------------------------------------------------------------
# Test 11: duplicate client names in IND sheet → DUPLICATE_CLIENT error
# ---------------------------------------------------------------------------


def test_duplicate_client_names_in_india_sheet_fails_parse() -> None:
    """'AlphaClient Ltd' appears twice in India sheet → DUPLICATE_CLIENT error."""
    india_rows: list[list[Any]] = [
        _make_ind_row("AlphaClient Ltd", 30),
        _make_ind_row("BetaClient Inc", 0),
        _make_ind_row("AlphaClient Ltd", 15),  # duplicate
    ]
    file_bytes = _build_wb_bytes(
        india_rows=india_rows,
        uae_rows=[_make_uae_row("ValidUaeClient LLC", 30)],
    )
    result = parse_credit_period_master(file_bytes)
    assert result.is_valid is False, "Duplicate client in IND sheet must make is_valid False"

    dup_errs = [e for e in result.errors if e.code == "DUPLICATE_CLIENT"]
    assert dup_errs, (
        f"Expected DUPLICATE_CLIENT in errors. " f"errors={[e.code for e in result.errors]}"
    )
    err = dup_errs[0]
    assert err.detail is not None
    # Spec §4.3 rule 5: detail must have "entity" key (not "sheet")
    assert (
        err.detail.get("entity") == "IND"
    ), f"detail['entity'] must be 'IND'; got {err.detail.get('entity')!r}"
    # detail['duplicates'] is a list of dicts {name: ..., row_indices: [...]}
    dup_names = [d["name"] for d in err.detail.get("duplicates", [])]
    assert "AlphaClient Ltd" in dup_names, (
        f"detail['duplicates'] must contain the duplicated name; " f"got names: {dup_names}"
    )


# ---------------------------------------------------------------------------
# Test 12: duplicate per-sheet only — same name in IND + UAE is valid
# ---------------------------------------------------------------------------


def test_same_client_name_in_both_sheets_is_valid() -> None:
    """'CommonClient' appearing in both India and UAE sheets is valid (no error).

    Duplicate detection is strictly within a single sheet.
    """
    shared_name = "CommonClient Ltd"
    file_bytes = _build_wb_bytes(
        india_rows=[
            _make_ind_row(shared_name, 30),
            _make_ind_row("OtherIndClient Ltd", 0),
        ],
        uae_rows=[
            _make_uae_row(shared_name, 15, "Same group"),
            _make_uae_row("OtherUaeClient LLC", 30),
        ],
    )
    result = parse_credit_period_master(file_bytes)
    assert result.is_valid is True, (
        f"Same client in both sheets must not be a duplicate. "
        f"errors={[e.code for e in result.errors]}"
    )
    dup_errs = [e for e in result.errors if e.code == "DUPLICATE_CLIENT"]
    assert not dup_errs, (
        f"Unexpected DUPLICATE_CLIENT error for cross-sheet name: "
        f"{[e.detail for e in dup_errs]}"
    )


# ---------------------------------------------------------------------------
# Test 13: missing India sheet → SHEET_NOT_FOUND in errors; UAE still processed
# ---------------------------------------------------------------------------


def test_missing_india_sheet_emits_sheet_not_found_and_uae_processed() -> None:
    """Workbook without India sheet → MISSING_SHEET error (spec §4.3 rule 1); UAE parsed.

    detail["missing"] must list ["India"].
    UAE rows must still be emitted (parser continues with present sheets).
    """
    file_bytes = _build_wb_bytes(
        india_rows=[],
        uae_rows=[_make_uae_row("UaeOnlyClient LLC", 30)],
        include_india=False,
        include_uae=True,
    )
    result = parse_credit_period_master(file_bytes)
    assert result.is_valid is False, "Missing India sheet must produce an error"

    sheet_errs = [e for e in result.errors if e.code == "MISSING_SHEET"]
    assert sheet_errs, f"Expected MISSING_SHEET error. errors={[e.code for e in result.errors]}"
    err = sheet_errs[0]
    assert err.detail is not None
    assert "India" in err.detail.get(
        "missing", []
    ), f"detail['missing'] must contain 'India'; got {err.detail.get('missing')!r}"

    # UAE sheet must still have been processed.
    uae_rows = [cp for cp in result.credit_periods if cp.entity_code == "UAE"]
    assert len(uae_rows) == 1, (
        f"UAE sheet must still be processed even when India is missing; "
        f"got {len(uae_rows)} UAE rows"
    )


# ---------------------------------------------------------------------------
# Test 14: missing Credit Period column in India → MISSING_REQUIRED_COLUMN
# ---------------------------------------------------------------------------


def test_missing_credit_period_column_emits_error() -> None:
    """India sheet without 'Credit Period' column header → MISSING_REQUIRED_COLUMN."""
    wb = openpyxl.Workbook()
    del wb["Sheet"]

    # India sheet with only Client Name (Credit Period column absent)
    ws_india = wb.create_sheet("India")
    ws_india.append(["Client Name"])  # no Credit Period col
    ws_india.append(["SomeClient Ltd"])

    ws_uae = wb.create_sheet("UAE")
    ws_uae.append(_UAE_HEADER)
    ws_uae.append(_make_uae_row("ValidUaeClient LLC", 30))

    buf = io.BytesIO()
    wb.save(buf)
    file_bytes = buf.getvalue()

    result = parse_credit_period_master(file_bytes)
    assert result.is_valid is False, "Missing Credit Period column must make is_valid False"
    missing_errs = [e for e in result.errors if e.code == "MISSING_REQUIRED_COLUMN"]
    assert missing_errs, (
        f"Expected MISSING_REQUIRED_COLUMN error. " f"errors={[e.code for e in result.errors]}"
    )


# ---------------------------------------------------------------------------
# Test 15: deterministic file_sha256 across two parse calls
# ---------------------------------------------------------------------------


def test_deterministic_file_sha256(simple_cp_bytes: bytes) -> None:
    r1 = parse_credit_period_master(simple_cp_bytes)
    r2 = parse_credit_period_master(simple_cp_bytes)
    assert (
        r1.file_sha256 == r2.file_sha256
    ), "Two parse calls on the same bytes must produce the same file_sha256"


# ---------------------------------------------------------------------------
# Test 16: duplicate client names in UAE sheet → DUPLICATE_CLIENT, entity="UAE"
# ---------------------------------------------------------------------------


def test_duplicate_client_names_in_uae_sheet_fails_parse() -> None:
    """'AlphaUae LLC' appears twice in UAE sheet → DUPLICATE_CLIENT with entity='UAE'.

    Spec §4.3 rule 5: detail["entity"] == "UAE", is_valid False.
    """
    uae_rows: list[list[Any]] = [
        _make_uae_row("AlphaUae LLC", 30, "Extended contract"),
        _make_uae_row("BetaUae LLC", 0),
        _make_uae_row("AlphaUae LLC", 15),  # duplicate
    ]
    file_bytes = _build_wb_bytes(
        india_rows=[_make_ind_row("ValidIndClient Ltd", 30)],
        uae_rows=uae_rows,
    )
    result = parse_credit_period_master(file_bytes)
    assert result.is_valid is False, "Duplicate client in UAE sheet must make is_valid False"

    dup_errs = [e for e in result.errors if e.code == "DUPLICATE_CLIENT"]
    assert dup_errs, (
        f"Expected DUPLICATE_CLIENT in errors. " f"errors={[e.code for e in result.errors]}"
    )
    uae_dup_errs = [e for e in dup_errs if e.detail is not None and e.detail.get("entity") == "UAE"]
    assert uae_dup_errs, (
        f"Expected DUPLICATE_CLIENT with entity='UAE'. "
        f"dup errors detail: {[e.detail for e in dup_errs]}"
    )
    err = uae_dup_errs[0]
    assert err.detail is not None
    dup_names = [d["name"] for d in err.detail.get("duplicates", [])]
    assert (
        "AlphaUae LLC" in dup_names
    ), f"detail['duplicates'] must contain the duplicated name; got {dup_names}"


# ---------------------------------------------------------------------------
# Test 17: missing UAE sheet → MISSING_SHEET, detail["missing"]==["UAE"], India processed
# ---------------------------------------------------------------------------


def test_missing_uae_sheet_emits_missing_sheet_and_india_processed() -> None:
    """Workbook without UAE sheet → MISSING_SHEET error (spec §4.3 rule 1).

    detail["missing"] must contain "UAE".
    India rows must still be emitted (parser continues with present sheets).
    """
    file_bytes = _build_wb_bytes(
        india_rows=[_make_ind_row("IndOnlyClient Ltd", 30)],
        uae_rows=[],
        include_india=True,
        include_uae=False,
    )
    result = parse_credit_period_master(file_bytes)
    assert result.is_valid is False, "Missing UAE sheet must produce an error"

    sheet_errs = [e for e in result.errors if e.code == "MISSING_SHEET"]
    assert sheet_errs, f"Expected MISSING_SHEET error. errors={[e.code for e in result.errors]}"
    err = sheet_errs[0]
    assert err.detail is not None
    assert "UAE" in err.detail.get(
        "missing", []
    ), f"detail['missing'] must contain 'UAE'; got {err.detail.get('missing')!r}"

    # India sheet must still have been processed.
    ind_rows = [cp for cp in result.credit_periods if cp.entity_code == "IND"]
    assert len(ind_rows) == 1, (
        f"India sheet must still be processed even when UAE is missing; "
        f"got {len(ind_rows)} IND rows"
    )


# ---------------------------------------------------------------------------
# Test 18: real fixture — emitted count per entity (skip if fixture absent)
# ---------------------------------------------------------------------------


def test_real_fixture_emitted_count_per_entity(credit_period_file_bytes: bytes) -> None:
    """Pin the exact number of credit_periods emitted per entity on the real fixture.

    Fixture cleansed 2026-04-19 per policy:
      - Identical duplicates: kept first occurrence (Fraction AI, Ziffy).
      - Conflicting duplicates: all rows for that client deleted (MakeMyTrip
        comma-variants, Sanjay Electricals, Peak Tourism with mixed values +
        blanks). Client falls to entity default via D8.
      - Blank credit_period rows: deleted (META ORYX, Vitamin Tea, Grootan,
        Intech SG).
      - MakeMyTrip row 146 (clean name, no trailing comma) preserved — parser
        treats it as a distinct client.

    Post-clean counts (measured against parser output):
      India: 178 rows emitted, 0 parse errors.
      UAE:    76 rows emitted, 0 parse errors.
    """
    result = parse_credit_period_master(credit_period_file_bytes)
    ind_count = sum(1 for cp in result.credit_periods if cp.entity_code == "IND")
    uae_count = sum(1 for cp in result.credit_periods if cp.entity_code == "UAE")
    assert result.is_valid, f"Expected is_valid=True on cleansed fixture; errors={result.errors}"
    assert ind_count == 178, (
        f"Expected 178 IND rows on cleansed fixture; got {ind_count}. "
        "If fixture was edited again, re-measure and update."
    )
    assert uae_count == 76, (
        f"Expected 76 UAE rows on cleansed fixture; got {uae_count}. "
        "If fixture was edited again, re-measure and update."
    )


# ---------------------------------------------------------------------------
# Test 19: unparseable credit_days row doesn't block other rows in same sheet
# ---------------------------------------------------------------------------


def test_unparseable_credit_days_row_does_not_block_valid_rows() -> None:
    """A row with credit_days='thirty' emits UNPARSEABLE_CREDIT_DAYS error.

    The bad row is NOT emitted.  Valid rows before AND after it ARE emitted.
    is_valid is False (error present).
    """
    india_rows: list[list[Any]] = [
        _make_ind_row("ValidBeforeClient Ltd", 30),  # row 1 — OK
        _make_ind_row("BadValueClient Ltd", "thirty"),  # row 2 — UNPARSEABLE
        _make_ind_row("ValidAfterClient Ltd", 15),  # row 3 — OK
    ]
    file_bytes = _build_wb_bytes(
        india_rows=india_rows,
        uae_rows=[_make_uae_row("ValidUaeClient LLC", 0)],
    )
    result = parse_credit_period_master(file_bytes)
    assert result.is_valid is False, "UNPARSEABLE_CREDIT_DAYS error must make is_valid False"

    unparseable_errs = [e for e in result.errors if e.code == "UNPARSEABLE_CREDIT_DAYS"]
    assert unparseable_errs, (
        f"Expected UNPARSEABLE_CREDIT_DAYS in errors. " f"errors={[e.code for e in result.errors]}"
    )
    # Bad row not emitted; valid rows are.
    ind_rows = [cp for cp in result.credit_periods if cp.entity_code == "IND"]
    assert len(ind_rows) == 2, f"Expected 2 IND rows (1 bad row skipped); got {len(ind_rows)}"
    ind_days = {cp.credit_days for cp in ind_rows}
    assert (
        30 in ind_days and 15 in ind_days
    ), f"Expected credit_days {{30, 15}} in emitted rows; got {ind_days}"


# ---------------------------------------------------------------------------
# Test 20: populated name + empty credit_days → UNPARSEABLE_CREDIT_DAYS (FIX-6)
# ---------------------------------------------------------------------------


def test_empty_credit_days_with_populated_name_emits_unparseable_error() -> None:
    """Distinguish 'empty name → skip' from 'populated name + empty credit_days → error'.

    Spec §4.3: empty Client Name rows are SKIPPED silently (cosmetic separators).
    But a row with a populated Client Name and empty credit_days value is a
    data-quality issue the analyst must see — emit UNPARSEABLE_CREDIT_DAYS.

    India sheet:
      Row 1: name="ClientAlpha", credit_days=30     (valid → emitted)
      Row 2: name="ClientBeta",  credit_days=None   (populated name + empty days → error)
    UAE sheet: minimal valid (so MISSING_SHEET doesn't fire).
    """
    india_rows: list[list[Any]] = [
        _make_ind_row("ClientAlpha", 30),
        _make_ind_row("ClientBeta", None),  # populated name, empty credit_days → UNPARSEABLE
    ]
    file_bytes = _build_wb_bytes(
        india_rows=india_rows,
        uae_rows=[_make_uae_row("ValidUaeClient LLC", 30)],
    )
    result = parse_credit_period_master(file_bytes)

    # is_valid must be False (one error present).
    assert result.is_valid is False, (
        "A populated name with empty credit_days must produce an error and set is_valid=False. "
        f"errors={[e.code for e in result.errors]}"
    )

    # Exactly one UNPARSEABLE_CREDIT_DAYS error expected (from ClientBeta row).
    unparseable = [e for e in result.errors if e.code == "UNPARSEABLE_CREDIT_DAYS"]
    assert len(unparseable) == 1, (
        f"Expected exactly 1 UNPARSEABLE_CREDIT_DAYS error; got {len(unparseable)}. "
        f"All errors: {[e.code for e in result.errors]}"
    )

    # The error row_index must be a real row index (not the -1 file-level sentinel).
    err = unparseable[0]
    assert err.row_index > 0, (
        f"UNPARSEABLE_CREDIT_DAYS.row_index must be the actual data row, not -1; "
        f"got row_index={err.row_index}"
    )

    # The valid row (ClientAlpha, credit_days=30) must be emitted.
    ind_rows = [cp for cp in result.credit_periods if cp.entity_code == "IND"]
    assert len(ind_rows) == 1, f"Expected 1 IND row (ClientAlpha); got {len(ind_rows)}"
    assert (
        ind_rows[0].credit_days == 30
    ), f"ClientAlpha must have credit_days=30; got {ind_rows[0].credit_days}"

    # The bad row (ClientBeta) must NOT be emitted as a StagedCreditPeriod.
    assert not any(
        cp.credit_days == 0 and cp.entity_code == "IND" for cp in result.credit_periods
    ), "ClientBeta (empty credit_days) must not appear in credit_periods"
    # Structural check: only 1 IND row (already asserted above), so ClientBeta isn't in output.
