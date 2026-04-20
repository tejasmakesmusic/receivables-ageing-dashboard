"""Unit tests for app.services.snapshot_service (M3 Task 2).

Uses MagicMock to stub the DB session and parser calls.  No live DB needed.
Tests cover the orchestration logic in upload_snapshot().
"""

from __future__ import annotations

import io
import uuid
from datetime import date
from decimal import Decimal
from typing import Any, cast
from unittest.mock import MagicMock, patch

import openpyxl
import pytest
from fastapi import HTTPException

from app.core.rbac import Role
from app.db.models.entity import Entity
from app.db.models.snapshot import Snapshot
from app.db.models.user import User
from app.parsers.common import ParseError, ParseResult, StagedInvoice
from app.services.snapshot_service import upload_snapshot

# ---------------------------------------------------------------------------
# Builders — use proper ORM __init__ so InstrumentedAttribute works.
# ---------------------------------------------------------------------------


def _make_entity(code: str = "IND") -> Entity:
    """Return an Entity instance (no session needed for attribute access)."""
    e = Entity(code=code, name="Test Entity", country="IN", base_currency="INR")
    e.id = uuid.uuid4()
    return e


def _make_user(
    role: Role = Role.ANALYST,
    entity_id_scope: uuid.UUID | None = None,
) -> User:
    """Return a User instance (no session needed for attribute access)."""
    u = User(email=f"test_{uuid.uuid4().hex[:6]}@example.com")
    u.id = uuid.uuid4()
    u.role = role
    u.entity_id_scope = entity_id_scope
    u.is_active = True
    return u


def _make_parse_result(
    source: str = "TALLY",
    has_errors: bool = False,
    has_warnings: bool = False,
    invoice_count: int = 2,
) -> ParseResult:
    """Build a minimal ParseResult for testing."""
    invoices = [
        StagedInvoice(
            row_index=i,
            source_currency="INR",
            party_name_raw=f"Party{i}",
            invoice_ref=f"INV-00{i}",
            invoice_date=date(2026, 1, 15),
            amount=Decimal("1000.00"),
            raw_row_json={"row": i},
        )
        for i in range(invoice_count)
    ]
    errors = (
        [ParseError(row_index=-1, code="SHEET_NOT_FOUND", message="Sheet missing")]
        if has_errors
        else []
    )
    warnings_list = (
        [ParseError(row_index=-1, code="GRAND_TOTAL_MISMATCH", message="Total off by 1")]
        if has_warnings
        else []
    )
    return ParseResult(
        invoices=invoices,
        errors=errors,
        warnings=warnings_list,
        as_of_date=date(2026, 1, 31) if source == "XERO" else None,
        file_sha256="a" * 64,
        source_hint=source,
    )


def _make_tally_xlsx() -> bytes:
    """Minimal Tally-shaped XLSX (sheet name only; parser not called in mocked tests)."""
    wb = openpyxl.Workbook()
    wb.active.title = "Sundry Debtors"  # openpyxl: active is non-None on a fresh workbook
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_xero_xlsx(as_of_str: str | None = "As at 31 March 2026") -> bytes:
    """Minimal Xero-shaped XLSX."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aged Receivables Detail"  # openpyxl: active is non-None on a fresh workbook
    ws.append(["Aged Receivables Detail"])  # row 1
    ws.append(["Test Company"])  # row 2
    ws.append([as_of_str])  # row 3 — sniffed
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_cp_xlsx() -> bytes:
    """Minimal Credit Period XLSX."""
    wb = openpyxl.Workbook()
    wb.active.title = "India"  # openpyxl: active is non-None on a fresh workbook
    wb.create_sheet("UAE")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Shared mock factory
# ---------------------------------------------------------------------------


def _mock_db(
    entity: Any = None,
    existing_snapshot: Any = None,
) -> MagicMock:
    """Return a mock Session pre-configured with scalar() side effects."""
    db = MagicMock()

    def _scalar(stmt: Any) -> Any:
        # Heuristic: detect which query is being run by inspecting the statement.
        stmt_str = str(stmt)
        if "entities" in stmt_str.lower():
            return entity
        if "snapshots" in stmt_str.lower():
            return existing_snapshot
        return None

    db.scalar.side_effect = _scalar

    # Track objects added via db.add() so flush can assign IDs.
    _added_objects: list[Any] = []

    def _add(obj: Any) -> None:
        _added_objects.append(obj)

    db.add.side_effect = _add

    def _flush() -> None:
        # Assign uuid.uuid4() to any added ORM object that lacks an id,
        # mirroring what a real flush+server-default would do.
        for obj in _added_objects:
            if getattr(obj, "id", None) is None:
                obj.id = uuid.uuid4()

    db.flush.side_effect = _flush
    db.commit.return_value = None
    db.refresh.return_value = None
    return db


# ---------------------------------------------------------------------------
# Tests: happy paths
# ---------------------------------------------------------------------------


class TestUploadSnapshotTally:
    @patch("app.services.snapshot_service.invoice_snapshots_has_partition_for", return_value=True)
    @patch("app.services.snapshot_service.parse_tally_grpbills")
    def test_tally_happy_path(
        self,
        mock_parse: MagicMock,
        mock_partition: MagicMock,
    ) -> None:
        entity = _make_entity("IND")
        user = _make_user(Role.ADMIN)
        db = _mock_db(entity=entity)
        mock_parse.return_value = _make_parse_result("TALLY")

        file_bytes = _make_tally_xlsx()
        resp = upload_snapshot(
            db=db,
            file_bytes=file_bytes,
            entity_code="IND",
            source_hint_form="TALLY",
            as_of_date_form=date(2026, 1, 31),
            current_user=user,
            request_ip="127.0.0.1",
        )

        assert resp.status == "STAGED"
        assert resp.source_hint == "TALLY"
        assert resp.as_of_date == date(2026, 1, 31)
        assert resp.parse_summary.invoices_parsed == 2
        assert resp.parse_summary.parse_error_count == 0
        mock_partition.assert_called_once()
        mock_parse.assert_called_once()


class TestUploadSnapshotXero:
    @patch("app.services.snapshot_service.invoice_snapshots_has_partition_for", return_value=True)
    @patch("app.services.snapshot_service.parse_xero_aged_receivables")
    def test_xero_sniff_wins(
        self,
        mock_parse: MagicMock,
        mock_partition: MagicMock,
    ) -> None:
        """Xero: sniffed date from file takes precedence over form value."""
        entity = _make_entity("UAE")
        user = _make_user(Role.ADMIN)
        db = _mock_db(entity=entity)
        result = _make_parse_result("XERO")
        mock_parse.return_value = result

        file_bytes = _make_xero_xlsx("As at 31 March 2026")
        resp = upload_snapshot(
            db=db,
            file_bytes=file_bytes,
            entity_code="UAE",
            source_hint_form="XERO",
            as_of_date_form=date(2026, 4, 1),  # form differs from sniffed
            current_user=user,
            request_ip="127.0.0.1",
        )
        # Sniffed wins (31 March 2026 = date(2026, 3, 31))
        assert resp.as_of_date == date(2026, 3, 31)

    @patch("app.services.snapshot_service.invoice_snapshots_has_partition_for", return_value=True)
    @patch("app.services.snapshot_service.parse_xero_aged_receivables")
    def test_xero_fallback_to_form(
        self,
        mock_parse: MagicMock,
        mock_partition: MagicMock,
    ) -> None:
        """Xero: falls back to form value when sniff fails."""
        entity = _make_entity("UAE")
        user = _make_user(Role.ADMIN)
        db = _mock_db(entity=entity)
        mock_parse.return_value = _make_parse_result("XERO")

        # No "As at" row in row 3
        file_bytes = _make_xero_xlsx(None)
        resp = upload_snapshot(
            db=db,
            file_bytes=file_bytes,
            entity_code="UAE",
            source_hint_form="XERO",
            as_of_date_form=date(2026, 3, 31),
            current_user=user,
            request_ip="127.0.0.1",
        )
        assert resp.as_of_date == date(2026, 3, 31)


class TestUploadSnapshotCreditPeriod:
    @patch("app.services.snapshot_service.parse_credit_period_master")
    def test_credit_period_happy_path(self, mock_parse: MagicMock) -> None:
        entity = _make_entity("IND")
        user = _make_user(Role.ADMIN)
        db = _mock_db(entity=entity)
        result = ParseResult(
            invoices=[],
            credit_periods=[],
            errors=[],
            warnings=[],
            as_of_date=None,
            file_sha256="b" * 64,
            source_hint="CREDIT_PERIOD",
        )
        mock_parse.return_value = result

        file_bytes = _make_cp_xlsx()
        resp = upload_snapshot(
            db=db,
            file_bytes=file_bytes,
            entity_code="IND",
            source_hint_form="CREDIT_PERIOD",
            as_of_date_form=None,
            current_user=user,
            request_ip="127.0.0.1",
        )
        assert resp.as_of_date is None
        assert resp.source_hint == "CREDIT_PERIOD"


# ---------------------------------------------------------------------------
# Tests: error paths
# ---------------------------------------------------------------------------


class TestUploadSnapshotErrors:
    def test_unknown_entity_code_raises_400(self) -> None:
        db = _mock_db(entity=None)
        user = _make_user(Role.ADMIN)
        with pytest.raises(HTTPException) as exc_info:
            upload_snapshot(
                db=db,
                file_bytes=_make_tally_xlsx(),
                entity_code="XXX",
                source_hint_form="TALLY",
                as_of_date_form=date(2026, 1, 31),
                current_user=user,
                request_ip="127.0.0.1",
            )
        assert exc_info.value.status_code == 400

    def test_analyst_wrong_entity_raises_403(self) -> None:
        entity = _make_entity("IND")
        other_entity_id = uuid.uuid4()
        user = _make_user(Role.ANALYST, entity_id_scope=other_entity_id)
        db = _mock_db(entity=entity)
        with pytest.raises(HTTPException) as exc_info:
            upload_snapshot(
                db=db,
                file_bytes=_make_tally_xlsx(),
                entity_code="IND",
                source_hint_form="TALLY",
                as_of_date_form=date(2026, 1, 31),
                current_user=user,
                request_ip="127.0.0.1",
            )
        assert exc_info.value.status_code == 403

    def test_duplicate_sha256_raises_409(self) -> None:
        entity = _make_entity("IND")
        user = _make_user(Role.ADMIN)
        existing = Snapshot(
            entity_id=entity.id,
            uploaded_by=uuid.uuid4(),
            as_of_date=None,
            source_hint="TALLY",
            status="STAGED",
            upload_file_sha256="a" * 64,
        )
        existing.id = uuid.uuid4()

        db = _mock_db(entity=entity, existing_snapshot=existing)
        with pytest.raises(HTTPException) as exc_info:
            upload_snapshot(
                db=db,
                file_bytes=_make_tally_xlsx(),
                entity_code="IND",
                source_hint_form="TALLY",
                as_of_date_form=date(2026, 1, 31),
                current_user=user,
                request_ip="127.0.0.1",
            )
        assert exc_info.value.status_code == 409
        assert cast(dict[str, Any], exc_info.value.detail)["code"] == "DUPLICATE_FILE"

    def test_tally_missing_as_of_date_raises_422(self) -> None:
        entity = _make_entity("IND")
        user = _make_user(Role.ADMIN)
        db = _mock_db(entity=entity)
        with pytest.raises(HTTPException) as exc_info:
            upload_snapshot(
                db=db,
                file_bytes=_make_tally_xlsx(),
                entity_code="IND",
                source_hint_form="TALLY",
                as_of_date_form=None,  # missing!
                current_user=user,
                request_ip="127.0.0.1",
            )
        assert exc_info.value.status_code == 422
        assert cast(dict[str, Any], exc_info.value.detail)["code"] == "AS_OF_DATE_MISSING"

    def test_xero_both_absent_raises_422(self) -> None:
        entity = _make_entity("UAE")
        user = _make_user(Role.ADMIN)
        db = _mock_db(entity=entity)
        # Xero file with no "As at" row and no form value.
        with pytest.raises(HTTPException) as exc_info:
            upload_snapshot(
                db=db,
                file_bytes=_make_xero_xlsx(None),
                entity_code="UAE",
                source_hint_form="XERO",
                as_of_date_form=None,
                current_user=user,
                request_ip="127.0.0.1",
            )
        assert exc_info.value.status_code == 422
        assert cast(dict[str, Any], exc_info.value.detail)["code"] == "AS_OF_DATE_MISSING"

    @patch("app.services.snapshot_service.invoice_snapshots_has_partition_for", return_value=False)
    def test_missing_partition_raises_422(self, mock_partition: MagicMock) -> None:
        entity = _make_entity("IND")
        user = _make_user(Role.ADMIN)
        db = _mock_db(entity=entity)
        with pytest.raises(HTTPException) as exc_info:
            upload_snapshot(
                db=db,
                file_bytes=_make_tally_xlsx(),
                entity_code="IND",
                source_hint_form="TALLY",
                as_of_date_form=date(2025, 12, 31),  # no partition
                current_user=user,
                request_ip="127.0.0.1",
            )
        assert exc_info.value.status_code == 422
        assert cast(dict[str, Any], exc_info.value.detail)["code"] == "MISSING_PARTITION"

    @patch("app.services.snapshot_service.invoice_snapshots_has_partition_for", return_value=True)
    @patch("app.services.snapshot_service.parse_tally_grpbills")
    def test_file_level_parse_errors_raise_422(
        self, mock_parse: MagicMock, mock_partition: MagicMock
    ) -> None:
        """ParseResult with errors → 422, no snapshot created."""
        entity = _make_entity("IND")
        user = _make_user(Role.ADMIN)
        db = _mock_db(entity=entity)
        mock_parse.return_value = _make_parse_result("TALLY", has_errors=True)

        with pytest.raises(HTTPException) as exc_info:
            upload_snapshot(
                db=db,
                file_bytes=_make_tally_xlsx(),
                entity_code="IND",
                source_hint_form="TALLY",
                as_of_date_form=date(2026, 1, 31),
                current_user=user,
                request_ip="127.0.0.1",
            )
        assert exc_info.value.status_code == 422
        assert cast(dict[str, Any], exc_info.value.detail)["code"] == "PARSE_ERROR"
        # Verify snapshot was NOT added to DB.
        db.add.assert_not_called()

    @patch("app.services.snapshot_service.invoice_snapshots_has_partition_for", return_value=True)
    @patch("app.services.snapshot_service.parse_tally_grpbills")
    def test_warnings_do_not_block_staging(
        self, mock_parse: MagicMock, mock_partition: MagicMock
    ) -> None:
        """Warnings do NOT block snapshot creation (spec §4.4 / is_valid property)."""
        entity = _make_entity("IND")
        user = _make_user(Role.ADMIN)
        db = _mock_db(entity=entity)
        mock_parse.return_value = _make_parse_result("TALLY", has_warnings=True)

        resp = upload_snapshot(
            db=db,
            file_bytes=_make_tally_xlsx(),
            entity_code="IND",
            source_hint_form="TALLY",
            as_of_date_form=date(2026, 1, 31),
            current_user=user,
            request_ip="127.0.0.1",
        )
        assert resp.status == "STAGED"
        assert len(resp.parse_summary.warnings) == 1

    def test_ambiguous_source_raises_400(self) -> None:
        entity = _make_entity("IND")
        user = _make_user(Role.ADMIN)
        db = _mock_db(entity=entity)
        # File with both Tally + Xero sheets → ambiguous.
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sundry Debtors"  # openpyxl: active is non-None on a fresh workbook
        wb.create_sheet("Aged Receivables Detail")
        buf = io.BytesIO()
        wb.save(buf)
        ambiguous_bytes = buf.getvalue()

        with pytest.raises(HTTPException) as exc_info:
            upload_snapshot(
                db=db,
                file_bytes=ambiguous_bytes,
                entity_code="IND",
                source_hint_form=None,
                as_of_date_form=None,
                current_user=user,
                request_ip="127.0.0.1",
            )
        assert exc_info.value.status_code == 400

    def test_unknown_source_hint_raises_400(self) -> None:
        entity = _make_entity("IND")
        user = _make_user(Role.ADMIN)
        db = _mock_db(entity=entity)

        with pytest.raises(HTTPException) as exc_info:
            upload_snapshot(
                db=db,
                file_bytes=_make_tally_xlsx(),
                entity_code="IND",
                source_hint_form="BOGUS",
                as_of_date_form=None,
                current_user=user,
                request_ip="127.0.0.1",
            )
        assert exc_info.value.status_code == 400

    def test_admin_bypasses_entity_scope(self) -> None:
        """ADMIN user can upload to any entity regardless of entity_id_scope."""
        entity = _make_entity("UAE")
        user = _make_user(Role.ADMIN, entity_id_scope=uuid.uuid4())  # some other scope
        db = _mock_db(entity=entity)
        # This should NOT raise 403 — admin bypasses scope check.
        # It will raise at a later step (partition check or parser), which is fine.
        # We patch partition + parser to avoid needing the full chain.
        with (
            patch(
                "app.services.snapshot_service.invoice_snapshots_has_partition_for",
                return_value=True,
            ),
            patch("app.services.snapshot_service.parse_tally_grpbills") as mock_parse,
        ):
            mock_parse.return_value = _make_parse_result("TALLY")
            resp = upload_snapshot(
                db=db,
                file_bytes=_make_tally_xlsx(),
                entity_code="UAE",
                source_hint_form="TALLY",
                as_of_date_form=date(2026, 1, 31),
                current_user=user,
                request_ip="127.0.0.1",
            )
        assert resp.status == "STAGED"
